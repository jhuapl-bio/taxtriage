#!/usr/bin/env python3
"""
annotate_report.py
------------------
Match DIAMOND BLASTx tabular output (format 6) against an annotation
metadata file on the accession / source_id column, then produce:

  1. A merged TSV report with all DIAMOND columns plus matched metadata.
  2. An XLSX workbook with:
       - "Matched_Hits"   : full merged table
       - "Summary_Matrix"  : pivot of gene_name x classification with hit counts
       - "Metadata_Counts" : per-metadata-field value counts for matched entries
"""

import argparse
import sys

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── DIAMOND format-6 default column names ──────────────────────────────
BLAST6_COLS = [
    "qseqid",
    "sseqid",
    "pident",
    "length",
    "mismatch",
    "gapopen",
    "qstart",
    "qend",
    "sstart",
    "send",
    "evalue",
    "bitscore",
]


def parse_args():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--diamond", required=True, help="DIAMOND BLASTx tabular output (format 6)")
    ap.add_argument("--meta", required=True, help="Annotation metadata TSV (must have source_id column)")
    ap.add_argument("--prefix", default="annotate_report", help="Output file prefix")
    ap.add_argument("--evalue-cutoff", type=float, default=1e-5, help="E-value cutoff for filtering hits")
    ap.add_argument("--pident", type=float, default=90, help="Minimum percent identity (0-100 scale) for filtering hits")
    return ap.parse_args()


def load_diamond(path):
    df = pd.read_csv(path, sep="\t", header=None, names=BLAST6_COLS, comment="#")
    df["evalue"] = pd.to_numeric(df["evalue"], errors="coerce")
    df["bitscore"] = pd.to_numeric(df["bitscore"], errors="coerce")
    df["pident"] = pd.to_numeric(df["pident"], errors="coerce")
    return df


def load_metadata(path):
    df = pd.read_csv(path, sep="\t", dtype=str)
    df.columns = df.columns.str.strip()
    if "source_id" not in df.columns:
        sys.exit("ERROR: metadata file must contain a 'source_id' column")
    return df


def format_xlsx(wb):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.auto_filter.ref = ws.dimensions

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row + 1, 102))),
                default=8,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"


def main():
    args = parse_args()

    diamond_df = load_diamond(args.diamond)
    meta_df = load_metadata(args.meta)
    # drop the sequence_protein column if it exists since it's not needed and can cause confusion with sseqid
    if "sequence_protein" in meta_df.columns:
        meta_df = meta_df.drop(columns=["sequence_protein"])

    # Filter by e-value
    diamond_df = diamond_df[diamond_df["evalue"] <= args.evalue_cutoff].copy()
    diamond_df = diamond_df[diamond_df["pident"] >= args.pident].copy()
    # remove the sequence_protein columns

    if diamond_df.empty:
        print("WARNING: No DIAMOND hits passed the e-value cutoff. Writing empty report.")

    # ── Merge on accession ──────────────────────────────────────────────
    merged = diamond_df.merge(meta_df, left_on="sseqid", right_on="source_id", how="inner")

    # ── TSV report ──────────────────────────────────────────────────────
    tsv_path = f"{args.prefix}.annotate_report.tsv"
    merged.to_csv(tsv_path, sep="\t", index=False)

    # ── XLSX workbook ───────────────────────────────────────────────────
    xlsx_path = f"{args.prefix}.annotate_report.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # Sheet 1: Full matched hits
        if not merged.empty:
            merged.to_excel(writer, sheet_name="Matched_Hits", index=False)
        else:
            pd.DataFrame(columns=BLAST6_COLS + list(meta_df.columns)).to_excel(
                writer, sheet_name="Matched_Hits", index=False
            )

        # Sheet 2: Summary matrix — gene_name x classification, values = count
        if not merged.empty and "gene_name" in merged.columns and "classification" in merged.columns:
            pivot = merged.pivot_table(
                index="gene_name",
                columns="classification",
                values="qseqid",
                aggfunc="count",
                fill_value=0,
            )
            pivot.to_excel(writer, sheet_name="Summary_Matrix")
        else:
            pd.DataFrame({"info": ["No gene_name/classification matches found"]}).to_excel(
                writer, sheet_name="Summary_Matrix", index=False
            )

        # Sheet 3: Metadata counts — value counts for key annotation columns
        count_cols = [
            "property", "source", "gene_name", "genus", "species",
            "organism", "product", "classification", "antibiotics_class",
            "antibiotics", "level", "host_name",
        ]
        available = [c for c in count_cols if c in merged.columns]
        if not merged.empty and available:
            counts_rows = []
            for col in available:
                vc = merged[col].value_counts()
                for val, cnt in vc.items():
                    counts_rows.append({"field": col, "value": val, "count": cnt})
            counts_df = pd.DataFrame(counts_rows)
            counts_df.to_excel(writer, sheet_name="Metadata_Counts", index=False)
        else:
            pd.DataFrame({"info": ["No metadata matches to count"]}).to_excel(
                writer, sheet_name="Metadata_Counts", index=False
            )

        # Sheet 4: Per-contig best hit summary
        if not merged.empty:
            best = merged.sort_values("bitscore", ascending=False).drop_duplicates(subset=["qseqid"], keep="first")
            summary_cols = ["qseqid", "sseqid", "pident", "evalue", "bitscore"]
            meta_summary = ["gene_name", "organism", "product", "classification", "antibiotics_class"]
            summary_cols += [c for c in meta_summary if c in best.columns]
            best[summary_cols].to_excel(writer, sheet_name="Best_Hits_Per_Contig", index=False)

        format_xlsx(writer.book)

    print(f"Wrote {tsv_path}")
    print(f"Wrote {xlsx_path}")
    print(f"Total DIAMOND hits (post-filter): {len(diamond_df)}")
    print(f"Matched to metadata: {len(merged)}")
    if not merged.empty:
        print(f"Unique accessions matched: {merged['sseqid'].nunique()}")
        print(f"Unique contigs with hits: {merged['qseqid'].nunique()}")


if __name__ == "__main__":
    main()
