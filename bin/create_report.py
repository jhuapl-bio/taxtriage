#!/usr/bin/env python3

# ---------------------------------------------------------------------------
# FIPS-safe environment setup (must run before ANY imports that touch OpenSSL)
# On HPC systems with FIPS mode enforced at the kernel level, the OpenSSL
# library inside containers may fail the FIPS self-test and SIGABRT the
# process.  Setting OPENSSL_CONF to /dev/null skips the FIPS provider
# entirely.  We also ensure matplotlib and fontconfig have writable cache
# directories so they don't emit warnings on read-only home filesystems.
# ---------------------------------------------------------------------------
import os as _os
import tempfile as _tempfile
# import FONT
import matplotlib.font_manager as _font_manager

if not _os.environ.get("OPENSSL_CONF"):
    _os.environ["OPENSSL_CONF"] = "/dev/null"

if not _os.environ.get("MPLCONFIGDIR"):
    _mpl_tmp = _os.path.join(_tempfile.gettempdir(), "matplotlib_cache")
    _os.makedirs(_mpl_tmp, exist_ok=True)
    _os.environ["MPLCONFIGDIR"] = _mpl_tmp

if not _os.environ.get("FONTCONFIG_PATH"):
    _os.environ["FONTCONFIG_PATH"] = _tempfile.gettempdir()

if not _os.environ.get("XDG_CACHE_HOME"):
    _os.environ["XDG_CACHE_HOME"] = _tempfile.gettempdir()
# ---------------------------------------------------------------------------

import io
import json
import argparse
import os
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import numpy as np
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.platypus.flowables import AnchorFlowable, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
from map_taxid import load_taxdump, load_names, load_merged
from body_site_normalization import normalize_body_site

# ── Unicode font registration ──────────────────────────────────────────────
# Built-in PDF fonts (Helvetica, etc.) only cover WinAnsi/Latin-1.
# Symbols such as ⚠ (U+26A0) render as blank blocks without a TTF font.
from reportlab.pdfbase import pdfmetrics as _rl_pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont as _rl_TTFont

_UNICODE_FONT = None


# HTML snippet for a red ⚠ warning symbol in Paragraph markup.
# Uses the registered Unicode font when available; falls back to bold ASCII.
_WARN_SYMBOL_HTML = (
    f'<font name="{_UNICODE_FONT}" color="#E85F50">&#9888;</font>'
    if _UNICODE_FONT
    else '<font color="#E85F50"><b>(!)</b></font>'
)


def get_all_strains(species_group):
    """Extract all strain-level members from a species group.

    Handles both the new 3-level hierarchy (toplevelkey → subkey groups → strains)
    and the legacy flat hierarchy (toplevelkey → strains) for backwards compatibility.

    Returns a flat list of strain dicts.
    """
    strains = []
    for member in species_group.get('members', []):
        if 'members' in member and member['members']:
            # New structure: member is a subkey group with nested strains
            strains.extend(member['members'])
        else:
            # Legacy flat structure: member is a strain directly
            strains.append(member)
    return strains


def get_subkey_groups(species_group):
    """Extract subkey-level groups from a species group.

    Returns the members list directly (each element is a subkey group with
    its own 'members' list of strains).  For legacy flat data, wraps strains
    into synthetic subkey groups.
    """
    members = species_group.get('members', [])
    if not members:
        return []
    # Check if this is the new nested structure
    if any('members' in m and m['members'] for m in members):
        return members
    # Legacy: wrap flat strains into subkey groups
    from collections import defaultdict
    by_subkey = defaultdict(list)
    for s in members:
        sk = str(s.get('subkey', s.get('key', 'unknown')))
        by_subkey[sk].append(s)
    groups = []
    for sk, strain_list in by_subkey.items():
        best = max(strain_list, key=lambda x: x.get('tass_score', 0))
        groups.append({
            'key': sk,
            'subkey': sk,
            'subkeyname': best.get('subkeyname', best.get('name', '')),
            'toplevelkey': best.get('toplevelkey', ''),
            'toplevelname': best.get('toplevelname', ''),
            'members': strain_list,
        })
    return groups


def get_qualifying_strains(species_group, args, min_conf, min_reads=1):
    """Return strains that qualify for display, including subkey-promoted strains.

    A strain qualifies if:
      1. It passes the confidence threshold on its own, OR
      2. Its parent subkey group passes the confidence threshold AND the strain
         belongs to an includable microbial category.

    This ensures that species-level aggregations (which may have a higher TASS
    than any single strain) can promote their strains into the report.
    """
    qualified = []
    seen = set()

    # Pass 1: strains that qualify on their own merit
    for strain in get_all_strains(species_group):
        if should_include_strain(strain, args) and has_min_reads(strain, min_reads):
            if passes_confidence_threshold(strain, min_conf):
                qualified.append(strain)
                seen.add(id(strain))

    # Pass 2: subkey promotion — include strains whose parent subkey group
    # passes the threshold even if the individual strain doesn't
    for sk_grp in get_subkey_groups(species_group):
        if not passes_confidence_threshold(sk_grp, min_conf):
            continue
        if not should_include_strain(sk_grp, args):
            continue
        for strain in sk_grp.get('members', []):
            if id(strain) in seen:
                continue
            if should_include_strain(strain, args) and has_min_reads(strain, min_reads):
                qualified.append(strain)
                seen.add(id(strain))

    return qualified


def strain_passes_with_subkey_promotion(strain, species_group, min_conf):
    """Check if a strain passes the confidence threshold, either directly
    or via its parent subkey group's TASS score."""
    if passes_confidence_threshold(strain, min_conf):
        return True
    # Check if the parent subkey group passes
    sk = str(strain.get('subkey', strain.get('key', '')))
    for sk_grp in get_subkey_groups(species_group):
        grp_sk = str(sk_grp.get('subkey', sk_grp.get('key', '')))
        if grp_sk == sk and passes_confidence_threshold(sk_grp, min_conf):
            return True
    return False


def get_direct_qualifying_strains(species_group, args, min_conf, min_reads=1):
    """Return only strain members that pass the confidence threshold directly."""
    qualified = []
    for strain in get_all_strains(species_group):
        if not should_include_strain(strain, args):
            continue
        if not has_min_reads(strain, min_reads):
            continue
        if passes_confidence_threshold(strain, min_conf):
            qualified.append(strain)
    return qualified


def _category_severity(microbial_category):
    category = str(microbial_category or 'Unknown')
    if 'Primary' in category:
        return 4
    if 'Opportunistic' in category:
        return 3
    if 'Potential' in category:
        return 2
    if 'Unknown' in category:
        return 1
    if 'Commensal' in category:
        return 0
    return 1


def resolve_rollup_annotation(species_record=None, qualifying_strains=None):
    """Resolve the display annotation for a species/subkey summary row."""
    candidates = []
    if species_record:
        candidates.append(species_record)
    candidates.extend(qualifying_strains or [])
    if not candidates:
        return {
            'microbial_category': 'Unknown',
            'annClass': '',
            'high_cons': False,
        }

    best = max(
        candidates,
        key=lambda rec: (
            _category_severity(rec.get('microbial_category', 'Unknown')),
            float(rec.get('tass_score', 0) or 0),
            float(rec.get('numreads', 0) or 0),
        ),
    )
    return {
        'microbial_category': best.get('microbial_category', 'Unknown'),
        'annClass': best.get('annClass', ''),
        'high_cons': any(rec.get('high_cons', False) for rec in candidates),
    }


def build_subkey_display_lookup(species_group, args, min_conf, min_reads=1,
                                max_members=None):
    """Build visible species/subkey rows plus directly qualifying child strains."""
    lookup = {}
    for sk_grp in get_subkey_groups(species_group):
        sk = str(sk_grp.get('subkey', sk_grp.get('key', 'unknown')))
        species_visible = (
            should_include_strain(sk_grp, args)
            and has_min_reads(sk_grp, min_reads)
            and passes_confidence_threshold(sk_grp, min_conf)
        )
        visible_strains = [
            strain for strain in get_direct_qualifying_strains(sk_grp, args, min_conf, min_reads=min_reads)
            if str(strain.get('key', '')) != sk
        ]
        visible_strains.sort(key=lambda s: s.get('tass_score', 0), reverse=True)
        if max_members is not None and max_members > 0:
            visible_strains = visible_strains[:max_members]

        if not species_visible and not visible_strains:
            continue

        resolved = resolve_rollup_annotation(
            sk_grp if species_visible else None,
            visible_strains,
        )
        display_name = (
            sk_grp.get('name')
            or sk_grp.get('subkeyname')
            or species_group.get('subkeyname')
            or species_group.get('name')
            or 'Unknown'
        )
        display_score = max(
            float(sk_grp.get('tass_score', 0) or 0) if species_visible else 0.0,
            max((float(s.get('tass_score', 0) or 0) for s in visible_strains), default=0.0),
        )
        # Strains that are Primary-category but did NOT pass the threshold.
        # These are hidden from the main table but indicated via a warning badge.
        # Only shown when no strain-level members are already visible above threshold.
        _visible_keys = {str(s.get('key', '')) for s in visible_strains}
        _visible_keys.add(sk)  # exclude the species-level key itself
        if visible_strains:
            # A qualifying strain is already shown — suppress the below-threshold badge
            below_threshold_primary = []
        else:
            below_threshold_primary = sorted(
                [
                    s for s in get_all_strains(sk_grp)
                    if 'Primary' in str(s.get('microbial_category', '') or '')
                    and not passes_confidence_threshold(s, min_conf)
                    and str(s.get('key', '')) not in _visible_keys
                ],
                key=lambda s: float(s.get('tass_score', 0) or 0),
                reverse=True,
            )

        lookup[sk] = {
            'species_key': str(species_group.get('toplevelkey', species_group.get('key', 'unknown'))),
            'subkey': sk,
            'species_record': sk_grp,
            'species_passes': species_visible,
            'display_name': display_name,
            'visible_strains': visible_strains,
            'display_score': display_score,
            'microbial_category': resolved['microbial_category'],
            'annClass': resolved['annClass'],
            'high_cons': resolved['high_cons'],
            'harmful_followup': has_harmful_followup_signal(visible_strains),
            'below_threshold_primary': below_threshold_primary,
        }

    return lookup


class _OutlineCollector:
    """Accumulates PDF outline entries and flushes them in sorted order.

    Entries are collected during the ReportLab build phase (as each
    ``OutlineDest`` flowable is drawn) and flushed once at the very end by
    ``FinalizeOutlines``.  Flushing sorts level-0 entries alphabetically and
    sorts level-1 children alphabetically within each parent group, so the
    PDF sidebar bookmark panel is alphabetically ordered.
    """

    def __init__(self):
        self.entries = []  # [(sort_key, dest_key, title, level, closed)]

    def add(self, dest_key, title, level=0, closed=True, sort_key=None):
        if sort_key is None:
            sort_key = title.lower()
        self.entries.append((sort_key, dest_key, title, level, closed))

    def flush_sorted(self, canv):
        """Sort entries respecting hierarchy and add to the canvas outline."""
        # Group level-1+ entries under their preceding level-0 entry
        groups = []  # [(level0_entry, [child_entries])]
        current_group = None
        for entry in self.entries:
            _sort_key, _dest_key, _title, level, _closed = entry
            if level == 0:
                current_group = (entry, [])
                groups.append(current_group)
            elif current_group is not None:
                current_group[1].append(entry)

        # Sort level-0 groups alphabetically
        groups.sort(key=lambda g: g[0][0])

        # Sort children within each group alphabetically
        for _parent, children in groups:
            children.sort(key=lambda e: e[0])

        # Flatten and add to the PDF outline
        for parent, children in groups:
            _, dest_key, title, level, closed = parent
            canv.addOutlineEntry(title, dest_key, level=level, closed=closed)
            for child in children:
                _, dest_key, title, level, closed = child
                canv.addOutlineEntry(title, dest_key, level=level, closed=closed)

        self.entries.clear()


class OutlineDest(Flowable):
    """Zero-size flowable that bookmarks the current Y position and registers
    an outline entry with the collector.

    Unlike the previous ``OutlineEntry``, this flowable uses
    ``bookmarkHorizontalAbsolute`` so the PDF viewer jumps to the *exact
    vertical position* on the page (not just the page top).  The actual
    ``addOutlineEntry`` call is deferred to ``FinalizeOutlines`` so all
    entries can be emitted in sorted order.
    """

    width = 0
    height = 0

    def __init__(self, key, title, level=0, closed=True, collector=None,
                 sort_key=None):
        super().__init__()
        self.key = key
        self.title = title
        self.level = level
        self.closed = closed
        self.collector = collector
        self.sort_key = sort_key

    def draw(self):
        self.canv.bookmarkHorizontalAbsolute(self.key, self.canv._y)
        if self.collector is not None:
            self.collector.add(
                self.key, self.title, self.level, self.closed, self.sort_key)


class AbsoluteAnchorFlowable(Flowable):
    """Zero-size flowable that registers a named destination using
    ``bookmarkHorizontalAbsolute`` so the bookmark resolves to the
    correct page position even when embedded inside a Table cell.

    ReportLab's built-in ``AnchorFlowable`` uses ``bookmarkHorizontal``
    which relies on the current transformation matrix.  Inside table cells
    the CTM is local to the cell, so bookmarks can point to wrong locations
    — especially when the table spans multiple pages.  This class avoids
    that problem by using the canvas's absolute Y directly.
    """

    width = 0
    height = 0

    def __init__(self, name):
        super().__init__()
        self._name = name

    def draw(self):
        self.canv.bookmarkHorizontalAbsolute(self._name, self.canv._y)


class FinalizeOutlines(Flowable):
    """Zero-size flowable placed at the very end of the story.

    When drawn it tells the collector to sort all accumulated entries
    alphabetically and write them to the PDF outline in one batch.
    """

    width = 0
    height = 0

    def __init__(self, collector):
        super().__init__()
        self.collector = collector

    def draw(self):
        self.collector.flush_sorted(self.canv)


class PageTracker(Flowable):
    """Zero-size flowable that records ``(page_index, y_position)`` when drawn.

    ``page_index`` is 0-based so it aligns with pikepdf's ``pdf.pages``
    indexing.  ``y_position`` is in PDF user-space points measured from the
    bottom of the page (standard PDF coordinate system).
    """

    width = 0
    height = 0

    def __init__(self, record_dict, key):
        super().__init__()
        self._record = record_dict
        self._key = key

    def draw(self):
        # ReportLab page numbers start at 1; convert to 0-based for pikepdf
        pg = self.canv.getPageNumber() - 1
        y = self.canv._y          # current Y in PDF points (0 = page bottom)
        self._record[self._key] = (pg, y)


class ControlSparkBar(Flowable):
    """Inline spark bar showing where a sample's TASS score falls relative
    to negative and positive control distributions.

    The bar is rendered as a compact horizontal strip:

    - Full bar represents the TASS range 0–1.
    - Red shaded zone: 0 → max(negative control TASS values).
    - Green shaded zone: min(positive control TASS) → 1.0  (if positives exist).
    - Small tick marks for each individual control value.
    - Black triangle marker for the sample's own TASS score.
    - If the sample falls within the negative (red) zone, the triangle is red.
    """

    def __init__(self, sample_tass, neg_values=None, pos_values=None,
                 control_flag=None, bar_width=72, bar_height=10,
                 tass_fold=None, reads_fold=None,
                 pos_tass_fold=None, pos_reads_fold=None,
                 insilico_values=None, insilico_tass_fold=None,
                 insilico_reads_fold=None, missing_from_insilico=False):
        super().__init__()
        self.sample_tass = float(sample_tass or 0)
        self.neg_values = neg_values or []
        self.pos_values = pos_values or []
        self.insilico_values = insilico_values or []
        self.control_flag = control_flag or "no_neg_controls"
        self.tass_fold = tass_fold          # fold-change over neg max TASS
        self.reads_fold = reads_fold        # fold-change over neg max reads
        self.pos_tass_fold = pos_tass_fold  # fold-change over pos min TASS
        self.pos_reads_fold = pos_reads_fold  # fold-change over pos min reads
        self.insilico_tass_fold = insilico_tass_fold
        self.insilico_reads_fold = insilico_reads_fold
        self.missing_from_insilico = missing_from_insilico
        self._row_h = 6                    # height per label row
        _has_neg_label = tass_fold is not None and bool(neg_values)
        _has_pos_label = pos_tass_fold is not None and bool(pos_values)
        _has_isil_label = (insilico_tass_fold is not None and bool(insilico_values)) or missing_from_insilico
        _label_rows = int(_has_neg_label) + int(_has_pos_label) + int(_has_isil_label)
        _label_rows = max(_label_rows, 1)  # always reserve at least 1 row
        self._label_h = self._row_h * _label_rows
        self._has_neg_label = _has_neg_label
        self._has_pos_label = _has_pos_label
        self._has_isil_label = _has_isil_label
        self.width = bar_width
        self.height = bar_height + self._label_h

    def draw(self):
        c = self.canv
        w = self.width
        label_h = self._label_h
        bar_h = self.height - label_h   # actual bar height (above the label)

        # The bar is drawn with its bottom edge at y=label_h so the fold
        # label fits underneath.
        by = label_h  # bar y-origin

        # ── Background bar ───────────────────────────────────────────────
        c.setFillColor(colors.Color(0.93, 0.93, 0.93, 1))
        c.rect(0, by, w, bar_h, fill=1, stroke=0)

        # ── Red zone (negative control range) ────────────────────────────
        neg_tass_vals = [float(v.get("tass_score", 0) or 0) for v in self.neg_values]
        if neg_tass_vals:
            neg_max = max(neg_tass_vals)
            red_w = neg_max * w
            if red_w > 0:
                c.setFillColor(colors.Color(0.95, 0.7, 0.7, 0.5))
                c.rect(0, by, red_w, bar_h, fill=1, stroke=0)
            # Tick marks for individual neg values
            c.setStrokeColor(colors.Color(0.8, 0.2, 0.2, 0.7))
            c.setLineWidth(0.5)
            for val in neg_tass_vals:
                x = val * w
                c.line(x, by, x, by + bar_h)

        # ── Green zone (positive control range) ──────────────────────────
        pos_tass_vals = [float(v.get("tass_score", 0) or 0) for v in self.pos_values]
        if pos_tass_vals:
            pos_min = min(pos_tass_vals)
            pos_max = max(pos_tass_vals)
            green_x = pos_min * w
            green_w = max(1, (pos_max - pos_min) * w)
            # Extend to right edge if pos_max is close to 1
            if pos_max >= 0.95:
                green_w = w - green_x
            c.setFillColor(colors.Color(0.7, 0.92, 0.7, 0.5))
            c.rect(green_x, by, green_w, bar_h, fill=1, stroke=0)
            # Tick marks for individual pos values
            c.setStrokeColor(colors.Color(0.2, 0.7, 0.2, 0.7))
            c.setLineWidth(0.5)
            for val in pos_tass_vals:
                x = val * w
                c.line(x, by, x, by + bar_h)

        # ── Teal zone (in-silico control range) ─────────────────────────
        isil_tass_vals = [float(v.get("tass_score", 0) or 0) for v in self.insilico_values]
        if isil_tass_vals:
            isil_min = min(isil_tass_vals)
            isil_max = max(isil_tass_vals)
            teal_x = isil_min * w
            teal_w = max(1, (isil_max - isil_min) * w)
            if isil_max >= 0.95:
                teal_w = w - teal_x
            # Light teal with transparency
            c.setFillColor(colors.Color(0.6, 0.88, 0.88, 0.45))
            c.rect(teal_x, by, teal_w, bar_h, fill=1, stroke=0)
            # Tick marks for individual insilico values
            c.setStrokeColor(colors.Color(0.0, 0.55, 0.55, 0.7))
            c.setLineWidth(0.5)
            for val in isil_tass_vals:
                x = val * w
                c.line(x, by, x, by + bar_h)

        # ── Sample marker (triangle) ─────────────────────────────────────
        sx = self.sample_tass * w
        # Clamp to bar bounds
        sx = max(2, min(w - 2, sx))
        tri_h = bar_h * 0.7
        tri_half = 3

        if self.control_flag == "within_negative":
            c.setFillColor(colors.Color(0.85, 0.15, 0.15, 1))
        else:
            c.setFillColor(colors.Color(0.1, 0.1, 0.1, 1))

        p = c.beginPath()
        p.moveTo(sx - tri_half, by)
        p.lineTo(sx + tri_half, by)
        p.lineTo(sx, by + tri_h)
        p.close()
        c.drawPath(p, fill=1, stroke=0)

        # ── Thin border ──────────────────────────────────────────────────
        c.setStrokeColor(colors.Color(0.6, 0.6, 0.6, 1))
        c.setLineWidth(0.3)
        c.rect(0, by, w, bar_h, fill=0, stroke=1)

        # ── Fold-change labels (tiny text below the bar) ─────────────────
        def _fold_str(v):
            """Format a fold value as string, handling inf and None."""
            if v is None:
                return None
            if v == float("inf"):
                return "\u221e\u00d7"   # ∞×
            return f"{v:.1f}\u00d7"

        c.setFont("Helvetica", 5)

        # Determine which row each label set occupies:
        # When both exist: neg on row 0 (bottom), pos on row 1 (above).
        # When only one exists: it goes on row 0.
        _current_row = 0

        # Negative control folds — red/dark-gray text, prefixed with "−"
        if self._has_neg_label and neg_tass_vals:
            row_y = _current_row * self._row_h + 0.5
            fold_txt = "\u2212 " + _fold_str(self.tass_fold)  # − prefix
            if self.control_flag == "within_negative":
                c.setFillColor(colors.Color(0.75, 0.1, 0.1, 1))
            else:
                c.setFillColor(colors.Color(0.35, 0.35, 0.35, 1))
            c.drawString(1, row_y, fold_txt)

            if self.reads_fold is not None:
                rd_txt = _fold_str(self.reads_fold)
                if rd_txt:
                    rd_txt = rd_txt.rstrip("\u00d7") + "\u00d7 rd"  # e.g. "28.5× rd"
                tass_txt_w = c.stringWidth(fold_txt, "Helvetica", 5)
                c.setFillColor(colors.Color(0.5, 0.5, 0.5, 1))
                c.drawString(tass_txt_w + 4, row_y, rd_txt)
            _current_row += 1

        # Positive control folds — green text, prefixed with "+"
        pos_tass_vals = [float(v.get("tass_score", 0) or 0) for v in self.pos_values]
        if self._has_pos_label and pos_tass_vals:
            row_y = _current_row * self._row_h + 0.5
            pos_fold_txt = "+ " + _fold_str(self.pos_tass_fold)  # + prefix
            c.setFillColor(colors.Color(0.1, 0.55, 0.1, 1))
            c.drawString(1, row_y, pos_fold_txt)

            if self.pos_reads_fold is not None:
                pos_rd_txt = _fold_str(self.pos_reads_fold)
                if pos_rd_txt:
                    pos_rd_txt = pos_rd_txt.rstrip("\u00d7") + "\u00d7 rd"
                pos_txt_w = c.stringWidth(pos_fold_txt, "Helvetica", 5)
                c.setFillColor(colors.Color(0.3, 0.6, 0.3, 1))
                c.drawString(pos_txt_w + 4, row_y, pos_rd_txt)
            _current_row += 1

        # In-silico control folds — teal text, prefixed with "∞"
        if self._has_isil_label:
            row_y = _current_row * self._row_h + 0.5
            if self.missing_from_insilico:
                c.setFillColor(colors.Color(0.0, 0.45, 0.45, 1))
                c.drawString(1, row_y, "\u221e not in sim")  # ∞ prefix
            elif isil_tass_vals:
                isil_fold_txt = "\u221e " + _fold_str(self.insilico_tass_fold)  # ∞ prefix
                c.setFillColor(colors.Color(0.0, 0.45, 0.45, 1))
                c.drawString(1, row_y, isil_fold_txt)

                if self.insilico_reads_fold is not None:
                    isil_rd_txt = _fold_str(self.insilico_reads_fold)
                    if isil_rd_txt:
                        isil_rd_txt = isil_rd_txt.rstrip("\u00d7") + "\u00d7 rd"
                    isil_txt_w = c.stringWidth(isil_fold_txt, "Helvetica", 5)
                    c.setFillColor(colors.Color(0.2, 0.55, 0.55, 1))
                    c.drawString(isil_txt_w + 4, row_y, isil_rd_txt)


class SubthresholdWarningFlowable(Flowable):
    """Compact amber badge showing the highest-scoring Primary strain that is
    below the TASS threshold.

    Displays: &#9830; {Name} (TASS: {score}, {reads} reads)
    """

    def __init__(self, below_threshold_strains, badge_height=11):
        super().__init__()
        # below_threshold_strains is already sorted descending by TASS; [0] is highest
        self.below_threshold_strains = below_threshold_strains
        self._badge_h = badge_height
        if below_threshold_strains:
            top = below_threshold_strains[0]
            name = top.get('name', 'Unknown')
            tass = float(top.get('tass_score', 0) or 0) * 100
            reads = int(top.get('numreads', 0) or 0)
            self._label = f'{name} (TASS: {tass:.1f}, {reads:,} reads) \u2014 highest below threshold'
        else:
            self._label = ''
        self.height = badge_height if below_threshold_strains else 0
        self.width = 0  # updated in wrap()
        self._avail_w = None

    def wrap(self, availWidth, availHeight):
        if not self.below_threshold_strains:
            self.width = 0
            self.height = 0
            return (0, 0)
        self._avail_w = availWidth
        self.width = availWidth
        return (availWidth, self._badge_h)

    def draw(self):
        canv = self.canv
        if not self.below_threshold_strains:
            return

        bh = self._badge_h
        font_size = 6.5
        icon_size = font_size + 1.0
        x_off = 4                  # left padding
        gap = 2                    # gap between icon and label text

        canv.saveState()
        canv.setFont('Helvetica-Bold', font_size)
        icon_w = canv.stringWidth('\u2757', 'Helvetica-Bold', icon_size)
        max_text_w = (self._avail_w or 9999) - x_off - icon_w - gap - x_off

        # Truncate label with ellipsis if it would overflow the available width
        label = self._label
        if canv.stringWidth(label, 'Helvetica-Bold', font_size) > max_text_w:
            while label and canv.stringWidth(label + '\u2026', 'Helvetica-Bold', font_size) > max_text_w:
                label = label[:-1]
            label = label.rstrip() + '\u2026'

        text_w = canv.stringWidth(label, 'Helvetica-Bold', font_size)
        bw = min(x_off + icon_w + gap + text_w + x_off, self._avail_w or 9999)

        # Amber rounded badge background
        canv.setFillColorRGB(0.85, 0.42, 0.02)
        canv.roundRect(0, 0, bw, bh, 2, fill=1, stroke=0)

        # White ⚠ icon
        canv.setFont('Helvetica-Bold', icon_size)
        canv.setFillColorRGB(1, 1, 1)
        canv.drawString(x_off, (bh - icon_size) / 2 + 0.5, '\u2757')

        # White label text after the icon
        canv.setFont('Helvetica-Bold', font_size)
        canv.drawString(x_off + icon_w + gap, (bh - font_size) / 2 + 0.5, label)

        canv.restoreState()


_STERILE_TYPES = {"sterile", "blood", "csf", "serum"}


def _commensal_site_tag(strain, sample_type):
    """Return an HTML tag like ' <font color="#e67e22">[skin flora]</font>'
    if the organism is a known commensal at a non-sterile site and the
    sample is from a sterile site.  Returns '' otherwise."""
    if not sample_type:
        return ''
    norm_st = sample_type.lower().strip()
    if norm_st not in _STERILE_TYPES:
        return ''
    sites = strain.get('commensal_sites', [])
    if not sites:
        return ''
    # Flatten lists (body_site_map can return lists) and drop blanks
    flat = []
    for s in sites:
        if isinstance(s, list):
            flat.extend(x for x in s if x)
        elif s:
            flat.append(s)
    if not flat:
        return ''
    label = ', '.join(sorted(set(flat)))
    return f' <font size="6" color="#e67e22">[{label} flora]</font>'


def _is_flora_on_sterile(strain, sample_type):
    """Return True if this organism has commensal flora sites (e.g. skin,
    nasal) AND the sample type is sterile (blood, csf, etc.).
    Used to slightly fade background colours for these rows."""
    if not sample_type:
        return False
    norm_st = sample_type.lower().strip()
    if norm_st not in _STERILE_TYPES:
        return False
    sites = strain.get('commensal_sites', [])
    if not sites:
        return False
    flat = []
    for s in sites:
        if isinstance(s, list):
            flat.extend(x for x in s if x)
        elif s:
            flat.append(s)
    return bool(flat)


def _has_control_data(species_groups):
    """Check if any organism in the dataset has control_comparison or insilico_comparison data."""
    for sg in species_groups:
        if sg.get('control_comparison') or sg.get('insilico_comparison'):
            return True
        for sk_m in sg.get('members', []):
            if sk_m.get('control_comparison') or sk_m.get('insilico_comparison'):
                return True
            for strain in sk_m.get('members', []):
                if strain.get('control_comparison') or strain.get('insilico_comparison'):
                    return True
    return False


def _build_spark_bar_from_member(member, bar_width=72, bar_height=10):
    """Build a ControlSparkBar from a member's control_comparison and/or insilico_comparison dict.

    Returns None if no control or insilico data is present.
    """
    cc = member.get('control_comparison')
    ic = member.get('insilico_comparison')
    if not cc and not ic:
        return None
    if not cc:
        cc = {}  # allow insilico-only rendering

    # Extract fold values — handle inf stored as string in JSON round-trips
    def _safe_fold(v):
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    # Extract insilico comparison data (if any)
    ic = member.get('insilico_comparison') or {}
    isil_values = ic.get('pos_control_values', [])
    isil_tass_fold = _safe_fold(ic.get('tass_fold_over_insilico'))
    isil_reads_fold = _safe_fold(ic.get('reads_fold_over_insilico'))
    missing_from_isil = bool(ic.get('missing_from_insilico'))

    return ControlSparkBar(
        sample_tass=member.get('tass_score', 0),
        neg_values=cc.get('neg_control_values', []),
        pos_values=cc.get('pos_control_values', []),
        control_flag=cc.get('control_flag', 'no_neg_controls'),
        bar_width=bar_width,
        bar_height=bar_height,
        tass_fold=_safe_fold(cc.get('tass_fold_over_neg')),
        reads_fold=_safe_fold(cc.get('reads_fold_over_neg')),
        pos_tass_fold=_safe_fold(cc.get('tass_fold_over_pos')),
        pos_reads_fold=_safe_fold(cc.get('reads_fold_over_pos')),
        insilico_values=isil_values,
        insilico_tass_fold=isil_tass_fold,
        insilico_reads_fold=isil_reads_fold,
        missing_from_insilico=missing_from_isil,
    )


def get_high_ani_matches(member):
    """Return the pre-computed high-ANI match list embedded in the member dict.

    match_paths.py populates each member's ``high_ani_matches`` field as a list
    of dicts::

        [{"key": "<taxid>", "ani_pct": <float 0-100>}, ...]

    sorted descending by ani_pct.  Returns an empty list when the field is
    absent (e.g. ANI matrix was not enabled during match_paths run).
    """
    return member.get('high_ani_matches') or []


def check_if_any_high_ani_in_dataset(strains):
    """Return True if any strain in the dataset has at least one high-ANI match."""
    for strain in strains:
        if get_high_ani_matches(strain):
            return True
    return False


def check_if_k2_reads_present(strains):
    for strain in strains:
        if strain.get('k2_reads') is not None and strain.get('k2_reads', 0) > 0:
            return True
    return False


def should_include_strain(strain, args, sampletype=None):
    category = str(strain.get('microbial_category', 'Unknown'))
    # Resolve sample type from argument or from the strain's own data
    _st = sampletype or strain.get('sampletype') or strain.get('normalized_sample_site') or ''
    # For sterile sample types, anything listed from the pathogen sheet
    # (i.e. any annotated organism) should always be shown regardless of category
    if _st:
        normalized_st = normalize_body_site(str(_st).lower())
        if normalized_st == "sterile":
            is_annotated = strain.get('is_annotated', 'No')
            if is_annotated == 'Yes' or 'Primary' in category or 'sterile' in category.lower():
                return True
    if 'Primary' in category:
        return True
    if "Opportunistic" in category and args.show_opportunistic:
        return True
    if 'Potential' in category and args.show_potentials:
        return True
    if 'Commensal' in category and args.show_commensals:
        return True
    if category == 'Unknown' and args.show_unidentified:
        return True
    return False


def passes_confidence_threshold(strain, threshold):
    if threshold > 1.0:
        threshold = threshold / 100.0
    tass_score = strain.get('tass_score', 0)
    return tass_score >= threshold


# ── Per-sample-type confidence defaults ──────────────────────────────────────
_SAMPLETYPE_CONF_MAP = {
    'sterile': 0.1,
    'blood':   0.1,
    'csf':     0.1,
    'stool':   0.5,
    'oral':    0.45,
    'nasal':   0.45,
    'skin':    0.55,
    'wound':   0.55,
    "vaginal": 0.5,
}
_DEFAULT_CONF = 0.5


def get_sample_min_conf(sample_name, species_groups, explicit_conf,
                        input_metadata=None):
    """Return ``(min_conf, source_label)`` for a sample.

    Resolution order:

    1. ``--min_conf`` CLI flag (user-provided, overrides everything).
    2. ``best_cutoffs`` from the per-sample metadata written by
       ``match_paths.py`` (derived from the thresholds JSON).
       We prefer the **toplevelkey** cutoff because the report groups
       organisms at that level.
    3. Hard-coded ``_SAMPLETYPE_CONF_MAP`` based on the body-site type.

    *source_label* is a human-readable string describing where the value
    came from (displayed in the report header).
    """
    # 1. Explicit CLI flag
    if explicit_conf is not None:
        return explicit_conf, "user-specified (--min_conf)"

    # 2. Best cutoff from thresholds JSON (via match_paths metadata)
    meta = (input_metadata or {}).get(sample_name, {})
    best_cutoffs = meta.get('best_cutoffs')
    if best_cutoffs:
        # Use the preferred granularity from match_paths (default: subkey),
        # then fall back through the remaining levels.
        _pref = meta.get('preferred_granularity', 'subkey')
        _fallback_order = [_pref] + [
            g for g in ("subkey", "key", "toplevelkey") if g != _pref
        ]
        for _level in _fallback_order:
            _lc = best_cutoffs.get(_level, {})
            _bt = _lc.get("best_threshold") if isinstance(_lc, dict) else None
            if _bt is not None:
                _src = meta.get('best_cutoffs_source', 'thresholds JSON')
                return float(_bt), f"derived from optimised thresholds ({', '.join(_src.split('|'))}, {_level})"

    # 3. Sampletype-based defaults
    raw_st = (species_groups[0].get('sampletype', '') if species_groups else '').strip()
    norm_st = normalize_body_site(raw_st.lower()) if raw_st else 'unknown'
    conf = _SAMPLETYPE_CONF_MAP.get(norm_st, _DEFAULT_CONF)
    return conf, f"default for sample type '{norm_st}'"


def load_json_samples(input_files):
    """Load organism data from one or more JSON files.

    Supports structured format:

        {"metadata": {...}, "organisms": [...]}

    and the legacy plain format::

        [{organism}, {organism}, ...]

    Returns (all_sample_data, all_metadata) where all_metadata is a list
    of metadata dicts (one per input file, empty dict for legacy files).
    """
    all_sample_data = []
    all_metadata = []
    for input_file in input_files:
        with open(input_file, 'r') as f:
            data = json.load(f)
        if isinstance(data, dict) and 'organisms' in data:
            all_sample_data.extend(data['organisms'])
            all_metadata.append(data.get('metadata', {}))
        elif isinstance(data, list):
            all_sample_data.extend(data)
            all_metadata.append({})
        else:
            print(f"WARNING: Unexpected JSON format in {input_file}, skipping.")
    return all_sample_data, all_metadata


def organize_data_by_sample(sample_data):
    samples_dict = {}
    for species_group in sample_data:
        sample_name = species_group.get('sample_name', 'Unknown Sample')
        if sample_name not in samples_dict:
            samples_dict[sample_name] = []
        samples_dict[sample_name].append(species_group)
    return samples_dict


def sanitize_bookmark_name(name):
    import re
    sanitized = re.sub(r'[^\w\-]', '_', str(name))
    sanitized = re.sub(r'_+', '_', sanitized)
    sanitized = sanitized.strip('_')
    return sanitized


def collect_all_bookmarks(samples_dict, low_confidence_strains):
    bookmarks = set()
    bookmarks.add('color_key')
    bookmarks.add('column_explanations')
    if low_confidence_strains:
        bookmarks.add('low_confidence')
    bookmarks.add('additional_info')
    # Strain-detail appendix anchors (always registered so links validate)
    bookmarks.add('strain_detail_table')
    for sample_name, species_groups in samples_dict.items():
        sample_bookmark = f"sample_{sanitize_bookmark_name(sample_name)}"
        bookmarks.add(sample_bookmark)
        strain_sample_bm = f"strain_sample_{sanitize_bookmark_name(sample_name)}"
        bookmarks.add(strain_sample_bm)
        for species_group in species_groups:
            species_key = species_group.get('toplevelkey', species_group.get('key', 'unknown'))
            species_bookmark = f"species_{sanitize_bookmark_name(sample_name)}_{species_key}"
            bookmarks.add(species_bookmark)
            strain_row_bm = f"strain_row_{sanitize_bookmark_name(sample_name)}_{species_key}"
            bookmarks.add(strain_row_bm)
    return bookmarks


def create_safe_link(text, bookmark, valid_bookmarks, color="blue"):
    if bookmark in valid_bookmarks:
        return f'<link href="#{bookmark}" color="{color}">{text}</link>'
    else:
        print(f"Warning: Bookmark '{bookmark}' does not exist, skipping link")
        return text


def build_taxid_to_bookmark_map(samples_dict):
    taxid_map = {}
    for sample_name, species_groups in samples_dict.items():
        sample_bookmark_part = sanitize_bookmark_name(sample_name)
        for species_group in species_groups:
            species_key = species_group.get('toplevelkey', species_group.get('key', 'unknown'))
            species_bookmark = f"species_{sample_bookmark_part}_{species_key}"
            taxid_map[str(species_key)] = species_bookmark
            for strain in get_all_strains(species_group):
                strain_key = str(strain.get('key', ''))
                if strain_key:
                    taxid_map[strain_key] = species_bookmark
    return taxid_map


def get_sample_stats(species_groups):
    total_alignments = sum(sg.get('numreads', 0) for sg in species_groups)
    primary_pathogen_count = 0
    for sg in species_groups:
        for strain in get_all_strains(sg):
            if 'Primary' in str(strain.get('microbial_category', '')):
                primary_pathogen_count += 1
    return total_alignments, primary_pathogen_count


def has_min_reads(strain, min_reads=1):
    try:
        reads = float(strain.get("numreads", 0) or 0)
    except Exception:
        reads = 0
    return reads >= min_reads


def get_species_group_stats(species_group):
    all_strains = get_all_strains(species_group)
    primary_count = sum(1 for m in all_strains if 'Primary' in str(m.get('microbial_category', '')))
    group_tass = species_group.get('tass_score', 0)
    tass_scores = [m.get('tass_score', 0) for m in all_strains]
    max_member_tass = max(tass_scores) if tass_scores else group_tass
    return group_tass, primary_count, max_member_tass


def group_members_by_subkey(members):
    """
    Group member strains by their 'subkey' field.
    Falls back to the strain's own 'key' if 'subkey' is absent.
    Returns: {subkey: [strain, ...sorted by TASS desc]}
    """
    subkey_groups = {}
    for strain in members:
        sk = str(strain.get('subkey', strain.get('key', 'unknown')))
        subkey_groups.setdefault(sk, []).append(strain)
    for sk in subkey_groups:
        subkey_groups[sk].sort(key=lambda s: s.get('tass_score', 0), reverse=True)
    return subkey_groups


def get_category_color(microbial_category, ann_class, alpha=1.0):
    val = str(microbial_category)
    derived = str(ann_class)
    if "Primary" in val and derived == "Direct":
        base_color = colors.HexColor("#E85F50")   # crimson red
    elif "Primary" in val:
        base_color = colors.HexColor('#E67E22')    # orange
    elif "Commensal" in val:
        base_color = colors.lightgreen
    elif "Opportunistic" in val:
        base_color = colors.HexColor('#ffe6a8')
    elif "Potential" in val:
        base_color = colors.lightblue
    else:
        base_color = colors.white
    if alpha < 1.0:
        return colors.Color(base_color.red, base_color.green, base_color.blue, alpha=alpha)
    return base_color


def get_derived_row_colors(alpha=1.0):
    indicator = colors.HexColor('#6F889B')
    row = colors.HexColor('#EAF1F5')
    if alpha < 1.0:
        indicator = colors.Color(indicator.red, indicator.green, indicator.blue, alpha=alpha)
        row = colors.Color(row.red, row.green, row.blue, alpha=alpha)
    return indicator, row


def has_harmful_followup_signal(members):
    """Return True when a derived row summarizes potentially harmful descendants."""
    for member in members or []:
        category = str(member.get('microbial_category', 'Unknown'))
        if member.get('high_cons', False):
            return True
        if any(tag in category for tag in ('Primary', 'Opportunistic', 'Potential')):
            return True
    return False


def collect_derived_subkey_alerts(all_strains, species_group_map, subkey_group_lookup=None):
    """Identify visible derived subkey rows that summarize lower-level strain calls."""
    species_groups = {}
    for strain in all_strains:
        species_group = species_group_map[id(strain)]
        species_key = str(species_group.get('toplevelkey', species_group.get('key', 'unknown')))
        species_groups.setdefault(species_key, []).append(strain)

    alerts = []
    for species_key, members in species_groups.items():
        if len(members) <= 1:
            continue
        subkey_members_map = group_members_by_subkey(members)
        for sk, subkey_members in subkey_members_map.items():
            best = subkey_members[0]
            best_key = str(best.get('key', ''))
            if best_key == str(sk):
                continue
            if not has_harmful_followup_signal(subkey_members):
                continue
            sk_group = (subkey_group_lookup or {}).get((species_key, str(sk))) or {}
            display_name = (
                sk_group.get('name')
                or best.get('subkeyname')
                or best.get('name', 'Unknown')
            )
            alerts.append({
                'species_key': species_key,
                'subkey': str(sk),
                'display_name': display_name,
            })
    return alerts

def _valid_num(x):
    try:
        if x is None:
            return None
        v = float(x)
        if v == 0:
            return None
        return v
    except Exception:
        return None


def strip_species_prefix(strain_name, species_name):
    """
    Remove the leading species name from a strain name so only the
    distinguishing part is shown.
    E.g. strip_species_prefix("Escherichia coli ETEC", "Escherichia coli")
         -> "ETEC"
    Returns the original name if the result would be empty or the prefix
    does not match.
    """
    import re
    if not species_name or not strain_name:
        return strain_name
    # Escape special regex chars in the species name and allow variable whitespace
    pattern = r'^\s*' + re.escape(species_name) + r'\s*'
    result = re.sub(pattern, '', strain_name, count=1, flags=re.IGNORECASE).strip()
    # Strip leading punctuation/separators left behind (e.g. "-", "_", "/")
    result = re.sub(r'^[\-_/,;:\s]+', '', result).strip()
    return result if result else strain_name


def build_insilico_metrics_table(species_groups, min_conf, available_width,
                                 missing_insilico=None, comparison_key='insilico_comparison'):
    """Build a summary table of TP/FP/precision/recall/F1 by microbial category.

    An organism is a True Positive (TP) if:
      - It exists in the insilico simulation (has comparison data with pos_control_values)
      - AND its TASS score >= min_conf (the passing threshold)

    A False Positive (FP) if:
      - It does NOT exist in the insilico simulation (missing_from_insilico or no insilico data)
      - AND its TASS score >= min_conf

    A False Negative (FN) if:
      - It exists in the insilico simulation
      - BUT its TASS score < min_conf (or it's missing from the sample entirely)

    A True Negative (TN) if:
      - It does NOT exist in the insilico simulation
      - AND its TASS score < min_conf

    Parameters
    ----------
    comparison_key : str
        The key to look up on each member/group for insilico comparison data.
        Default 'insilico_comparison' (combined). For per-type tables use
        'insilico_comparison_iss' or 'insilico_comparison_nanosim'.

    Returns a ReportLab Table or None if no insilico data present.
    """
    from collections import defaultdict

    # Categories to track
    categories = ['Primary', 'Opportunistic', 'Potential', 'Commensal', 'Unknown']
    cat_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'fn': 0, 'tn': 0,
                                      'tp_reads': 0, 'fp_reads': 0,
                                      'fn_reads': 0, 'tn_reads': 0})

    has_any_insilico = False

    for grp in species_groups:
        for member in get_all_strains(grp):
            cat_raw = str(member.get('microbial_category', 'Unknown'))
            # Map to our category bins
            cat = 'Unknown'
            for c in categories:
                if c in cat_raw:
                    cat = c
                    break

            tass = float(member.get('tass_score', 0) or 0)
            reads = float(member.get('numreads', 0) or 0)
            passes = tass >= min_conf

            ic = member.get(comparison_key) or {}
            # Determine if this organism was in the simulation
            in_sim = bool(ic.get('pos_control_values')) and not ic.get('missing_from_insilico')
            is_missing_from_insilico = bool(ic.get('missing_from_insilico'))

            if in_sim or (ic and not is_missing_from_insilico and
                          (ic.get('insilico_tass') is not None or ic.get('tass_fold_over_insilico') is not None)):
                has_any_insilico = True
                in_sim = True

            if ic or in_sim or is_missing_from_insilico:
                has_any_insilico = True

            if in_sim and passes:
                cat_stats[cat]['tp'] += 1
                cat_stats[cat]['tp_reads'] += reads
            elif in_sim and not passes:
                cat_stats[cat]['fn'] += 1
                cat_stats[cat]['fn_reads'] += reads
            elif not in_sim and passes:
                cat_stats[cat]['fp'] += 1
                cat_stats[cat]['fp_reads'] += reads
            else:
                cat_stats[cat]['tn'] += 1
                cat_stats[cat]['tn_reads'] += reads

    # Also count missing insilico organisms as FN — use their actual category
    if missing_insilico:
        has_any_insilico = True
        for entry in missing_insilico:
            cat_raw = str(entry.get('microbial_category', 'Unknown'))
            cat = 'Unknown'
            for c in categories:
                if c in cat_raw:
                    cat = c
                    break
            cat_stats[cat]['fn'] += 1
            cat_stats[cat]['fn_reads'] += float(entry.get('pos_numreads', 0) or 0)

    if not has_any_insilico:
        return None

    # Build the table
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    styles = getSampleStyleSheet()

    hdr_style = ParagraphStyle('IsilHdr', parent=styles['Normal'],
                                fontSize=7, leading=9,
                                fontName='Helvetica-Bold',
                                textColor=colors.white)
    cell_style = ParagraphStyle('IsilCell', parent=styles['Normal'],
                                 fontSize=7, leading=9,
                                 fontName='Helvetica')
    bold_cell = ParagraphStyle('IsilBold', parent=cell_style,
                                fontName='Helvetica-Bold')

    headers = ['Category', 'TP Taxa', 'FP Taxa', 'FN Taxa', 'TN Taxa',
               'TP Reads', 'FP Reads', 'Precision', 'Recall', 'F1', 'Accuracy']

    table_data = [[Paragraph(h, hdr_style) for h in headers]]

    def _fmt_pct(val):
        if val is None:
            return '-'
        return f'{val*100:.1f}%'

    def _compute_metrics(tp, fp, fn, tn):
        precision = tp / (tp + fp) if (tp + fp) > 0 else None
        recall = tp / (tp + fn) if (tp + fn) > 0 else None
        accuracy = (tp + tn) / (tp + fp + fn + tn) if (tp + fp + fn + tn) > 0 else None
        if precision is not None and recall is not None and (precision + recall) > 0:
            f1 = 2 * precision * recall / (precision + recall)
        else:
            f1 = None
        return precision, recall, f1, accuracy

    total = {'tp': 0, 'fp': 0, 'fn': 0, 'tn': 0,
             'tp_reads': 0, 'fp_reads': 0, 'fn_reads': 0, 'tn_reads': 0}

    for cat in categories:
        s = cat_stats[cat]
        if s['tp'] + s['fp'] + s['fn'] + s['tn'] == 0:
            continue
        for k in total:
            total[k] += s[k]
        prec, rec, f1, acc = _compute_metrics(s['tp'], s['fp'], s['fn'], s['tn'])
        row = [
            Paragraph(cat, cell_style),
            Paragraph(str(s['tp']), cell_style),
            Paragraph(str(s['fp']), cell_style),
            Paragraph(str(s['fn']), cell_style),
            Paragraph(str(s['tn']), cell_style),
            Paragraph(f"{s['tp_reads']:,.0f}", cell_style),
            Paragraph(f"{s['fp_reads']:,.0f}", cell_style),
            Paragraph(_fmt_pct(prec), cell_style),
            Paragraph(_fmt_pct(rec), cell_style),
            Paragraph(_fmt_pct(f1), cell_style),
            Paragraph(_fmt_pct(acc), cell_style),
        ]
        table_data.append(row)

    # Total row
    prec, rec, f1, acc = _compute_metrics(total['tp'], total['fp'], total['fn'], total['tn'])
    total_row = [
        Paragraph('<b>TOTAL</b>', bold_cell),
        Paragraph(f"<b>{total['tp']}</b>", bold_cell),
        Paragraph(f"<b>{total['fp']}</b>", bold_cell),
        Paragraph(f"<b>{total['fn']}</b>", bold_cell),
        Paragraph(f"<b>{total['tn']}</b>", bold_cell),
        Paragraph(f"<b>{total['tp_reads']:,.0f}</b>", bold_cell),
        Paragraph(f"<b>{total['fp_reads']:,.0f}</b>", bold_cell),
        Paragraph(f"<b>{_fmt_pct(prec)}</b>", bold_cell),
        Paragraph(f"<b>{_fmt_pct(rec)}</b>", bold_cell),
        Paragraph(f"<b>{_fmt_pct(f1)}</b>", bold_cell),
        Paragraph(f"<b>{_fmt_pct(acc)}</b>", bold_cell),
    ]
    table_data.append(total_row)

    n_cols = len(headers)
    w = available_width or 500
    col_w = [w * 0.12] + [w * (0.88 / (n_cols - 1))] * (n_cols - 1)

    tbl = Table(table_data, colWidths=col_w)
    tbl_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#008B8B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.Color(0.7, 0.7, 0.7)),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.Color(0.97, 0.97, 0.97),
                                                colors.white]),
        # Total row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E0F7F7')),
        ('LINEABOVE', (0, -1), (-1, -1), 0.8, colors.HexColor('#008B8B')),
    ]
    tbl.setStyle(TableStyle(tbl_styles))
    return tbl


def build_missing_insilico_detail_table(missing_insilico, available_width):
    """Build a detail table listing specific organisms present in insilico but missing from sample.

    Each row shows the organism name, taxid, category, and the insilico TASS/reads.
    Returns a ReportLab Table or None if no missing organisms.
    """
    if not missing_insilico:
        return None

    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    styles = getSampleStyleSheet()

    hdr_style = ParagraphStyle('MissIsilHdr', parent=styles['Normal'],
                                fontSize=7, leading=9,
                                fontName='Helvetica-Bold',
                                textColor=colors.white)
    cell_style = ParagraphStyle('MissIsilCell', parent=styles['Normal'],
                                 fontSize=7, leading=9,
                                 fontName='Helvetica')

    headers = ['Organism', 'Taxid', 'Category', 'InSilico TASS', 'InSilico Reads', 'Status']
    table_data = [[Paragraph(h, hdr_style) for h in headers]]

    for entry in missing_insilico:
        name = entry.get('name', 'Unknown')
        taxid = str(entry.get('id', '-'))
        cat = entry.get('microbial_category', 'Unknown')
        tass = entry.get('pos_tass_score', 0)
        reads = entry.get('pos_numreads', 0)
        tass_str = f"{float(tass) * 100:.2f}" if tass else '-'
        reads_str = f"{float(reads):,.0f}" if reads else '-'

        row = [
            Paragraph(name, cell_style),
            Paragraph(taxid, cell_style),
            Paragraph(cat, cell_style),
            Paragraph(tass_str, cell_style),
            Paragraph(reads_str, cell_style),
            Paragraph('<font color="red">Missing from sample (FN)</font>', cell_style),
        ]
        table_data.append(row)

    w = available_width or 500
    col_w = [w * 0.28, w * 0.10, w * 0.14, w * 0.14, w * 0.14, w * 0.20]

    tbl = Table(table_data, colWidths=col_w)
    tbl_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#B22222')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.Color(0.7, 0.7, 0.7)),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.Color(1.0, 0.95, 0.95),
                                                colors.white]),
    ]
    tbl.setStyle(TableStyle(tbl_styles))
    return tbl


def create_combined_sample_table(all_strains, species_group_map, small_style,
                                  show_ani_column, show_k2_column,
                                  taxid_to_bookmark, valid_bookmarks,
                                  sample_total_reads=0, sample_name=None,
                                  available_width=None, use_subkey=True,
                                  show_strains_table=True,
                                  outline_collector=None,
                                  zscore_threshold=None,
                                  zscore_separator_index=None,
                                  sampletype=None,
                                  show_control_bar=False,
                                  missing_pos_controls=None,
                                  subkey_group_lookup=None,
                                  subkey_display_lookup=None):
    """
    Create a single table combining all strains from all species groups.

    When use_subkey=True (default):
            - Each genus/toplevel group is followed by a species/subkey summary row.
            - Any child strains that also pass the cutoff are rendered immediately
                below that species/subkey row.

    When use_subkey=False:
      - Original flat per-strain layout (one row per strain, original behaviour).
    """
    # ── Styles ────────────────────────────────────────────────────────────────
    strain_name_style = ParagraphStyle(
        'StrainName', parent=small_style, fontSize=10, leading=10,
        wordWrap='CJK')
    data_style = ParagraphStyle(
        'DataStyle', parent=small_style, fontSize=8, leading=9,
        wordWrap='CJK')
    strain_name_style_small = ParagraphStyle(
        'StrainNameSmall', parent=small_style, fontSize=8, leading=9,
        wordWrap='CJK')
    data_style_small = ParagraphStyle(
        'DataStyleSmall', parent=small_style, fontSize=7, leading=8,
        wordWrap='CJK')
    ani_style_small = ParagraphStyle(
        'ANIStyleSmall', parent=small_style, fontSize=5, leading=7,
        wordWrap='CJK')
    ani_style = ParagraphStyle(
        'ANIStyle', parent=small_style, fontSize=6, leading=8,
        wordWrap='CJK')
    species_name_style = ParagraphStyle(
        'SpeciesName', parent=small_style, fontSize=10.5, leading=11.5,
        fontName='Helvetica-Bold', leftIndent=6, wordWrap='CJK')
    species_name_style_small = ParagraphStyle(
        'SpeciesNameSmall', parent=small_style, fontSize=9.5, leading=10.5,
        fontName='Helvetica-Bold', leftIndent=6, wordWrap='CJK')
    species_data_style = ParagraphStyle(
        'SpeciesData', parent=small_style, fontSize=9.5, leading=10.5,
        wordWrap='CJK')
    child_strain_name_style = ParagraphStyle(
        'ChildStrainName', parent=small_style, fontSize=7.2, leading=8.2,
        leftIndent=10, wordWrap='CJK')
    child_strain_name_style_small = ParagraphStyle(
        'ChildStrainNameSmall', parent=small_style, fontSize=6.7, leading=7.6,
        leftIndent=10, wordWrap='CJK')
    child_strain_data_style = ParagraphStyle(
        'ChildStrainData', parent=small_style, fontSize=6.8, leading=7.8,
        wordWrap='CJK')
    group_header_style = ParagraphStyle(
        'GroupHeader', parent=small_style, fontSize=12, leading=13,
        fontName='Helvetica-Bold', wordWrap='CJK')
    group_strain_summary_style = ParagraphStyle(
        'GroupStrainSummary', parent=small_style, fontSize=7, leading=9,
        alignment=TA_RIGHT, fontName='Helvetica-Oblique', wordWrap='CJK')
    indicator_para_style = ParagraphStyle(
        'IndicatorPara', parent=small_style, fontSize=14, leading=16,
        spaceBefore=0, spaceAfter=0,
        alignment=TA_CENTER, fontName='Helvetica-Bold')
    mini_style = ParagraphStyle(
        'MiniStyle', parent=small_style, fontSize=7, leading=9,
        wordWrap='CJK')
    mini_header_style = ParagraphStyle(
        'MiniHeaderStyle', parent=small_style, fontSize=7, leading=9,
        fontName='Helvetica-Bold', wordWrap='CJK')

    if available_width is None:
        available_width = 8.5 * inch - 0.02 * 8.5 * inch

    def _has_any_inline_strain_tables():
        if not (use_subkey and show_strains_table):
            return False
        if subkey_display_lookup:
            return False
        strains_by_group = {}
        for s in all_strains:
            sg = species_group_map[id(s)]
            gk = sg.get('toplevelkey', sg.get('key', 'unknown'))
            strains_by_group.setdefault(gk, []).append(s)
        for gk, group_strains in strains_by_group.items():
            if not group_strains:
                continue
            total_group_members = len(group_strains)
            subkey_groups = {}
            for s in group_strains:
                sk = str(s.get('subkey', s.get('key', 'unknown')))
                subkey_groups.setdefault(sk, []).append(s)
            for sk, members in subkey_groups.items():
                has_multiple_members = len(members) > 1
                has_different_keys = any(str(m.get('key', '')) != str(sk) for m in members)
                sublevel_members = [m for m in members if str(m.get('key', '')) != str(sk)]
                is_name_switch = total_group_members == 1 and has_different_keys
                has_appendix_entry = (
                    (has_multiple_members or has_different_keys)
                    and sublevel_members
                    and not is_name_switch
                )
                if has_appendix_entry:
                    return True
        return False

    show_inline_table = use_subkey and show_strains_table and _has_any_inline_strain_tables()

    # ── Headers ──────────────────────────────────────────────────────────────
    header_style = ParagraphStyle(
        'HeaderStyle', parent=small_style, fontSize=8, leading=9,
        fontName='Helvetica-Bold', textColor=colors.whitesmoke,
        alignment=TA_CENTER, wordWrap='CJK')
    base_headers = ['', Paragraph('Organism', header_style), Paragraph('TASS', header_style)]
    if show_k2_column:
        base_headers.append(Paragraph('K2<br/>Reads', header_style))
    base_headers += [
        Paragraph('Reads', header_style),
        Paragraph('RPM', header_style),
        Paragraph('Cov.', header_style),
    ]
    if show_ani_column:
        base_headers.append(Paragraph('High<br/>ANI', header_style))
    if show_control_bar:
        base_headers.append(Paragraph('Ctrl', header_style))
    n_base = len(base_headers)

    # Strain detail column in subkey mode (single column holding a nested mini-table).
    # In compact/appendix mode the arrow is embedded inline in the name cell,
    # so no extra column is needed.
    if show_inline_table:
        headers = base_headers + ['']
    else:
        headers = base_headers
    n_total = len(headers)

    # ── Column widths ─────────────────────────────────────────────────────────
    #   Columns: [indicator, Organism, TASS, (K2 Reads), Reads, RPM, Coverage, (High ANI), (Ctrl)]
    def _base_col_widths(w):
        if show_control_bar:
            # When control bar is present, allocate ~12% for it and shrink organism
            _ctrl_w = 0.12
            if show_k2_column and show_ani_column:
                base = [w*0.03, w*0.18, w*0.07, w*0.09, w*0.12, w*0.09, w*0.09, w*0.15]
            elif show_k2_column:
                base = [w*0.03, w*0.25, w*0.08, w*0.11, w*0.13, w*0.11, w*0.14]
            elif show_ani_column:
                base = [w*0.03, w*0.23, w*0.08, w*0.13, w*0.10, w*0.10, w*0.17]
            else:
                base = [w*0.03, w*0.34, w*0.09, w*0.14, w*0.12, w*0.14]
            base.append(w * _ctrl_w)
            return base
        else:
            if show_k2_column and show_ani_column:
                base = [w*0.03, w*0.24, w*0.08, w*0.11, w*0.14, w*0.10, w*0.10, w*0.20]
            elif show_k2_column:
                base = [w*0.03, w*0.30, w*0.09, w*0.13, w*0.15, w*0.13, w*0.17]
            elif show_ani_column:
                base = [w*0.03, w*0.28, w*0.09, w*0.15, w*0.12, w*0.12, w*0.21]
            else:
                base = [w*0.03, w*0.40, w*0.10, w*0.17, w*0.14, w*0.16]
            return base

    def build_metrics_mini_table(max_cds, max_mmbert, max_width=None):
        """
        Create a proper table showing CDS and/or mmbert with column headers.
        Returns None if neither metric is present.
        Hides columns when data isn't available.
        max_width constrains the total table width to fit within a container.
        """
        cds_v = _valid_num(max_cds)
        mm_v = _valid_num(max_mmbert)

        # If neither metric is present, return None
        if cds_v is None and mm_v is None:
            return None

        # Build header and data rows based on what's available
        rows = []

        # Determine which columns to include
        has_cds = cds_v is not None
        has_mmbert = mm_v is not None

        # Build header
        header_row = []
        if has_cds:
            header_row.append(Paragraph(f"<b>CDS</b>", mini_header_style))
        if has_mmbert:
            header_row.append(Paragraph(f"<b>mmbert %</b>", mini_header_style))
        rows.append(header_row)

        # Build data row
        data_row = []
        if has_cds:
            data_row.append(Paragraph(f"{cds_v:.0f}", mini_style))
        if has_mmbert:
            mmbert_pct = mm_v * 100  # Convert to percentage
            data_row.append(Paragraph(f"{mmbert_pct:.2f}%", mini_style))
        rows.append(data_row)

        # Calculate column widths, respecting max_width if provided
        desired_cds_w = 0.5 * inch
        desired_mm_w = 0.65 * inch
        col_widths_metrics = []
        if has_cds:
            col_widths_metrics.append(desired_cds_w)
        if has_mmbert:
            col_widths_metrics.append(desired_mm_w)

        # Scale down if total exceeds max_width
        if max_width is not None:
            total_desired = sum(col_widths_metrics)
            if total_desired > max_width:
                scale = max_width / total_desired
                col_widths_metrics = [w * scale for w in col_widths_metrics]

        # Create table with only the necessary columns
        t = Table(rows, colWidths=col_widths_metrics)
        t.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D5E8FF')),  # Header background
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        return t
    if show_inline_table:
        left_w = available_width * 0.62
        right_w = available_width * 0.38
        col_widths = _base_col_widths(left_w) + [right_w]
    else:
        # Compact/appendix mode: no separate right column — arrow is inline in name cell
        right_w = available_width * 0.38  # kept for metrics_max_w calc but not added as column
        col_widths = _base_col_widths(available_width)

    # ── Table data ────────────────────────────────────────────────────────────
    table_data = [headers]
    table_styles = []
    group_row_indices = []

    def _build_high_ani_paragraph(record, style):
        if not show_ani_column:
            return None
        matches = get_high_ani_matches(record)
        if not matches:
            return Paragraph('-', style)
        ani_links = []
        for m in matches[:3]:
            taxid = str(m.get('key', ''))
            ani_pct = float(m.get('ani_pct', 0))
            if taxid in taxid_to_bookmark:
                bm = taxid_to_bookmark[taxid]
                if bm in valid_bookmarks:
                    ani_links.append(
                        f'<link href="#{bm}" color="blue">{taxid} ({ani_pct:.1f}%)</link>')
                else:
                    ani_links.append(f"{taxid} ({ani_pct:.1f}%)")
            else:
                ani_links.append(f"{taxid} ({ani_pct:.1f}%)")
        return Paragraph(', '.join(ani_links), style)

    def _row_zscore_state(base_record, extra_records=None):
        if zscore_threshold is None:
            return False, ''
        records = [base_record]
        records.extend(extra_records or [])
        max_z = max(float(rec.get('zscore', 0) or 0) for rec in records)
        row_below = max_z < zscore_threshold
        if not row_below:
            return False, ''
        ref_record = base_record
        n_samples = int(ref_record.get('hmp_num_samples', 0) or 0)
        n_sites = int(ref_record.get('hmp_site_count', 0) or 0)
        if n_sites > 0:
            pct = (n_samples / n_sites) * 100
            pct_str = '&lt;0.001%' if pct < 0.001 else f'{pct:.1f}%'
            sample_label = f' <font color="#999999" size="6">{n_samples} ({pct_str})</font>'
        else:
            sample_label = ''
        return True, f' <font color="#999999">&#9830;</font>{sample_label}<font color="#999999">&#9830;</font>'

    def _append_species_summary_row(species_entry):
        nonlocal row_idx

        species_record = species_entry['species_record']
        visible_strains = species_entry['visible_strains']
        microbial_category = species_entry['microbial_category']
        ann_class = species_entry['annClass']
        indicator_text = '★' if species_entry['high_cons'] else ''
        row_below_zscore, zscore_sym = _row_zscore_state(species_record, visible_strains)
        row_name_style = species_name_style_small if row_below_zscore else species_name_style
        row_data_style = data_style_small if row_below_zscore else species_data_style
        row_ani_style = ani_style_small if row_below_zscore else ani_style
        species_key = species_entry['subkey']
        display_name = species_entry['display_name']
        # followup_symbol = ''
        # if species_entry['harmful_followup']:
        #     followup_symbol = f' {_WARN_SYMBOL_HTML}'
        if species_entry['species_passes']:
            row_marker = ' <font color="#666666" size="7"><i>species</i></font>'
        else:
            row_marker = ' <font color="#666666" size="7"><i>species from qualifying strains</i></font>'
        flora_tag = _commensal_site_tag(species_record, species_record.get('normalized_sample_site', ''))
        name_html = (
            f'{display_name} '
            f'(<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={species_key}" '
            f'color="blue">{species_key}</link>)'
            f'{row_marker}{flora_tag}{zscore_sym}'
        )

        species_reads = float(species_record.get('numreads', 0) or 0)
        pct = (species_reads / sample_total_reads * 100.0) if sample_total_reads else 0.0
        rpm = species_record.get('rpm', 0) or 0
        high_ani_text = _build_high_ani_paragraph(species_record, row_ani_style)

        _below_primary = species_entry.get('below_threshold_primary', [])
        if _below_primary:
            _bt = _below_primary[0]
            _bt_name = _bt.get('name', 'Unknown')
            _bt_tass = float(_bt.get('tass_score', 0) or 0) * 100
            _bt_reads = int(_bt.get('numreads', 0) or 0)
            name_html += (
                f'<br/><font color="#808080" size="6">'
                f'&nbsp;&nbsp;&nbsp;<font color="#8B0000">!</font> &nbsp;{_bt_name}</font>'
                f'<br/><font color="#808080" size="5.5">'
                f'&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;'
                f'TASS {_bt_tass:.1f}&nbsp;&nbsp;{_bt_reads:,} reads</font>'
            )
        _name_cell = Paragraph(name_html, row_name_style)

        row = [
            Paragraph(indicator_text, indicator_para_style) if indicator_text else '',
            _name_cell,
            Paragraph(f"{species_record.get('tass_score', 0)*100:.1f}", row_data_style),
        ]
        if show_k2_column:
            row.append(Paragraph(f"{species_record.get('k2_reads', 0):,.0f}", row_data_style))
        row.append(Paragraph(f"{species_reads:,.0f} ({pct:.1f}%)", row_data_style))
        row.append(Paragraph(f"{rpm:,.0f}", row_data_style))
        row.append(Paragraph(f"{min(100, species_record.get('coverage', 0)*100):.1f}%", row_data_style))
        if show_ani_column:
            row.append(high_ani_text)
        if show_control_bar:
            spark = _build_spark_bar_from_member(species_record, bar_width=int(col_widths[-2 if show_inline_table else -1] - 6) if col_widths else 72, bar_height=10)
            row.append(spark if spark else Paragraph('-', row_data_style))
        if show_inline_table:
            row.append('')

        table_data.append(row)

        flora_fade = _is_flora_on_sterile(species_record, species_record.get('normalized_sample_site', ''))
        flora_mult = 0.70 if flora_fade else 1.0
        if 'Primary' in str(microbial_category):
            # Use light red (#E8A0A0) ONLY when the species-level passes the threshold but
            # ALL individual strains are below it (summary row represents an inferred call).
            # Use dark red (#E85F50) when at least one qualifying strain also passes.
            _species_above_strains_below = (
                species_entry.get('species_passes') and not species_entry.get('visible_strains')
            )
            _primary_hex = '#E8A0A0' if _species_above_strains_below else '#E85F50'
            _SUMMARY_PRIMARY = colors.HexColor(_primary_hex)
            if row_below_zscore:
                ind_color = colors.Color(_SUMMARY_PRIMARY.red, _SUMMARY_PRIMARY.green, _SUMMARY_PRIMARY.blue, alpha=0.35 * flora_mult)
                row_color = colors.Color(_SUMMARY_PRIMARY.red, _SUMMARY_PRIMARY.green, _SUMMARY_PRIMARY.blue, alpha=0.08 * flora_mult)
            else:
                ind_color = colors.Color(_SUMMARY_PRIMARY.red, _SUMMARY_PRIMARY.green, _SUMMARY_PRIMARY.blue, alpha=1.0 * flora_mult)
                row_color = colors.Color(_SUMMARY_PRIMARY.red, _SUMMARY_PRIMARY.green, _SUMMARY_PRIMARY.blue, alpha=0.20 * flora_mult)
        else:
            if row_below_zscore:
                ind_color = get_category_color(microbial_category, ann_class, alpha=0.35 * flora_mult)
                row_color = get_category_color(microbial_category, ann_class, alpha=0.08 * flora_mult)
            else:
                ind_color = get_category_color(microbial_category, ann_class, alpha=1.0 * flora_mult)
                row_color = get_category_color(microbial_category, ann_class, alpha=0.20 * flora_mult)
        table_styles.append(('BACKGROUND', (0, row_idx), (0, row_idx), ind_color))
        table_styles.append(('BACKGROUND', (1, row_idx), (-1, row_idx), row_color))
        table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx), 0.8, colors.HexColor('#D8D8D8')))
        table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 4))
        table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 4))
        # Species row: slight indent from genus header
        table_styles.append(('LEFTPADDING', (1, row_idx), (1, row_idx), 14))
        row_idx += 1

    def _append_child_strain_row(strain, promoted=False, below_threshold_primary=None):
        nonlocal row_idx

        microbial_category = strain.get('microbial_category', 'Unknown')
        ann_class = strain.get('annClass', '')
        indicator_text = '★' if strain.get('high_cons', False) else ''
        row_below_zscore, zscore_sym = _row_zscore_state(strain)
        if promoted:
            row_name_style = species_name_style_small if row_below_zscore else species_name_style
            row_data_style = data_style_small if row_below_zscore else species_data_style
        else:
            row_name_style = child_strain_name_style_small if row_below_zscore else child_strain_name_style
            row_data_style = data_style_small if row_below_zscore else child_strain_data_style
        row_ani_style = ani_style_small if row_below_zscore else ani_style
        strain_key = strain.get('key', '')
        flora_tag = _commensal_site_tag(strain, strain.get('normalized_sample_site', ''))
        if promoted:
            followup_symbol = ''
            if strain.get('harmful_followup'):
                followup_symbol = f' {_WARN_SYMBOL_HTML}'
            name_html = (
                f'{followup_symbol}'
                f'{strain.get("name", "Unknown")} '
                f'(<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={strain_key}" '
                f'color="blue">{strain_key}</link>)'
                f' <font color="#666666" size="7"><i>strain</i></font>'
                f'{flora_tag}{zscore_sym}'
            )
        else:
            name_html = (
                f'! {strain.get("name", "Unknown")} '
                f'(<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={strain_key}" '
                f'color="blue">{strain_key}</link>) '
                f'<font color="#666666" size="6"><i>qualifying strain</i></font>'
                f'{flora_tag}{zscore_sym}'
            )
        high_ani_text = _build_high_ani_paragraph(strain, row_ani_style)
        strain_reads = float(strain.get('numreads', 0) or 0)
        pct = (strain_reads / sample_total_reads * 100.0) if sample_total_reads else 0.0
        rpm = strain.get('rpm', 0) or 0

        _name_cell_s = Paragraph(name_html, row_name_style)

        row = [
            Paragraph(indicator_text, indicator_para_style) if indicator_text else '',
            _name_cell_s,
            Paragraph(f"{strain.get('tass_score', 0)*100:.1f}", row_data_style),
        ]
        if show_k2_column:
            row.append(Paragraph(f"{strain.get('k2_reads', 0):,.0f}", row_data_style))
        row.append(Paragraph(f"{strain_reads:,.0f} ({pct:.1f}%)", row_data_style))
        row.append(Paragraph(f"{rpm:,.0f}", row_data_style))
        row.append(Paragraph(f"{min(100, strain.get('coverage', 0)*100):.1f}%", row_data_style))
        if show_ani_column:
            row.append(high_ani_text)
        if show_control_bar:
            spark = _build_spark_bar_from_member(strain, bar_width=int(col_widths[-2 if show_inline_table else -1] - 6) if col_widths else 72, bar_height=10)
            row.append(spark if spark else Paragraph('-', row_data_style))
        if show_inline_table:
            row.append('')

        table_data.append(row)

        flora_fade = _is_flora_on_sterile(strain, strain.get('normalized_sample_site', ''))
        flora_mult = 0.70 if flora_fade else 1.0
        if promoted:
            # Promoted strains are always above the threshold (they come from visible_strains),
            # so use dark red (#E85F50) for Primary — never the light summary shade (#E8A0A0).
            if 'Primary' in str(microbial_category):
                _PROMOTED_PRIMARY = colors.HexColor('#E85F50')
                if row_below_zscore:
                    ind_color = colors.Color(_PROMOTED_PRIMARY.red, _PROMOTED_PRIMARY.green, _PROMOTED_PRIMARY.blue, alpha=0.35 * flora_mult)
                    row_color = colors.Color(_PROMOTED_PRIMARY.red, _PROMOTED_PRIMARY.green, _PROMOTED_PRIMARY.blue, alpha=0.08 * flora_mult)
                else:
                    ind_color = colors.Color(_PROMOTED_PRIMARY.red, _PROMOTED_PRIMARY.green, _PROMOTED_PRIMARY.blue, alpha=1.0 * flora_mult)
                    row_color = colors.Color(_PROMOTED_PRIMARY.red, _PROMOTED_PRIMARY.green, _PROMOTED_PRIMARY.blue, alpha=0.20 * flora_mult)
            else:
                if row_below_zscore:
                    ind_color = get_category_color(microbial_category, ann_class, alpha=0.35 * flora_mult)
                    row_color = get_category_color(microbial_category, ann_class, alpha=0.08 * flora_mult)
                else:
                    ind_color = get_category_color(microbial_category, ann_class, alpha=1.0 * flora_mult)
                    row_color = get_category_color(microbial_category, ann_class, alpha=0.20 * flora_mult)
        elif row_below_zscore:
            ind_color = get_category_color(microbial_category, ann_class, alpha=0.30 * flora_mult)
            row_color = get_category_color(microbial_category, ann_class, alpha=0.04 * flora_mult)
        else:
            ind_color = get_category_color(microbial_category, ann_class, alpha=1.0 * flora_mult)
            row_color = get_category_color(microbial_category, ann_class, alpha=0.10 * flora_mult)
        table_styles.append(('BACKGROUND', (0, row_idx), (0, row_idx), ind_color))
        table_styles.append(('BACKGROUND', (1, row_idx), (-1, row_idx), row_color))
        if promoted:
            table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx), 0.8, colors.HexColor('#D8D8D8')))
            table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 4))
            table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 4))
            table_styles.append(('LEFTPADDING', (1, row_idx), (1, row_idx), 14))
        else:
            table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx), 0.5, colors.HexColor('#E3E3E3')))
            table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 3))
            table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 3))
            # Strain row: deeper indent than species, clearly subordinate
            table_styles.append(('LEFTPADDING', (1, row_idx), (1, row_idx), 28))
        row_idx += 1

    def _append_subkey_row(best, detail_members, species_key, sk,
                           single_strain_in_group, row_variant='standard',
                           display_name_override=None, display_key_override=None,
                           harmful_followup=False):
        nonlocal row_idx

        # Prefer subkey group's annotations over individual strain's
        _sk_grp_ann = (subkey_group_lookup or {}).get((species_key, sk)) or {}
        microbial_category = _sk_grp_ann.get('microbial_category') or best.get('microbial_category', 'Unknown')
        ann_class = _sk_grp_ann.get('annClass') or best.get('annClass', '')
        is_hc = _sk_grp_ann.get('high_cons', best.get('high_cons', False))
        indicator_text = '★' if is_hc else ''

        _row_below_zscore = False
        if zscore_threshold is not None:
            _all_member_zscores = [float(best.get('zscore', 0) or 0)]
            _all_member_zscores.extend(
                float(m.get('zscore', 0) or 0) for m in detail_members)
            _row_max_z = max(_all_member_zscores)
            _row_below_zscore = _row_max_z < zscore_threshold

        best_key = str(best.get('key', ''))
        # When a subkey group exists, prefer its species-level name/key
        _sk_grp_info = (subkey_group_lookup or {}).get((species_key, sk))
        _has_sk_name = bool(_sk_grp_info and _sk_grp_info.get('name'))
        if display_name_override is not None:
            display_name = display_name_override
        elif _has_sk_name and str(best.get('subkey', best_key)) != best_key:
            # Subkey group available and strain is below species level — use species name
            display_name = _sk_grp_info.get('name', best.get('subkeyname', best.get('name', 'Unknown')))
        elif single_strain_in_group:
            display_name = best.get('name', 'Unknown')
        elif str(best.get('subkey', best_key)) != best_key:
            display_name = best.get('subkeyname', best.get('name', 'Unknown'))
        else:
            display_name = best.get('name', 'Unknown')

        if display_key_override is not None:
            display_key = str(display_key_override)
        elif _has_sk_name and str(best.get('subkey', best_key)) != best_key:
            display_key = str(best.get('subkey', best_key))
        elif single_strain_in_group:
            display_key = best_key
        elif str(best.get('subkey', best_key)) != best_key:
            display_key = str(best.get('subkey', best_key))
        else:
            display_key = best_key

        if row_variant == 'species_level':
            row_marker = ' <font color="#666666" size="7"><i>species level</i></font>'
        elif row_variant == 'derived_subkey':
            row_marker = ' <font color="#6F889B" size="7"><i>derived from lower-level strains</i></font>'
        elif row_variant == 'derived':
            row_marker = ' <font color="#6F889B" size="7"><i>derived strain call</i></font>'
        else:
            row_marker = ''

        followup_symbol = ''
        if harmful_followup:
            followup_symbol = f' {_WARN_SYMBOL_HTML}'

        if _row_below_zscore:
            _n_samples = int(best.get('hmp_num_samples', 0) or 0)
            _n_sites = int(best.get('hmp_site_count', 0) or 0)
            if _n_sites > 0:
                _pct = (_n_samples / _n_sites) * 100
                _pct_str = '&lt;0.001%' if _pct < 0.001 else f'{_pct:.1f}%'
                _sample_label = f' <font color="#999999" size="6">{_n_samples} ({_pct_str})</font>'
            else:
                _sample_label = ''
            _zscore_sym = f' <font color="#999999">&#9830;</font>{_sample_label}<font color="#999999">&#9830;</font>'
        else:
            _zscore_sym = ''

        _flora_tag = _commensal_site_tag(best, best.get('normalized_sample_site', ''))
        name_html_base = (
            f'{display_name} '
            f'(<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={display_key}" '
            f'color="blue">{display_key}</link>){followup_symbol}{row_marker}{_flora_tag}{_zscore_sym}'
        )

        has_appendix_entry = bool(detail_members)
        if not show_inline_table and has_appendix_entry:
            strain_row_bm = f"strain_row_{sanitize_bookmark_name(sample_name)}_{species_key}"
            arrow_link = create_safe_link('Strain detail \u2193', strain_row_bm,
                                          valid_bookmarks, color='blue')
            name_html = f'{name_html_base} {arrow_link}'
        else:
            name_html = name_html_base

        strain_reads = float(best.get('numreads', 0) or 0)
        pct = (strain_reads / sample_total_reads * 100.0) if sample_total_reads else 0.0
        rpm = best.get('rpm', 0) or 0

        high_ani_text = ''
        if show_ani_column:
            matches = get_high_ani_matches(best)
            if matches:
                ani_links = []
                for m in matches[:3]:
                    taxid = str(m.get('key', ''))
                    ani_pct = float(m.get('ani_pct', 0))
                    if taxid in taxid_to_bookmark:
                        bm = taxid_to_bookmark[taxid]
                        if bm in valid_bookmarks:
                            ani_links.append(
                                f'<link href="#{bm}" color="blue">{taxid} ({ani_pct:.1f}%)</link>')
                        else:
                            ani_links.append(f"{taxid} ({ani_pct:.1f}%)")
                    else:
                        ani_links.append(f"{taxid} ({ani_pct:.1f}%)")
                high_ani_text = Paragraph(", ".join(ani_links), ani_style)
            else:
                high_ani_text = Paragraph("-", ani_style)

        row_name_style = strain_name_style_small if _row_below_zscore else strain_name_style
        row_data_style = data_style_small if _row_below_zscore else data_style
        row_ani_style = ani_style_small if _row_below_zscore else ani_style

        # Use subkey group's TASS score when available (species-level aggregation)
        _display_tass = best.get('tass_score', 0)
        if subkey_group_lookup:
            _sk_grp = subkey_group_lookup.get((species_key, sk))
            if _sk_grp:
                _display_tass = _sk_grp.get('tass_score', _display_tass)

        row = [
            Paragraph(indicator_text, indicator_para_style) if indicator_text else '',
            Paragraph(name_html, row_name_style),
            Paragraph(f"{_display_tass*100:.1f}", row_data_style),
        ]
        if show_k2_column:
            row.append(Paragraph(f"{best.get('k2_reads', 0):,.0f}", row_data_style))
        row.append(Paragraph(f"{strain_reads:,.0f} ({pct:.1f}%)", row_data_style))
        row.append(Paragraph(f"{rpm:,.0f}", row_data_style))
        row.append(Paragraph(f"{min(100, best.get('coverage', 0)*100):.1f}%", row_data_style))
        if show_ani_column:
            if isinstance(high_ani_text, Paragraph) and _row_below_zscore:
                high_ani_text = Paragraph(high_ani_text.text, row_ani_style)
            row.append(high_ani_text)
        if show_control_bar:
            _spark = _build_spark_bar_from_member(best, bar_width=int(col_widths[-2 if show_inline_table else -1] - 6) if col_widths else 72, bar_height=10)
            row.append(_spark if _spark else Paragraph('-', row_data_style))

        if show_inline_table:
            if has_appendix_entry:
                mini_rows = [[
                    Paragraph('<b>Additional Strain/Subsp.</b>', mini_header_style),
                    Paragraph('<b>TASS</b>', mini_header_style),
                    Paragraph('<b>Reads Aligned</b>', mini_header_style),
                ]]
                for m in detail_members:
                    m_reads = float(m.get('numreads', 0) or 0)
                    m_pct = (m_reads / sample_total_reads * 100.0) if sample_total_reads else 0.0
                    m_key = m.get('key', '')
                    m_name = m.get('name', 'Unknown')
                    m_star = '★ ' if m.get('high_cons', False) else ''
                    mini_rows.append([
                        Paragraph(
                            f'{m_star}<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={m_key}" '
                            f'color="blue">{m_name}</link>',
                            mini_style
                        ),
                        Paragraph(f"{m.get('tass_score', 0)*100:.1f}", mini_style),
                        Paragraph(f"{m_reads:,.0f} ({m_pct:.1f}%)", mini_style),
                    ])
                mini_avail = right_w - 6
                mini_col_widths = [mini_avail * 0.50, mini_avail * 0.20, mini_avail * 0.30]
                mini_tbl = Table(mini_rows, colWidths=mini_col_widths)
                mini_tbl.setStyle(TableStyle([
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D5E8FF')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (1, 0), (2, -1), 'CENTER'),
                ]))
                row.append(mini_tbl)
            else:
                row.append('')

        table_data.append(row)

        print(f"  Row {row_idx}: [{row_variant} sk={sk}] {best.get('name', '')[:40]} "
              f"({len(detail_members)} detail member(s)) - Cat: {microbial_category}")

        _flora_fade = _is_flora_on_sterile(best, best.get('normalized_sample_site', ''))
        _flora_mult = 0.70 if _flora_fade else 1.0
        if row_variant in {'derived', 'derived_subkey'}:
            if _row_below_zscore:
                ind_color, row_color = get_derived_row_colors(alpha=0.35 * _flora_mult)
            else:
                ind_color, row_color = get_derived_row_colors(alpha=1.0 * _flora_mult)
        else:
            if _row_below_zscore:
                ind_color = get_category_color(microbial_category, ann_class, alpha=0.35 * _flora_mult)
                row_color = get_category_color(microbial_category, ann_class, alpha=0.05 * _flora_mult)
            else:
                ind_color = get_category_color(microbial_category, ann_class, alpha=1.0 * _flora_mult)
                row_color = get_category_color(microbial_category, ann_class, alpha=0.15 * _flora_mult)
        table_styles.append(('BACKGROUND', (0, row_idx), (0, row_idx), ind_color))
        table_styles.append(('BACKGROUND', (1, row_idx), (-1, row_idx), row_color))
        if _row_below_zscore:
            table_styles.append(('TEXTCOLOR', (1, row_idx), (-1, row_idx),
                                 colors.Color(0.6, 0.6, 0.6, 1)))
            table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 3))
            table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 3))
            table_styles.append(('FONTSIZE', (1, row_idx), (-1, row_idx), 7))
        table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx), 1.5, colors.HexColor('#CCCCCC')))
        if not _row_below_zscore:
            table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 8))
            table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 6))
        row_idx += 1

    current_species_key = None
    row_idx = 1  # row 0 is the header
    emitted_subkeys_per_group = {}  # {species_key: set of already-emitted subkeys}

    # Pre-compute strain names per species group for group header summaries
    strains_per_group = {}
    for s in all_strains:
        sg = species_group_map[id(s)]
        gk = sg.get('toplevelkey', sg.get('key', 'unknown'))
        strains_per_group.setdefault(gk, []).append(s.get('name', 'Unknown'))
    group_mmbert_max = {}
    group_dmnd_cds_max = {}
    for s in all_strains:
        sg = species_group_map[id(s)]
        gk = sg.get('toplevelkey', sg.get('key', 'unknown'))

        # Use only the GROUP-level mmbert, not member-level
        group_mm = _valid_num(sg.get('mmbert', None))
        if group_mm is not None:
            group_mmbert_max[gk] = group_mm

        best_dmnd = None
        for c in [
            (sg.get('diamond') or {}).get('cds', None),
            (s.get('diamond') or {}).get('cds', None)
        ]:
            c_v = _valid_num(c)
            if c_v is not None:
                best_dmnd = c_v if best_dmnd is None else max(best_dmnd, c_v)
        if best_dmnd is not None:
            group_dmnd_cds_max[gk] = best_dmnd
    # ── Z-score separator tracking ───────────────────────────────────────────
    # When zscore_threshold is set, we insert a visual separator row between
    # the last "elevated abundance" group (zscore >= threshold) and the first
    # "within expected bounds" group.  Pre-compute which groups are elevated.
    _zscore_elevated_groups = set()
    _any_elevated = False
    _any_normal = False
    if zscore_threshold is not None:
        for s in all_strains:
            sg = species_group_map[id(s)]
            gk = sg.get('toplevelkey', sg.get('key', 'unknown'))
            # Check zscore on both the group and the member; take the max
            gz = max(float(sg.get('zscore', 0) or 0),
                     float(s.get('zscore', 0) or 0))
            if gz >= zscore_threshold:
                _zscore_elevated_groups.add(gk)
        _any_elevated = bool(_zscore_elevated_groups)
        # Check if there are any normal groups too (need both sides for a separator)
        for s in all_strains:
            sg = species_group_map[id(s)]
            gk = sg.get('toplevelkey', sg.get('key', 'unknown'))
            if gk not in _zscore_elevated_groups:
                _any_normal = True
                break
    _zscore_separator_inserted = False

    # ── Flora-on-sterile separator tracking ──────────────────────────────────
    # For sterile sample types (blood, csf, serum, etc.), pre-compute which
    # groups are commensal flora so we can insert a "Potential Contaminants"
    # separator bar before the first such group.
    _flora_groups_set = set()
    _any_non_flora_grp = False
    _any_flora_grp = False
    _flora_separator_inserted = False
    _is_sterile_sampletype = bool(sampletype and sampletype.lower().strip() in _STERILE_TYPES)
    if _is_sterile_sampletype:
        for s in all_strains:
            sg = species_group_map[id(s)]
            gk = sg.get('toplevelkey', sg.get('key', 'unknown'))
            if (_is_flora_on_sterile(sg, sampletype)
                    or _is_flora_on_sterile(s, sampletype)):
                _flora_groups_set.add(gk)
        _any_flora_grp = bool(_flora_groups_set)
        for s in all_strains:
            sg = species_group_map[id(s)]
            gk = sg.get('toplevelkey', sg.get('key', 'unknown'))
            if gk not in _flora_groups_set:
                _any_non_flora_grp = True
                break

    i = 0
    while i < len(all_strains):
        strain = all_strains[i]
        species_group = species_group_map[id(strain)]
        species_key = species_group.get('toplevelkey', species_group.get('key', 'unknown'))

        # ── Species group header row ──────────────────────────────────────────
        if current_species_key != species_key:
            # ── Insert z-score separator if transitioning from elevated to normal ──
            if (_any_elevated and _any_normal
                    and not _zscore_separator_inserted
                    and species_key not in _zscore_elevated_groups):
                # Insert a visual divider row
                sep_label = (
                    f'<i>— Below: organisms within expected abundance '
                    f'(z-score &lt; {zscore_threshold}) —</i>'
                )
                sep_style = ParagraphStyle(
                    'ZScoreSep', parent=small_style, fontSize=8, leading=10,
                    alignment=TA_CENTER, fontName='Helvetica-Oblique',
                    textColor=colors.HexColor('#666666'))
                sep_row = [''] * n_total
                sep_row[1] = Paragraph(sep_label, sep_style)
                table_data.append(sep_row)
                # Span the label across all columns except the indicator
                table_styles.append(('SPAN', (1, row_idx), (n_total - 1, row_idx)))
                table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx),
                                     colors.HexColor('#F5F5F5')))
                table_styles.append(('LINEABOVE', (0, row_idx), (-1, row_idx),
                                     2.0, colors.HexColor('#999999')))
                table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx),
                                     0.5, colors.HexColor('#CCCCCC')))
                table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 6))
                table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6))
                table_styles.append(('ALIGN', (1, row_idx), (1, row_idx), 'CENTER'))
                row_idx += 1
                _zscore_separator_inserted = True

            # ── Insert flora separator if entering a potential-contaminant group ──
            if (_any_flora_grp and _any_non_flora_grp
                    and not _flora_separator_inserted
                    and species_key in _flora_groups_set):
                flora_sep_label = '<i>— Potential Contaminants —</i>'
                flora_sep_style = ParagraphStyle(
                    'FloraSep', parent=small_style, fontSize=8, leading=10,
                    alignment=TA_CENTER, fontName='Helvetica-Oblique',
                    textColor=colors.HexColor('#e67e22'))
                flora_sep_row = [''] * n_total
                flora_sep_row[1] = Paragraph(flora_sep_label, flora_sep_style)
                table_data.append(flora_sep_row)
                table_styles.append(('SPAN', (1, row_idx), (n_total - 1, row_idx)))
                table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx),
                                     colors.HexColor('#FFF8F0')))
                table_styles.append(('LINEABOVE', (0, row_idx), (-1, row_idx),
                                     2.0, colors.HexColor('#e67e22')))
                table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx),
                                     0.5, colors.HexColor('#F5CBA7')))
                table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 6))
                table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 6))
                table_styles.append(('ALIGN', (1, row_idx), (1, row_idx), 'CENTER'))
                row_idx += 1
                _flora_separator_inserted = True

            current_species_key = species_key
            emitted_subkeys_per_group[species_key] = set()

            group_name = species_group.get('toplevelname', 'Unknown')
            group_reads = species_group.get('numreads', 0)
            group_k2_reads = species_group.get('k2_reads', 0)

            group_strain_names = strains_per_group.get(species_key, [])
            n_strains = len(group_strain_names)
            n_species_rows = len((subkey_display_lookup or {}).get(species_key, {}))

            _grp_flora = _commensal_site_tag(species_group, species_group.get('normalized_sample_site', ''))
            group_name_para = Paragraph(
                f'<b><link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={species_key}" '
                f'color="blue">{group_name}</link></b>{_grp_flora}',
                group_header_style
            )

            # Group-level metrics: CDS/mmbert (group-level only)
            gm = group_mmbert_max.get(species_key, None)
            dmd_cds = group_dmnd_cds_max.get(species_key, None)
            # Pass right column content width (minus outer cell padding) to cap mini-table
            metrics_max_w = (right_w - 6) if use_subkey else None
            group_metrics_tbl = build_metrics_mini_table(dmd_cds, gm, max_width=metrics_max_w)

            # The group name cell spans columns 1+2
            spanned_name_width = col_widths[1] + col_widths[2]

            # Name cell contains anchor + outline dest + name paragraph
            # Subtract outer cell padding (LEFT=4 + RIGHT=4) so text wraps within bounds
            name_cell_w = spanned_name_width - 8
            name_cell_rows = []
            # Embed the species-group anchor and outline destination directly
            # in the table cell so bookmarks jump to this exact row.
            # Uses AbsoluteAnchorFlowable (bookmarkHorizontalAbsolute) instead
            # of AnchorFlowable (bookmarkHorizontal) so the Y coordinate is
            # correct even when nested inside a table cell across page breaks.
            if sample_name is not None:
                sp_bm = f"species_{sanitize_bookmark_name(sample_name)}_{species_key}"
                name_cell_rows.append([AbsoluteAnchorFlowable(sp_bm)])
                if outline_collector is not None:
                    name_cell_rows.append([OutlineDest(
                        f"outline_{sp_bm}", group_name,
                        level=1, closed=True, collector=outline_collector)])
            name_cell_rows.append([group_name_para])
            group_name_cell = Table(
                name_cell_rows,
                colWidths=[name_cell_w]
            )
            group_name_cell.setStyle(TableStyle([
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))

            group_row = ['', group_name_cell, '']
            if show_k2_column:
                group_row.append(Paragraph(f'<b>{group_k2_reads:,.0f}</b>', mini_style))
            group_row.append(Paragraph(f'<b>{group_reads:,.0f}</b>', mini_style))
            group_row += ['', '']
            if show_ani_column:
                group_row.append('')
            if show_control_bar:
                # Group-level spark bar using the group's control_comparison
                _grp_spark = _build_spark_bar_from_member(species_group, bar_width=int(col_widths[-2 if show_inline_table else -1] - 6) if col_widths else 72, bar_height=10)
                group_row.append(_grp_spark if _grp_spark else '')
            # Right-side detail column (inline mode only — compact mode has no extra column)
            if show_inline_table:
                # Full inline mode: show CDS/mmbert metrics (organism count is now in the spanned area)
                if group_metrics_tbl is not None:
                    right_cell_rows = [[group_metrics_tbl]]
                    right_cell = Table(right_cell_rows, colWidths=[right_w - 6])
                    right_cell.setStyle(TableStyle([
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                        ('TOPPADDING', (0, 0), (-1, -1), 0),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ]))
                    group_row.append(right_cell)
                else:
                    group_row.append('')

            table_data.append(group_row)
            group_row_indices.append(row_idx)

            # Span group name across Organism + TASS columns only
            table_styles.append(('SPAN', (1, row_idx), (2, row_idx)))
            # Span trailing columns: stop before the right detail column when it exists
            rpm_col_idx = 5 if show_k2_column else 4
            span_end = (n_base - 1) if show_inline_table else (n_total - 1)
            if rpm_col_idx <= span_end:
                table_styles.append(('SPAN', (rpm_col_idx, row_idx), (span_end, row_idx)))
            # Always place organism count summary in the RPM slot (first cell of trailing span)
            if n_species_rows:
                summary_text = (
                    f'<i>{n_species_rows} species/subkey row{"s" if n_species_rows != 1 else ""}; '
                    f'{n_strains} qualifying strain{"s" if n_strains != 1 else ""}</i>'
                )
            else:
                summary_text = f'<i>{n_strains} likely detected organism{"s" if n_strains != 1 else ""}</i>'
            group_row[rpm_col_idx] = Paragraph(summary_text, group_strain_summary_style)
            table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#E8E8E8')))
            table_styles.append(('ALIGN', (1, row_idx), (1, row_idx), 'LEFT'))
            table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 8))
            table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 8))
            row_idx += 1

        # ── Data rows ─────────────────────────────────────────────────────────
        if use_subkey:
            sk = str(strain.get('subkey', strain.get('key', 'unknown')))

            # Skip if we already emitted this subkey for this group
            if sk in emitted_subkeys_per_group[species_key]:
                i += 1
                continue
            emitted_subkeys_per_group[species_key].add(sk)

            species_entry = ((subkey_display_lookup or {}).get(species_key, {})).get(sk)
            if not species_entry:
                i += 1
                continue

            visible_strains = species_entry['visible_strains']
            if len(visible_strains) == 1:
                # Single qualifying strain — promote it directly, skip species summary row
                _append_child_strain_row(
                    visible_strains[0], promoted=True,
                    below_threshold_primary=species_entry.get('below_threshold_primary', []),
                )
            else:
                _append_species_summary_row(species_entry)
                for visible_strain in visible_strains:
                    _append_child_strain_row(visible_strain)
            i += 1

        else:
            # ── Original flat per-strain mode ──────────────────────────────────
            microbial_category = strain.get('microbial_category', 'Unknown')
            ann_class = strain.get('annClass', '')
            is_hc = strain.get('high_cons', False)
            indicator_text = '★' if is_hc else ''

            # ── Early per-row zscore check (flat mode) ────────────────────
            _row_below_zscore = False
            if zscore_threshold is not None:
                _strain_z = float(strain.get('zscore', 0) or 0)
                _row_below_zscore = _strain_z < zscore_threshold

            strain_key = strain.get('key', '')
            if _row_below_zscore:
                _n_samples = int(strain.get('hmp_num_samples', 0) or 0)
                _n_sites = int(strain.get('hmp_site_count', 0) or 0)
                if _n_sites > 0:
                    _pct = (_n_samples / _n_sites) * 100
                    _pct_str = '&lt;0.001%' if _pct < 0.001 else f'{_pct:.3f}%'
                    _sample_label = f' <font color="#999999" size="6">{_n_samples} ({_pct_str})</font>'
                else:
                    _sample_label = ''
                _zscore_sym = f' <font color="#999999">&#9671;</font>{_sample_label}'
            else:
                _zscore_sym = ''
            _flora_tag = _commensal_site_tag(strain, strain.get('normalized_sample_site', ''))
            name_html = (
                f'{strain.get("name", "Unknown")} '
                f'(<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={strain_key}" '
                f'color="blue">{strain_key}</link>){_flora_tag}{_zscore_sym}'
            )

            high_ani_text = ''
            if show_ani_column:
                # Use the pre-computed matches embedded by match_paths.py
                matches = get_high_ani_matches(strain)
                if matches:
                    ani_links = []
                    for m in matches[:3]:
                        taxid = str(m.get('key', ''))
                        ani_pct = float(m.get('ani_pct', 0))
                        if taxid in taxid_to_bookmark:
                            bm = taxid_to_bookmark[taxid]
                            if bm in valid_bookmarks:
                                ani_links.append(
                                    f'<link href="#{bm}" color="blue">{taxid} ({ani_pct:.1f}%)</link>')
                            else:
                                ani_links.append(f"{taxid} ({ani_pct:.1f}%)")
                        else:
                            ani_links.append(f"{taxid} ({ani_pct:.1f}%)")
                    high_ani_text = Paragraph(", ".join(ani_links), ani_style)
                else:
                    high_ani_text = Paragraph("-", ani_style)

            strain_reads = float(strain.get('numreads', 0) or 0)
            pct = (strain_reads / sample_total_reads * 100.0) if sample_total_reads else 0.0
            rpm = strain.get('rpm', 0) or 0

            row_name_style = strain_name_style_small if _row_below_zscore else strain_name_style
            row_data_style = data_style_small if _row_below_zscore else data_style
            row_ani_style = ani_style_small if _row_below_zscore else ani_style

            row = [
                Paragraph(indicator_text, indicator_para_style) if indicator_text else '',
                Paragraph(name_html, row_name_style),
                Paragraph(f"{strain.get('tass_score', 0)*100:.1f}", row_data_style),
            ]
            if show_k2_column:
                row.append(Paragraph(f"{strain.get('k2_reads', 0):,.0f}", row_data_style))
            row.append(Paragraph(f"{strain_reads:,.0f} ({pct:.1f}%)", row_data_style))
            row.append(Paragraph(f"{rpm:,.0f}", row_data_style))
            row.append(Paragraph(f"{min(100, strain.get('coverage', 0)*100):.1f}%", row_data_style))
            if show_ani_column:
                if isinstance(high_ani_text, Paragraph) and _row_below_zscore:
                    high_ani_text = Paragraph(high_ani_text.text, row_ani_style)
                row.append(high_ani_text)
            if show_control_bar:
                _spark = _build_spark_bar_from_member(strain, bar_width=int(col_widths[-1] - 6) if col_widths else 72, bar_height=10)
                row.append(_spark if _spark else Paragraph('-', row_data_style))

            table_data.append(row)

            print(f"  Row {row_idx}: {strain.get('name', 'Unknown')[:40]} "
                  f"- Category: {microbial_category}, Class: {ann_class}")

            # ── Per-row zscore opacity (flat mode) ────────────────────────
            # _row_below_zscore was computed earlier (before name HTML building).
            # ── Flora-on-sterile fade (flat mode) ─────────────────────────
            _flora_fade = _is_flora_on_sterile(strain, strain.get('normalized_sample_site', ''))
            _flora_mult = 0.70 if _flora_fade else 1.0
            if _row_below_zscore:
                ind_color = get_category_color(microbial_category, ann_class, alpha=0.35 * _flora_mult)
                row_color = get_category_color(microbial_category, ann_class, alpha=0.05 * _flora_mult)
            else:
                ind_color = get_category_color(microbial_category, ann_class, alpha=1.0 * _flora_mult)
                row_color = get_category_color(microbial_category, ann_class, alpha=0.15 * _flora_mult)
            table_styles.append(('BACKGROUND', (0, row_idx), (0, row_idx), ind_color))
            table_styles.append(('BACKGROUND', (1, row_idx), (-1, row_idx), row_color))
            if _row_below_zscore:
                table_styles.append(('TEXTCOLOR', (1, row_idx), (-1, row_idx),
                                     colors.Color(0.6, 0.6, 0.6, 1)))
                table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 3))
                table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 3))
                table_styles.append(('FONTSIZE', (1, row_idx), (-1, row_idx), 7))
            # Horizontal separator below each strain row with padding
            table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx), 1.5, colors.HexColor('#CCCCCC')))
            if not _row_below_zscore:
                table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 8))
                table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 6))
            row_idx += 1
            i += 1

    # ── Missing positive control rows (tiny, faded) ─────────────────────────
    _miss_pos = missing_pos_controls or []
    if _miss_pos:
        _miss_style = ParagraphStyle(
            'MissingCtrl', parent=small_style,
            fontSize=6.5, leading=8, textColor=colors.Color(0.5, 0.3, 0.1))
        _miss_data_style = ParagraphStyle(
            'MissingCtrlData', parent=small_style,
            fontSize=6, leading=7, textColor=colors.Color(0.55, 0.55, 0.55))

        # Separator row labelling the section
        _sep_label = Paragraph(
            '<font size="6"><b>Not found but in the positive control</b></font>',
            ParagraphStyle('MissSep', parent=small_style, fontSize=6,
                           leading=7, textColor=colors.Color(0.45, 0.25, 0.1)))
        _sep_row = [''] * n_total
        _sep_row[1] = _sep_label
        table_data.append(_sep_row)
        table_styles.append(('SPAN', (1, row_idx), (n_total - 1, row_idx)))
        table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx),
                             colors.Color(0.97, 0.93, 0.87, 1)))
        table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 2))
        table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 2))
        row_idx += 1

        for mp in _miss_pos:
            _mname = mp.get('name') or mp.get('id') or '?'
            _mlevel = mp.get('level', 'toplevelkey')
            _mtass = mp.get('pos_tass_score', 0)
            _mreads = mp.get('pos_numreads', 0)
            _msrc = mp.get('source', '')

            _name_html = (
                f'<font color="#8B4513"><i>{_mname}</i></font>'
                f'&nbsp;<font size="5" color="#999999">'
                f'Control TASS {_mtass:.2f}, {int(_mreads):,} reads'
                f'</font>'
            )
            miss_row = [''] * n_total
            miss_row[0] = ''  # no category indicator
            miss_row[1] = Paragraph(_name_html, _miss_style)
            # Span organism name across all data columns
            table_data.append(miss_row)
            table_styles.append(('SPAN', (1, row_idx), (n_total - 1, row_idx)))
            table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx),
                                 colors.Color(0.98, 0.96, 0.92, 1)))
            table_styles.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 1))
            table_styles.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 1))
            table_styles.append(('LINEBELOW', (0, row_idx), (-1, row_idx), 0.5,
                                 colors.Color(0.85, 0.85, 0.85)))
            row_idx += 1

    # ── Base table style ──────────────────────────────────────────────────────
    base_ts = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('LEFTPADDING', (0, 0), (-1, 0), 3),
        ('RIGHTPADDING', (0, 0), (-1, 0), 3),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('VALIGN', (0, 1), (0, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 1), (0, -1), 1),
        ('RIGHTPADDING', (0, 1), (0, -1), 1),
        ('TOPPADDING', (0, 1), (0, -1), 2),
        ('BOTTOMPADDING', (0, 1), (0, -1), 2),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (1, 1), (1, -1), 'MIDDLE'),
        ('VALIGN', (2, 1), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        # Organism column (col 1): keep comfortable padding
        ('LEFTPADDING', (1, 1), (1, -1), 4),
        ('RIGHTPADDING', (1, 1), (1, -1), 4),
        ('TOPPADDING', (1, 1), (1, -1), 6),
        ('BOTTOMPADDING', (1, 1), (1, -1), 6),
        # Data columns (col 2 onwards): tight padding to prevent overflow
        ('LEFTPADDING', (2, 1), (n_base - 1, -1), 2),
        ('RIGHTPADDING', (2, 1), (n_base - 1, -1), 2),
        ('TOPPADDING', (2, 1), (n_base - 1, -1), 4),
        ('BOTTOMPADDING', (2, 1), (n_base - 1, -1), 4),
    ]

    if show_inline_table:
        # Right detail column styles + vertical delimiter (inline mode only)
        right_col_ts = [
            ('ALIGN', (n_base, 1), (n_total - 1, -1), 'LEFT'),
            ('VALIGN', (n_base, 1), (n_total - 1, -1), 'TOP'),
            ('LEFTPADDING', (n_base, 1), (n_total - 1, -1), 4),
            ('TOPPADDING', (n_base, 1), (n_total - 1, -1), 2),
            ('RIGHTPADDING', (n_base, 1), (n_total - 1, -1), 2),
            ('BOTTOMPADDING', (n_base, 1), (n_total - 1, -1), 2),
            ('LINEAFTER', (n_base - 1, 0), (n_base - 1, -1), 2, colors.HexColor('#3498DB')),
        ]
        base_ts += right_col_ts

    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle(base_ts + table_styles))
    return table


def create_low_confidence_table(low_confidence_strains, small_style, show_k2_column,
                                 available_width=None):
    strain_name_style = ParagraphStyle(
        'StrainName', parent=small_style, fontSize=8, leading=10)
    data_style = ParagraphStyle(
        'DataStyle', parent=small_style, fontSize=7, leading=9)

    headers = (['Sample', 'Organism', 'TASS', 'K2 Reads', 'Reads']
               if show_k2_column else ['Sample', 'Organism', 'TASS', 'Reads'])
    table_data = [headers]

    for item in low_confidence_strains:
        strain = item['strain']
        sample_name = item['sample_name']

        strain_name_text = strain.get('name', 'Unknown')
        strain_key = strain.get('key', '')
        if strain_key:
            strain_name_text = (
                f'{strain_name_text} '
                f'(<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={strain_key}" '
                f'color="blue">{strain_key}</link>)'
            )
        if strain.get('high_cons', False):
            strain_name_text = f"★ {strain_name_text}"

        row = [
            Paragraph(sample_name, strain_name_style),
            Paragraph(strain_name_text, strain_name_style),
            Paragraph(f"{strain.get('tass_score', 0)*100:.1f}", data_style),
        ]
        if show_k2_column:
            row.append(Paragraph(f"{strain.get('k2_reads', 0):,.0f}", data_style))
        row.append(Paragraph(f"{strain.get('numreads', 0):,.0f}", data_style))
        table_data.append(row)

    if available_width is None:
        available_width = 8.5*inch - 0.02*8.5*inch

    if show_k2_column:
        col_widths = [available_width*0.18, available_width*0.50,
                      available_width*0.10, available_width*0.10, available_width*0.12]
    else:
        col_widths = [available_width*0.20, available_width*0.56,
                      available_width*0.11, available_width*0.13]

    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 1), (-1, -1), 6),
        ('RIGHTPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]
    for idx, item in enumerate(low_confidence_strains):
        row_color = get_category_color(
            item['strain'].get('microbial_category', 'Unknown'),
            item['strain'].get('annClass', ''),
            alpha=0.25
        )
        table_styles.append(('BACKGROUND', (0, idx+1), (-1, idx+1), row_color))
    table.setStyle(TableStyle(table_styles))
    return table


def create_strain_detail_tables(samples_dict, sorted_groups_by_sample,
                                small_style, show_k2_column, valid_bookmarks,
                                args, available_width, outline_collector=None):
    """
    Build the strain-detail appendix story elements.

    Returns a list of ReportLab flowables to be appended to the main story.
    One flat table per sample; each row is a single strain member.
    Rows are compact (small font, tight padding) to save space.
    Each species-group block in the table is preceded by an AnchorFlowable
    so that links from the main table can jump directly to it.
    """
    story_items = []

    # ── Section heading ───────────────────────────────────────────────────────
    heading_style = ParagraphStyle(
        'StrainDetailHeading', parent=small_style, fontSize=14,
        textColor=colors.HexColor('#34495E'), spaceBefore=10, spaceAfter=4,
        fontName='Helvetica-Bold')
    note_style = ParagraphStyle(
        'StrainDetailNote', parent=small_style, fontSize=8, leading=10,
        textColor=colors.HexColor('#555555'))
    sample_heading_style = ParagraphStyle(
        'StrainSampleHeading', parent=small_style, fontSize=11,
        fontName='Helvetica-Bold', textColor=colors.HexColor('#2C3E50'),
        spaceBefore=8, spaceAfter=2)
    cell_style = ParagraphStyle(
        'StrainCell', parent=small_style, fontSize=7, leading=8, wordWrap='CJK')
    header_cell_style = ParagraphStyle(
        'StrainHeaderCell', parent=small_style, fontSize=7, leading=8,
        fontName='Helvetica-Bold', textColor=colors.whitesmoke,
        alignment=TA_CENTER, wordWrap='CJK')

    story_items.append(AnchorFlowable('strain_detail_table'))
    story_items.append(OutlineDest('outline_strain_detail', 'Strain Detail', level=0, closed=True, collector=outline_collector))
    story_items.append(Paragraph('<b>Additional Lower-Level Strain Detail</b>', heading_style))
    story_items.append(Paragraph(
        'This appendix lists additional lower-level strains, subspecies, and assemblies '
        'detected beneath each species-level entry in the main table above. '
        'Only species groups with one or more qualifying lower-level strain members are shown here. '
        'Links labeled "Strain detail" jump between the main table and this '
        'appendix when pathogenic strain-level entries are present.',
        note_style))
    story_items.append(Spacer(1, 0.06 * inch))

    for sample_name, sorted_groups in sorted_groups_by_sample:
        # Pre-scan: does this sample have any strains with CDS or mmbert?
        sample_has_cds = False
        sample_has_mmbert = False
        sample_strain_rows = []  # list of (sg, strain) pairs to render
        _mc = args._sample_min_conf.get(sample_name, _DEFAULT_CONF)

        for sg in sorted_groups:
            species_key = sg.get('toplevelkey', sg.get('key', 'unknown'))
            qualifying = get_qualifying_strains(sg, args, _mc, min_reads=args.min_reads)
            # Mirror the main table's right-column visibility logic exactly.
            # A group belongs in the appendix only when it has genuinely
            # additional lower-level strains beyond what is already shown in
            # the main table row.
            #
            # Two cases to exclude:
            #   1. Every member's key == its own subkey — nothing sub-level.
            #   2. Single-strain name-switch: exactly one qualifying member
            #      whose key != subkey.  In this case the main table row has
            #      already switched its display name to that strain, so there
            #      is nothing extra to list in the appendix.
            has_sublevel_members = any(
                str(m.get('key', '')) != str(m.get('subkey', m.get('key', '')))
                for m in qualifying
            )
            if not has_sublevel_members:
                continue
            # Detect single-strain name-switch (mirrors single_strain_in_group
            # logic from create_combined_sample_table): only one visible entry
            # in the entire species group and its key differs from its subkey.
            if len(qualifying) == 1 and str(qualifying[0].get('key', '')) != str(qualifying[0].get('subkey', qualifying[0].get('key', ''))):
                continue
            # Only append entries that are genuine sub-level strains (key != subkey).
            # entries where key == subkey are already shown as their own species row
            # in the main table and must not appear again in the appendix.
            sublevel = [
                m for m in qualifying
                if str(m.get('key', '')) != str(m.get('subkey', m.get('key', '')))
            ]
            for strain in sublevel:
                sample_strain_rows.append((sg, strain))
                cds_v = _valid_num((strain.get('diamond') or {}).get('cds', None))
                mm_v = _valid_num(strain.get('mmbert', None))
                if cds_v is not None:
                    sample_has_cds = True
                if mm_v is not None:
                    sample_has_mmbert = True

        if not sample_strain_rows:
            continue

        # ── Sample heading with back-link ─────────────────────────────────────
        strain_sample_bm = f"strain_sample_{sanitize_bookmark_name(sample_name)}"
        main_sample_bm = f"sample_{sanitize_bookmark_name(sample_name)}"
        back_link = create_safe_link('\u2191', main_sample_bm, valid_bookmarks, color='blue')
        _strain_meta = getattr(args, '_input_metadata', {}).get(sample_name, {})
        _strain_plat = _strain_meta.get('platform', '')
        _strain_plat_str = f" — {_strain_plat}" if _strain_plat and _strain_plat != 'unknown' else ''
        story_items.append(AnchorFlowable(strain_sample_bm))
        story_items.append(Paragraph(
            f'<b>{sample_name}{_strain_plat_str}</b> {back_link}', sample_heading_style))

        # ── Column headers ────────────────────────────────────────────────────
        col_headers = [
            '',  # indicator
            Paragraph('Organism', header_cell_style),
            Paragraph('TASS', header_cell_style),
        ]
        if show_k2_column:
            col_headers.append(Paragraph('K2<br/>Reads', header_cell_style))
        col_headers += [
            Paragraph('Reads', header_cell_style),
            Paragraph('Cov.', header_cell_style),
        ]
        if sample_has_cds:
            col_headers.append(Paragraph('CDS', header_cell_style))
        if sample_has_mmbert:
            col_headers.append(Paragraph('mmbert%', header_cell_style))

        # ── Column widths ─────────────────────────────────────────────────────
        n_extra = (1 if sample_has_cds else 0) + (1 if sample_has_mmbert else 0)
        n_k2 = 1 if show_k2_column else 0
        # Distribute: indicator(2%), organism(rest), TASS(7%), K2(8%), Reads(12%), Cov(7%), CDS(7%), mmbert(8%)
        fixed_pcts = 0.02 + 0.07 + n_k2 * 0.08 + 0.12 + 0.07 + n_extra * 0.075
        org_pct = max(0.25, 1.0 - fixed_pcts)

        w = available_width
        cw = [w * 0.02, w * org_pct, w * 0.07]
        if show_k2_column:
            cw.append(w * 0.08)
        cw += [w * 0.12, w * 0.07]
        if sample_has_cds:
            cw.append(w * 0.075)
        if sample_has_mmbert:
            cw.append(w * 0.075)

        # ── Build table data ──────────────────────────────────────────────────
        table_data = [col_headers]
        table_styles_det = []
        row_idx = 1

        emitted_species = set()

        for sg, strain in sample_strain_rows:
            species_key = sg.get('toplevelkey', sg.get('key', 'unknown'))
            species_name = sg.get('toplevelname', '')

            # Place per-species anchor before the first strain row for that group
            if species_key not in emitted_species:
                emitted_species.add(species_key)
                strain_row_bm = f"strain_row_{sanitize_bookmark_name(sample_name)}_{species_key}"
                story_items.append(AnchorFlowable(strain_row_bm))

            microbial_category = strain.get('microbial_category', 'Unknown')
            ann_class = strain.get('annClass', '')
            is_hc = strain.get('high_cons', False)
            indicator = '★' if is_hc else ''

            strain_key = strain.get('key', '')
            strain_name = strain.get('name', 'Unknown')

            # Back-link ↑ to the species row in the main table
            species_bm = f"species_{sanitize_bookmark_name(sample_name)}_{species_key}"
            up_link = create_safe_link('Main detail \u2191', species_bm,
                                       valid_bookmarks, color='blue')

            name_html = (
                f'<link href="https://www.ncbi.nlm.nih.gov/taxonomy/?term={strain_key}" '
                f'color="blue">{strain_name}</link> {up_link}'
            )

            strain_reads = float(strain.get('numreads', 0) or 0)
            # Use total reads from metadata (includes unaligned) when available
            _det_smeta = getattr(args, '_input_metadata', {}).get(sample_name, {})
            _det_meta_total = _det_smeta.get('total_reads')
            sample_total = max(1, int(_det_meta_total) if _det_meta_total else sum(
                sg2.get('numreads', 0)
                for sg2 in samples_dict.get(sample_name, [])
            ))
            pct = strain_reads / sample_total * 100.0

            tass_val = strain.get('tass_score', 0) * 100

            row = [
                indicator,
                Paragraph(name_html, cell_style),
                Paragraph(f'{tass_val:.1f}', cell_style),
            ]
            if show_k2_column:
                row.append(Paragraph(f"{strain.get('k2_reads', 0):,.0f}", cell_style))
            row += [
                Paragraph(f'{strain_reads:,.0f} ({pct:.1f}%)', cell_style),
                Paragraph(f'{min(100, strain.get("coverage", 0) * 100):.1f}%', cell_style),
            ]
            if sample_has_cds:
                cds_v = _valid_num((strain.get('diamond') or {}).get('cds', None))
                row.append(Paragraph(f'{cds_v:.0f}' if cds_v is not None else '—', cell_style))
            if sample_has_mmbert:
                mm_v = _valid_num(strain.get('mmbert', None))
                row.append(Paragraph(
                    f'{mm_v * 100:.2f}%' if mm_v is not None else '—', cell_style))

            table_data.append(row)

            # Row styling
            ind_color = get_category_color(microbial_category, ann_class, alpha=1.0)
            row_color = get_category_color(microbial_category, ann_class, alpha=0.12)
            table_styles_det.append(('BACKGROUND', (0, row_idx), (0, row_idx), ind_color))
            table_styles_det.append(('BACKGROUND', (1, row_idx), (-1, row_idx), row_color))
            table_styles_det.append(('LINEBELOW', (0, row_idx), (-1, row_idx),
                                     0.5, colors.HexColor('#DDDDDD')))
            table_styles_det.append(('TOPPADDING', (0, row_idx), (-1, row_idx), 2))
            table_styles_det.append(('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 2))
            row_idx += 1

        # ── Base table style ──────────────────────────────────────────────────
        base_ts = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('TOPPADDING', (0, 0), (-1, 0), 3),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
            ('LEFTPADDING', (0, 0), (-1, 0), 2),
            ('RIGHTPADDING', (0, 0), (-1, 0), 2),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#AAAAAA')),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (1, 1), (1, -1), 3),
            ('RIGHTPADDING', (1, 1), (1, -1), 3),
            ('LEFTPADDING', (2, 1), (-1, -1), 2),
            ('RIGHTPADDING', (2, 1), (-1, -1), 2),
            ('FONTSIZE', (0, 1), (0, -1), 11),  # indicator column larger for ★
        ]

        det_table = Table(table_data, repeatRows=1, colWidths=cw)
        det_table.setStyle(TableStyle(base_ts + table_styles_det))
        story_items.append(det_table)
        story_items.append(Spacer(1, 0.08 * inch))

    return story_items


def _collect_protein_annotations(species_groups):
    """
    Walk the 3-level organism hierarchy and collect all protein_annotations
    into a flat list of dicts.

    Key design points
    -----------------
    * Each row is tagged with ``_organism_name`` (the *detected* genus/species
      name from the sample) and ``_toplevel_name`` (the genus-level name).
      Downstream visualisations must use ``_organism_name`` / ``_toplevel_name``
      for grouping — NOT the ``genus`` field that comes from the reference
      database metadata, which reflects the source organism of the matched
      accession (e.g. a Salmonella CARD entry matched to an Escherichia contig
      would still carry genus='Salmonella').
    * Deduplication key now uses ``(gene_name or sseqid, taxid_scope)`` so that
      accessions with no gene name (e.g. TCDB transporters with only an sseqid)
      are not collapsed into a single placeholder per taxon.
    """
    rows = []
    seen = set()
    for sg in species_groups:
        tlk     = str(sg.get('toplevelkey', sg.get('key', '')))
        tl_name = sg.get('toplevelname', sg.get('name', ''))
        # genus-level annotations
        for annot in sg.get('protein_annotations_genus', []):
            gene      = annot.get('gene_name') or annot.get('sseqid', '')
            dedup_key = (gene, tlk)
            if dedup_key not in seen:
                seen.add(dedup_key)
                rows.append({
                    **annot,
                    '_organism_name': tl_name,
                    '_toplevel_name': tl_name,
                    '_taxid':         tlk,
                    '_level':         'genus',
                })
        # species/subkey level
        for sk_m in sg.get('members', []):
            sk      = str(sk_m.get('subkey', sk_m.get('key', '')))
            sk_name = sk_m.get('subkeyname', sk_m.get('name', ''))
            for annot in sk_m.get('protein_annotations', []):
                gene      = annot.get('gene_name') or annot.get('sseqid', '')
                dedup_key = (gene, sk)
                if dedup_key not in seen:
                    seen.add(dedup_key)
                    rows.append({
                        **annot,
                        '_organism_name': sk_name,
                        '_toplevel_name': tl_name,
                        '_taxid':         sk,
                        '_level':         'species',
                    })
            # strain level
            for strain in sk_m.get('members', []):
                for annot in strain.get('protein_annotations', []):
                    gene      = annot.get('gene_name') or annot.get('sseqid', '')
                    s_key     = str(strain.get('key', ''))
                    dedup_key = (gene, sk)   # deduplicate at species level
                    if dedup_key not in seen:
                        seen.add(dedup_key)
                        rows.append({
                            **annot,
                            '_organism_name': strain.get('name', sk_name),
                            '_toplevel_name': tl_name,
                            '_taxid':         s_key,
                            '_level':         'strain',
                        })
    return rows


def _filter_annot_rows_by_pident(rows, min_pident):
    """Return only rows where pident >= min_pident (handles str/float/NaN)."""
    filtered = []
    for r in rows:
        try:
            pident_val = float(r.get('pident', 0) or 0)
        except (ValueError, TypeError):
            pident_val = 0.0
        if pident_val >= min_pident:
            filtered.append(r)
    return filtered


def create_protein_annotation_table(species_groups, small_style, available_width, min_pident=0.0):
    """
    Build a ReportLab Table collapsed to (genus, property) rows.
    Each row lists all unique gene names comma-separated and shows
    avg/median e-value across all hits in that bucket.
    Rows with pident < min_pident are excluded when min_pident > 0.
    Returns None if no annotations are present.
    """
    import statistics

    rows = _collect_protein_annotations(species_groups)
    if min_pident > 0:
        rows = _filter_annot_rows_by_pident(rows, min_pident)
    if not rows:
        return None

    # ── Collapse to (genus, property) buckets ────────────────────────────
    # Each bucket accumulates: genes (set), evalues (list), best pident
    buckets = {}  # (genus, property) -> {'genes': set, 'evalues': [], 'best_pident': float}
    for r in rows:
        genus = str(r.get('genus', '') or r.get('_organism_name', '') or '').strip()
        prop  = str(r.get('property', '') or '').strip()
        if not genus or not prop:
            continue
        key = (genus, prop)
        if key not in buckets:
            buckets[key] = {'genes': set(), 'evalues': [], 'best_pident': 0.0}
        gene = str(r.get('gene_name', '') or r.get('sseqid', '') or '').strip()
        if gene:
            buckets[key]['genes'].add(gene)
        try:
            ev = float(r.get('evalue') or 0)
            buckets[key]['evalues'].append(ev)
        except (ValueError, TypeError):
            pass
        try:
            pid = float(r.get('pident') or 0)
            if pid > buckets[key]['best_pident']:
                buckets[key]['best_pident'] = pid
        except (ValueError, TypeError):
            pass

    if not buckets:
        return None

    # Sort by genus then property
    sorted_keys = sorted(buckets.keys(), key=lambda k: (k[0], k[1]))

    cell_font   = ParagraphStyle('AnnotCell',   parent=small_style, fontSize=6.5, leading=8)
    header_style = ParagraphStyle('AnnotHeader', parent=small_style, fontSize=7,   leading=9,
                                   fontName='Helvetica-Bold', textColor=colors.whitesmoke)

    header = [
        Paragraph('Genus',      header_style),
        Paragraph('Property',   header_style),
        Paragraph('Genes',      header_style),
        Paragraph('# Hits',     header_style),
        Paragraph('Best %id',   header_style),
        Paragraph('Avg E-val',  header_style),
        Paragraph('Med E-val',  header_style),
    ]
    table_data = [header]

    for key in sorted_keys:
        genus, prop = key
        b = buckets[key]
        genes_str = ', '.join(sorted(b['genes'])) if b['genes'] else '—'
        n_hits    = len(b['evalues'])
        best_pid  = f"{b['best_pident']:.1f}" if b['best_pident'] else ''
        if b['evalues']:
            avg_ev = statistics.mean(b['evalues'])
            med_ev = statistics.median(b['evalues'])
            avg_str = f"{avg_ev:.2e}"
            med_str = f"{med_ev:.2e}"
        else:
            avg_str = med_str = ''

        table_data.append([
            Paragraph(genus,    cell_font),
            Paragraph(prop,     cell_font),
            Paragraph(genes_str, cell_font),
            Paragraph(str(n_hits), cell_font),
            Paragraph(best_pid, cell_font),
            Paragraph(avg_str,  cell_font),
            Paragraph(med_str,  cell_font),
        ])

    col_widths = [
        available_width * 0.13,  # Genus
        available_width * 0.16,  # Property
        available_width * 0.42,  # Genes
        available_width * 0.06,  # # Hits
        available_width * 0.07,  # Best %id
        available_width * 0.08,  # Avg E-val
        available_width * 0.08,  # Med E-val
    ]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND',    (0, 0), (-1,  0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR',     (0, 0), (-1,  0), colors.whitesmoke),
        ('FONTNAME',      (0, 0), (-1,  0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1,  0), 7),
        ('BOTTOMPADDING', (0, 0), (-1,  0), 4),
        ('TOPPADDING',    (0, 0), (-1,  0), 4),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE',      (0, 1), (-1, -1), 6.5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
        ('TOPPADDING',    (0, 1), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
    ]
    # Alternate shading + green separator line between genera
    prev_genus = None
    for i, key in enumerate(sorted_keys, start=1):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.Color(0.95, 0.97, 0.95)))
        if prev_genus is not None and key[0] != prev_genus:
            style_cmds.append(('LINEABOVE', (0, i), (-1, i), 0.8, colors.HexColor('#2E7D32')))
        prev_genus = key[0]

    tbl.setStyle(TableStyle(style_cmds))
    return tbl


def create_protein_annotation_heatmap(species_groups, available_width, min_pident=90.0):
    """
    Build a heatmap (as a ReportLab Image flowable) showing protein annotation
    hit counts per (organism × classification) cell, filtered to rows where
    pident >= min_pident.

    X-axis: unique classification values.
    Y-axis: unique organism names (from the 'organism' field in each annotation).
    Cell value: number of distinct gene hits in that (organism, classification) bucket.

    Returns None if no rows survive the pident filter.
    """
    rows = _collect_protein_annotations(species_groups)
    if not rows:
        return None

    # ── Filter by pident ────────────────────────────────────────────────────
    filtered = _filter_annot_rows_by_pident(rows, min_pident)

    if not filtered:
        return None

    # ── Build pivot counts ──────────────────────────────────────────────────
    # X-axis: genus; Y-axis: property (high-level category — Drug Target,
    # Virulence Factor, Antibiotic Resistance, Transporter, etc.) which
    # is far fewer rows than the raw classification field.
    counts = {}  # (genus, property) -> set of gene_name
    for r in filtered:
        org = str(r.get('genus', '') or r.get('_organism_name', '') or '').strip()
        prop = str(r.get('property', '') or '').strip()
        gene = str(r.get('gene_name', '') or r.get('sseqid', '') or '').strip()
        if not org or not prop:
            continue
        counts.setdefault((org, prop), set()).add(gene)

    if not counts:
        return None

    # Convert to count integers
    count_values = {k: len(v) for k, v in counts.items()}

    # X-axis: genus; Y-axis: property (high-level grouping)
    genera = sorted({k[0] for k in count_values})
    properties = sorted({k[1] for k in count_values})

    # Build matrix (rows=properties, cols=genera)
    matrix = np.zeros((len(properties), len(genera)), dtype=float)
    gen_idx = {g: i for i, g in enumerate(genera)}
    prop_idx = {p: i for i, p in enumerate(properties)}
    for (org, prop), cnt in count_values.items():
        matrix[prop_idx[prop], gen_idx[org]] = cnt

    # ── Draw with matplotlib ────────────────────────────────────────────────
    n_rows = len(properties)
    n_cols = len(genera)

    cell_h = 0.45          # inches per row
    cell_w = max(1.2, available_width / 72 / max(n_cols, 1))  # pts→inches
    fig_h = max(3, n_rows * cell_h + 1.5)
    fig_w = min(available_width / 72, n_cols * cell_w + 2.5)  # cap at page width

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    cmap = plt.cm.YlOrRd
    cmap.set_under('white')
    im = ax.imshow(matrix, aspect='auto', cmap=cmap, vmin=0.5,
                   vmax=max(1, matrix.max()))

    ax.set_xticks(range(n_cols))
    ax.set_xticklabels(genera, rotation=45, ha='right', fontsize=6.5)
    ax.set_yticks(range(n_rows))
    ax.set_yticklabels(properties, fontsize=7)
    ax.set_xlabel('Genus', fontsize=8)
    ax.set_ylabel('Property', fontsize=8)
    ax.set_title(f'Protein Annotation Hits by Genus (pident ≥ {min_pident:.0f}%)', fontsize=9)

    # Add gridlines
    ax.set_xticks(np.arange(-0.5, n_cols, 1), minor=True)
    ax.set_yticks(np.arange(-0.5, n_rows, 1), minor=True)
    ax.grid(which='minor', color='white', linewidth=0.5, alpha=0.25)
    ax.tick_params(which='minor', bottom=False, left=False)

    # Annotate cells with count
    for i in range(n_rows):
        for j in range(n_cols):
            val = int(matrix[i, j])
            if val > 0:
                ax.text(j, i, str(val), ha='center', va='center',
                        fontsize=6, color='black' if matrix[i, j] < matrix.max() * 0.7 else 'white')

    cbar = fig.colorbar(im, ax=ax, shrink=0.6, pad=0.02)
    cbar.set_label('# Gene Hits', fontsize=7)
    cbar.ax.tick_params(labelsize=6)

    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)

    # Scale to fit available page width while respecting aspect ratio
    img_w_pt = available_width
    aspect = fig_h / fig_w
    img_h_pt = img_w_pt * aspect

    # Cap height to 90% of the usable page body (letter minus margins)
    max_img_h_pt = (11 * 72) - (2 * 0.5 * 72) - (2 * 0.5 * 72)  # ~648 pt
    if img_h_pt > max_img_h_pt:
        scale = max_img_h_pt / img_h_pt
        img_h_pt = max_img_h_pt
        img_w_pt = img_w_pt * scale

    return Image(buf, width=img_w_pt, height=img_h_pt)


# ── Property category display config (shared by bubble chart + genus cards) ──
_PROP_DISPLAY_ORDER = [
    'Antibiotic Resistance',
    'Virulence Factor',
    'Drug Target',
    'Transporter',
]
_PROP_LABELS = {
    'Antibiotic Resistance': 'AMR',
    'Virulence Factor':      'Virulence',
    'Drug Target':           'Drug Target',
    'Transporter':           'Transporter',
}
_PROP_MPL_COLORS = {
    'Antibiotic Resistance': '#C62828',
    'Virulence Factor':      '#E65100',
    'Drug Target':           '#1565C0',
    'Transporter':           '#2E7D32',
}
_PROP_RL_BADGE = {
    'Antibiotic Resistance': '#C62828',
    'Virulence Factor':      '#E65100',
    'Drug Target':           '#1565C0',
    'Transporter':           '#2E7D32',
}
_PROP_RL_ROW_BG = {
    'Antibiotic Resistance': '#FFEBEE',
    'Virulence Factor':      '#FFF3E0',
    'Drug Target':           '#E3F2FD',
    'Transporter':           '#E8F5E9',
}


def create_annotation_bubble_chart(species_groups, available_width, min_pident=0.0):
    """
    Bubble chart: Y = detected genus, X = property category, size ∝ unique gene count.

    Grouping fix: uses ``_organism_name`` (the genus detected in this sample)
    rather than the annotation's ``genus`` field (the reference-database source
    organism), so cross-genus reference matches are attributed correctly.

    Sizing fix: bubble radius is capped at a fixed maximum (28 pt) so the chart
    does not over-zoom when only one or two genera are present.
    """
    rows = _collect_protein_annotations(species_groups)
    if min_pident > 0:
        rows = _filter_annot_rows_by_pident(rows, min_pident)
    if not rows:
        return None

    # ── Collect (detected_genus, property) → gene set ───────────────────────
    # Group at the genus (toplevel) level so species/strain sub-rows don't
    # create separate duplicate entries.  Priority: _toplevel_name (always the
    # genus detected in the sample) > _organism_name > annotation genus field.
    buckets = {}
    for r in rows:
        genus = (
            str(r.get('_toplevel_name') or r.get('_organism_name') or
                r.get('genus', '') or '').strip()
        )
        prop  = str(r.get('property', '') or '').strip()
        gene  = str(r.get('gene_name', '') or r.get('sseqid', '') or '').strip()
        if not genus or not prop:
            continue
        buckets.setdefault((genus, prop), set()).add(gene or '—')

    if not buckets:
        return None

    # ── Axes ────────────────────────────────────────────────────────────────
    genus_totals = {}
    for (genus, prop), genes in buckets.items():
        genus_totals[genus] = genus_totals.get(genus, 0) + len(genes)
    genera = sorted(genus_totals.keys(), key=lambda g: -genus_totals[g])

    props_present = {k[1] for k in buckets}
    properties = [p for p in _PROP_DISPLAY_ORDER if p in props_present]
    for p in sorted(props_present):
        if p not in properties:
            properties.append(p)

    n_genera = len(genera)
    n_props  = len(properties)
    gen_idx  = {g: i for i, g in enumerate(genera)}
    prop_idx = {p: i for i, p in enumerate(properties)}

    # ── Figure geometry ─────────────────────────────────────────────────────
    # Use a fixed cell size so the chart doesn't balloon when there are few
    # genera or properties.
    CELL_W_IN  = 1.6   # inches per column — constant regardless of page width
    CELL_H_IN  = 1.05  # inches per row (room for bubble + gene annotation text)
    LEFT_PAD   = 1.4   # inches for y-axis labels
    RIGHT_PAD  = 0.3
    TOP_PAD    = 0.7
    BOT_PAD    = 0.6

    fig_w = min(available_width / 72,
                n_props * CELL_W_IN + LEFT_PAD + RIGHT_PAD)
    fig_h = n_genera * CELL_H_IN + TOP_PAD + BOT_PAD
    fig_h = max(2.8, min(fig_h, (11 * 72 - 4 * 0.5 * 72) / 72))  # cap at page

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(-0.6, n_props - 0.4)
    ax.set_ylim(-0.6, n_genera - 0.4)
    ax.set_facecolor('#FAFAFA')
    fig.patch.set_facecolor('white')

    for xi in range(n_props):
        ax.axvline(xi, color='#E0E0E0', linewidth=0.5, zorder=0)
    for yi in range(n_genera):
        ax.axhline(yi, color='#E0E0E0', linewidth=0.5, zorder=0)

    # ── Bubble sizing — hard cap prevents over-zoom with few data points ─────
    MAX_RADIUS_PT = 28.0          # absolute cap in matplotlib scatter points
    MIN_RADIUS_PT = 9.0
    max_count = max(len(v) for v in buckets.values())

    for (genus, prop), genes in buckets.items():
        xi = prop_idx[prop]
        yi = gen_idx[genus]
        count = len(genes)
        radius = MIN_RADIUS_PT + (MAX_RADIUS_PT - MIN_RADIUS_PT) * (count / max_count)
        area   = np.pi * radius ** 2
        color  = _PROP_MPL_COLORS.get(prop, '#9E9E9E')

        ax.scatter(xi, yi, s=area, color=color, alpha=0.85, zorder=3,
                   edgecolors='white', linewidths=0.9)

        # Count inside bubble — scale font with bubble size
        fs = max(6.5, min(10.0, 6.0 + 3.5 * (count / max_count)))
        ax.text(xi, yi, str(count), ha='center', va='center',
                fontsize=fs, color='white', fontweight='bold', zorder=4)

        # Gene names below bubble (up to 3 + "+N more")
        sorted_genes = sorted(g for g in genes if g and g != '—')
        if sorted_genes:
            shown = sorted_genes[:3]
            rest  = len(sorted_genes) - 3
            label = ', '.join(shown)
            if rest > 0:
                label += f'\n+{rest} more'
            ax.text(xi, yi - 0.40, label, ha='center', va='top',
                    fontsize=5.0, color='#424242', zorder=4,
                    linespacing=1.3, style='italic')

    # ── Axes formatting ─────────────────────────────────────────────────────
    ax.set_xticks(range(n_props))
    ax.set_xticklabels(
        [_PROP_LABELS.get(p, p) for p in properties],
        fontsize=8.5, fontweight='bold',
    )
    ax.set_yticks(range(n_genera))
    ax.set_yticklabels(genera, fontsize=8, fontstyle='italic')
    ax.set_title('Specialty Gene Hits by Genus and Category', fontsize=9, pad=8)
    ax.tick_params(axis='both', which='both', length=0)
    for sp in ['top', 'right']:
        ax.spines[sp].set_visible(False)
    for sp in ['left', 'bottom']:
        ax.spines[sp].set_color('#BDBDBD')

    # ── Legend ───────────────────────────────────────────────────────────────
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor=_PROP_MPL_COLORS.get(p, '#9E9E9E'),
              label=_PROP_LABELS.get(p, p), alpha=0.88)
        for p in properties
    ]
    ax.legend(handles=legend_elements, loc='lower right', fontsize=6.5,
              framealpha=0.92, edgecolor='#BDBDBD',
              title='Category', title_fontsize=7.0)

    plt.tight_layout(pad=0.5)

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)

    # Scale to page width, cap height at usable page body.
    # Use available_width - 12 to account for ReportLab's 6pt frame padding
    # on each side; img_w_pt must not exceed the frame's inner usable width.
    max_w_pt = available_width - 12
    img_w_pt = max_w_pt
    aspect   = fig_h / fig_w
    img_h_pt = img_w_pt * aspect
    max_h_pt = (11 * 72) - 4 * 0.5 * 72
    if img_h_pt > max_h_pt:
        img_w_pt = img_w_pt * (max_h_pt / img_h_pt)
        img_h_pt = max_h_pt

    return Image(buf, width=img_w_pt, height=img_h_pt)


def create_annotation_genus_cards(species_groups, small_style, available_width, min_pident=0.0):
    """
    Genus-card panel table: one coloured header per detected genus, then one
    row per property category showing ALL gene hits with full classification
    details (mechanism chain, source database, pident badge).

    Layout per genus
    ─────────────────────────────────────────────────────────────────────────
    │ Escherichia  ·  14 hits                                               │
    ├─────────────────┬─────────────────────────────────────────────────────┤
    │ AMR / 1 gene    │ cysB  aminocoumarin resistance gene  [CARD 33.7%]   │
    │ Virulence / 3   │ hlyB  Toxin · Membrane-damaging · RTX  [VFDB]       │
    │                 │ mppA  [Victors]  ·  Z2207  [Victors]                │
    │ Drug Tgt / 8    │ fdhF · fdnG · gor  …  +5 more  [DrugBank]           │
    │ Transporter / 1 │ 3.A.1.5.1  [TCDB]                                   │
    ─────────────────────────────────────────────────────────────────────────

    Grouping uses ``_organism_name`` (detected sample genus), not the
    annotation's ``genus`` field (reference-database origin).
    """
    rows = _collect_protein_annotations(species_groups)
    if min_pident > 0:
        rows = _filter_annot_rows_by_pident(rows, min_pident)
    if not rows:
        return None

    # ── Collect (detected_genus, property) → annotation detail list ─────────
    # Group at genus (toplevel) level — same priority rule as bubble chart.
    buckets = {}
    for r in rows:
        genus = (
            str(r.get('_toplevel_name') or r.get('_organism_name') or
                r.get('genus', '') or '').strip()
        )
        prop  = str(r.get('property', '') or '').strip()
        if not genus or not prop:
            continue

        gene    = str(r.get('gene_name', '') or r.get('sseqid', '') or '').strip() or '—'
        raw_cl  = str(r.get('classification', '') or '').strip()
        source  = str(r.get('source', '') or '').strip()
        abx_cls = str(r.get('antibiotics_class', '') or '').strip()
        pident  = float(r.get('pident', 0) or 0)

        # Build classification chain — skip trivial/generic entries
        _skip = {'Drug target', 'None', 'nan', ''}
        classif_parts = [
            c.strip() for c in raw_cl.split(';')
            if c.strip() and c.strip() not in _skip
        ] if raw_cl else []

        buckets.setdefault((genus, prop), []).append({
            'gene':        gene,
            'classif':     classif_parts,
            'source':      source,
            'abx_cls':     abx_cls,
            'pident':      pident,
        })

    if not buckets:
        return None

    # ── Genus order: most hits first ────────────────────────────────────────
    genus_totals = {}
    for (genus, prop), items in buckets.items():
        genus_totals[genus] = genus_totals.get(genus, 0) + len(items)
    genera = sorted(genus_totals.keys(), key=lambda g: -genus_totals[g])

    # ── Paragraph styles ─────────────────────────────────────────────────────
    genus_hdr_sty = ParagraphStyle(
        'GCGenusHdr2', parent=small_style, fontSize=8.5, leading=11,
        fontName='Helvetica-Bold', textColor=colors.whitesmoke)
    badge_sty = ParagraphStyle(
        'GCBadge2', parent=small_style, fontSize=6.5, leading=8.5,
        fontName='Helvetica-Bold', textColor=colors.whitesmoke)
    gene_sty = ParagraphStyle(
        'GCGene2', parent=small_style, fontSize=6.2, leading=8.8,
        textColor=colors.HexColor('#1A1A1A'))
    spacer_sty = ParagraphStyle(
        'GCSpacer2', parent=small_style, fontSize=2, leading=3)

    table_rows = []
    style_cmds = []
    row_idx = 0

    # Column widths are finalised at table-creation time using eff_w (see below)
    # so these placeholders are overwritten; initialise to avoid NameError.
    col_badge = (available_width - 12) * 0.155
    col_genes = (available_width - 12) * 0.845

    for genus in genera:
        total = genus_totals[genus]

        # ── Genus header (full-width span) ───────────────────────────────────
        header_text = (
            f'<b>{genus}</b>'
            f'<font size="7" color="#A5D6A7">  ·  '
            f'{total} specialty gene hit{"s" if total != 1 else ""}'
            f'</font>'
        )
        table_rows.append([Paragraph(header_text, genus_hdr_sty), Paragraph('', genus_hdr_sty)])
        style_cmds += [
            ('BACKGROUND',    (0, row_idx), (-1, row_idx), colors.HexColor('#1B5E20')),
            ('SPAN',          (0, row_idx), (-1, row_idx)),
            ('TOPPADDING',    (0, row_idx), (-1, row_idx), 5),
            ('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 5),
            ('LEFTPADDING',   (0, row_idx), (-1, row_idx), 6),
            ('RIGHTPADDING',  (0, row_idx), (-1, row_idx), 4),
        ]
        row_idx += 1

        # ── Property sub-rows ────────────────────────────────────────────────
        props_for_genus = [p for p in _PROP_DISPLAY_ORDER if (genus, p) in buckets]
        for p in sorted(k[1] for k in buckets if k[0] == genus):
            if p not in _PROP_DISPLAY_ORDER and (genus, p) not in [(g, pp) for g, pp in props_for_genus]:
                props_for_genus.append(p)

        for prop in props_for_genus:
            items = buckets.get((genus, prop), [])
            badge_bg = colors.HexColor(_PROP_RL_BADGE.get(prop, '#616161'))
            row_bg   = colors.HexColor(_PROP_RL_ROW_BG.get(prop, '#F5F5F5'))
            short_lbl = _PROP_LABELS.get(prop, prop)
            n = len(items)

            badge_cell = Paragraph(
                f'<b>{short_lbl}</b><br/>'
                f'<font size="5.5">{n} gene{"s" if n != 1 else ""}</font>',
                badge_sty)

            # ── Build gene detail entries ─────────────────────────────────────
            # Sort by gene name; show up to 10, then "+N more"
            sorted_items = sorted(items, key=lambda x: x['gene'].lower())
            MAX_SHOWN = 10
            shown = sorted_items[:MAX_SHOWN]
            rest  = len(sorted_items) - MAX_SHOWN

            # Collect unique sources for a summary badge at the end
            all_sources = sorted({it['source'] for it in sorted_items if it['source']})

            entry_parts = []
            for it in shown:
                gene   = it['gene']
                cl     = it['classif']
                abx    = it['abx_cls']
                pid    = it['pident']
                src    = it['source']

                # Gene name (bold)
                s = f'<b>{gene}</b>'

                # Classification chain (italic, muted) — e.g. "Toxin · Pore-forming · RTX toxin"
                if cl:
                    cl_str = ' · '.join(cl[:4])  # cap at 4 segments to stay readable
                    s += f' <font size="5.5" color="#5D4037"><i>{cl_str}</i></font>'

                # Antibiotic class for AMR rows
                if abx and prop == 'Antibiotic Resistance':
                    s += f' <font size="5.2" color="#880000">[{abx}]</font>'

                # Source DB + pident
                pid_str = f' {pid:.0f}%' if pid > 0 else ''
                if src:
                    s += f' <font size="5.2" color="#757575">[{src}{pid_str}]</font>'

                entry_parts.append(s)

            if rest > 0:
                entry_parts.append(
                    f'<font color="#757575" size="5.5"><i>+{rest} more</i></font>')

            # Source summary at end of row (deduplicated list)
            gene_cell = Paragraph('  ·  '.join(entry_parts), gene_sty)

            table_rows.append([badge_cell, gene_cell])
            style_cmds += [
                ('BACKGROUND',    (0, row_idx), (0, row_idx), badge_bg),
                ('BACKGROUND',    (1, row_idx), (1, row_idx), row_bg),
                ('TOPPADDING',    (0, row_idx), (-1, row_idx), 4),
                ('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 4),
                ('LEFTPADDING',   (0, row_idx), (0, row_idx), 5),
                ('LEFTPADDING',   (1, row_idx), (1, row_idx), 6),
                ('RIGHTPADDING',  (0, row_idx), (-1, row_idx), 4),
                ('VALIGN',        (0, row_idx), (-1, row_idx), 'TOP'),
            ]
            row_idx += 1

        # Thin spacer between genera
        table_rows.append([Paragraph('', spacer_sty), Paragraph('', spacer_sty)])
        style_cmds += [
            ('BACKGROUND',    (0, row_idx), (-1, row_idx), colors.white),
            ('TOPPADDING',    (0, row_idx), (-1, row_idx), 2),
            ('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 2),
        ]
        row_idx += 1

    if not table_rows:
        return None

    # Use available_width - 12 to stay within the frame's inner usable width
    # (ReportLab frames have 6pt padding on each side by default).
    eff_w     = available_width - 12
    col_badge = eff_w * 0.155
    col_genes = eff_w * 0.845

    tbl = Table(table_rows, colWidths=[col_badge, col_genes])
    style_cmds += [
        ('GRID',   (0, 0), (-1, -1), 0.3, colors.Color(0.82, 0.82, 0.82)),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]
    tbl.setStyle(TableStyle(style_cmds))
    return tbl


def create_protein_annotation_xlsx(output_path, samples_dict, args):
    """
    Write a multi-sheet Excel workbook with protein annotation data.

    Sheets
    ------
    1. Genus Summary     — one row per (sample, genus, property); gene list, # hits,
                           best pident, avg e-value, median e-value.
    2. Per-Gene Hits     — one row per raw annotation hit (flat, same fields as JSON).
    3. Sample Overview   — one row per detected organism with TASS score, reads, category.
    4. AMR Genes         — filtered view: Antibiotic Resistance rows only, with
                           antibiotics_class and antibiotics columns.
    """
    import statistics
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — skipping annotation XLSX output.")
        return

    wb = openpyxl.Workbook()

    # ── Helpers ──────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="2E7D32")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL   = PatternFill("solid", fgColor="EDF7EE")
    AMR_FILL   = PatternFill("solid", fgColor="FFF3E0")
    AMR_HDR    = PatternFill("solid", fgColor="BF360C")
    THIN       = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD'),
    )
    CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='top', wrap_text=True)

    def _write_header(ws, headers, fill=HDR_FILL):
        ws.append(headers)
        for cell in ws[1]:
            cell.font     = HDR_FONT
            cell.fill     = fill
            cell.alignment = CENTER
            cell.border   = THIN

    def _autofit(ws, min_w=8, max_w=50):
        for col_cells in ws.columns:
            length = max(
                (len(str(c.value or '')) for c in col_cells),
                default=min_w
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
                min(max_w, max(min_w, length + 2))
            )

    def _style_rows(ws, start=2, alt_fill=ALT_FILL):
        for i, row in enumerate(ws.iter_rows(min_row=start), start=start):
            for cell in row:
                cell.border    = THIN
                cell.alignment = LEFT
                if i % 2 == 0:
                    cell.fill = alt_fill

    def _safe_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    all_raw_rows  = []   # for Sheet 2 (per-gene) and Sheet 4 (AMR filter)
    genus_buckets = {}   # (sample, genus, property) -> {genes, evalues, pidents}

    for sample_name, species_groups in sorted(samples_dict.items()):
        rows = _collect_protein_annotations(species_groups)
        for r in rows:
            genus  = str(r.get('_toplevel_name') or r.get('genus', '') or '').strip()
            prop   = str(r.get('property', '') or '').strip()
            gene   = str(r.get('gene_name', '') or r.get('sseqid', '') or '').strip()
            ev     = _safe_float(r.get('evalue'))
            pid    = _safe_float(r.get('pident'))
            bs     = _safe_float(r.get('bitscore'))

            all_raw_rows.append({
                'sample':            sample_name,
                'genus':             genus,
                'species':           str(r.get('species', '') or ''),
                'gene_name':         gene,
                'product':           str(r.get('product', '') or ''),
                'property':          prop,
                'classification':    str(r.get('classification', '') or ''),
                'antibiotics_class': str(r.get('antibiotics_class', '') or ''),
                'antibiotics':       str(r.get('antibiotics', '') or ''),
                'source':            str(r.get('source', '') or ''),
                'source_id':         str(r.get('source_id', '') or ''),
                'pident':            pid,
                'evalue':            ev,
                'bitscore':          bs,
                'organism':          str(r.get('organism', '') or ''),
                'level':             str(r.get('level', '') or ''),
            })

            if genus and prop:
                key = (sample_name, genus, prop)
                if key not in genus_buckets:
                    genus_buckets[key] = {'genes': set(), 'evalues': [], 'pidents': []}
                if gene:
                    genus_buckets[key]['genes'].add(gene)
                if ev is not None:
                    genus_buckets[key]['evalues'].append(ev)
                if pid is not None:
                    genus_buckets[key]['pidents'].append(pid)

    # ── Sheet 1: Genus Summary ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Genus Summary'
    _write_header(ws1, [
        'Sample', 'Genus', 'Property', 'Genes', '# Hits',
        'Best %id', 'Avg %id', 'Avg E-value', 'Median E-value',
    ])
    for (sample, genus, prop), b in sorted(genus_buckets.items()):
        genes_str = ', '.join(sorted(b['genes']))
        n         = len(b['evalues']) or len(b['genes'])
        best_pid  = max(b['pidents']) if b['pidents'] else None
        avg_pid   = statistics.mean(b['pidents']) if b['pidents'] else None
        avg_ev    = statistics.mean(b['evalues']) if b['evalues'] else None
        med_ev    = statistics.median(b['evalues']) if b['evalues'] else None
        ws1.append([sample, genus, prop, genes_str, n,
                    round(best_pid, 2) if best_pid is not None else '',
                    round(avg_pid, 2)  if avg_pid  is not None else '',
                    avg_ev, med_ev])
    _style_rows(ws1)
    _autofit(ws1)

    # ── Sheet 2: Per-Gene Hits ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Per-Gene Hits')
    _write_header(ws2, [
        'Sample', 'Genus', 'Species', 'Gene', 'Product', 'Property',
        'Classification', 'Antibiotics Class', 'Antibiotics',
        'Source', 'Source ID', '%id', 'E-value', 'Bitscore',
        'Reference Organism', 'Level',
    ])
    for r in all_raw_rows:
        ws2.append([
            r['sample'], r['genus'], r['species'], r['gene_name'],
            r['product'], r['property'], r['classification'],
            r['antibiotics_class'], r['antibiotics'],
            r['source'], r['source_id'],
            r['pident'], r['evalue'], r['bitscore'],
            r['organism'], r['level'],
        ])
    _style_rows(ws2)
    _autofit(ws2)

    # ── Sheet 3: Sample Overview ──────────────────────────────────────────────
    ws3 = wb.create_sheet('Sample Overview')
    _write_header(ws3, [
        'Sample', 'Sample Type', 'Genus', 'Species/Subkey',
        'Microbial Category', 'Ann Class', 'High Consequence',
        'TASS Score', 'Reads', 'Coverage', 'RPM', 'RPKM',
        'Minhash Score', 'Gini', 'Passes Threshold',
    ])
    _mc_map = getattr(args, '_sample_min_conf', {})
    for sample_name, species_groups in sorted(samples_dict.items()):
        _mc = _mc_map.get(sample_name, _DEFAULT_CONF)
        for sg in species_groups:
            stype = sg.get('sampletype', '')
            for sk_m in sg.get('members', []):
                passes = 'TRUE' if passes_confidence_threshold(sk_m, _mc) else 'FALSE'
                ws3.append([
                    sample_name,
                    stype,
                    sg.get('toplevelname', sg.get('name', '')),
                    sk_m.get('subkeyname', sk_m.get('name', '')),
                    sk_m.get('microbial_category', ''),
                    sk_m.get('annClass', ''),
                    'TRUE' if sk_m.get('high_cons', False) else 'FALSE',
                    round(float(sk_m.get('tass_score', 0) or 0), 4),
                    int(sk_m.get('numreads', 0) or 0),
                    round(float(sk_m.get('coverage', 0) or 0), 4),
                    round(float(sk_m.get('rpm', 0) or 0), 2),
                    round(float(sk_m.get('rpkm', 0) or 0), 4),
                    round(float(sk_m.get('minhash_score', 0) or 0), 4),
                    round(float(sk_m.get('gini_coefficient', 0) or 0), 4),
                    passes,
                ])
    _style_rows(ws3)
    _autofit(ws3)

    # ── Sheet 4: AMR Genes ────────────────────────────────────────────────────
    ws4 = wb.create_sheet('AMR Genes')
    _write_header(ws4, [
        'Sample', 'Genus', 'Species', 'Gene', 'Product',
        'Antibiotics Class', 'Antibiotics', 'Classification',
        'Source', 'Source ID', '%id', 'E-value', 'Bitscore',
    ], fill=AMR_HDR)
    amr_rows = [r for r in all_raw_rows if 'resist' in r['property'].lower() or
                'amr' in r['property'].lower() or
                r['antibiotics_class'] or r['antibiotics']]
    for r in amr_rows:
        ws4.append([
            r['sample'], r['genus'], r['species'], r['gene_name'],
            r['product'], r['antibiotics_class'], r['antibiotics'],
            r['classification'], r['source'], r['source_id'],
            r['pident'], r['evalue'], r['bitscore'],
        ])
    _style_rows(ws4, alt_fill=AMR_FILL)
    _autofit(ws4)

    wb.save(output_path)
    print(f"Annotation XLSX written: {output_path}  "
          f"({len(all_raw_rows)} gene hits, {len(genus_buckets)} genus-property buckets)")


def create_pdf_template(output_path, samples_dict, args):
    """
    Create the full PDF report.
    """
    page_width, page_height = letter
    left_margin = page_width * 0.01
    right_margin = page_width * 0.01

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    available_width = doc.width
    story = []
    outline_collector = _OutlineCollector()
    styles = getSampleStyleSheet()
    taxid_to_bookmark = build_taxid_to_bookmark_map(samples_dict)

    legend_text_style = ParagraphStyle('LegendText', parent=styles['Normal'], fontSize=8, leading=10)
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=24,
        textColor=colors.HexColor('#2C3E50'), spaceAfter=20, alignment=TA_CENTER)
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'], fontSize=16,
        textColor=colors.HexColor('#34495E'), spaceAfter=0, spaceBefore=12)
    indent_style = ParagraphStyle(
        'IndentStyle', parent=styles['Normal'], fontSize=7, leading=10, leftIndent=20)
    small_style = ParagraphStyle('SmallText', parent=styles['Normal'], fontSize=10, leading=10)
    metadata_style = styles['Normal']

    # Collect all strains (flattened from the 3-level hierarchy)
    all_strains = []
    for sample_name, species_groups in samples_dict.items():
        for species_group in species_groups:
            all_strains.extend(get_all_strains(species_group))

    if not args.enable_matrix:
        show_ani_column = False
    else:
        show_ani_column = check_if_any_high_ani_in_dataset(all_strains)
    show_k2_column = check_if_k2_reads_present(all_strains)
    use_subkey = not args.no_subkey
    if use_subkey:
        # Auto-disable subkey columns if ALL strains have subkey == key
        any_subkey_differs = any(
            str(strain.get('subkey', strain.get('key', ''))) != str(strain.get('key', ''))
            for species_groups in samples_dict.values()
            for sg in species_groups
            for strain in get_all_strains(sg)
        )
        if not any_subkey_differs:
            use_subkey = False
            print("Auto-disabled subkey grouping: all subkeys match their keys")

    show_strains_table = getattr(args, 'show_strains_table', False)
    show_strain_appendix = (not show_strains_table) and (not use_subkey)

    print(f"\nHigh ANI column: {'SHOWN' if show_ani_column else 'HIDDEN'} (pre-computed from match_paths.py)")
    print(f"K2 Reads column: {'SHOWN' if show_k2_column else 'HIDDEN'}")
    print(f"Subkey grouping: {'ENABLED' if use_subkey else 'DISABLED'}")
    if use_subkey:
        print("Strain detail layout: species/subkey rows with qualifying strain rows inline in main table")
    else:
        print(f"Strain detail table: {'INLINE' if show_strains_table else 'APPENDIX'}")
    print(f"\nFiltering settings:")
    print(f"  Show Potentials: {args.show_potentials}")
    print(f"  Show Commensals: {args.show_commensals}")
    print(f"  Show Opportunistic: {args.show_opportunistic}")
    print(f"  Show Unidentified: {args.show_unidentified}")
    print(f"  Sorting mode: {'Alphabetical' if args.sort_alphabetical else 'TASS Score (descending)'}")
    print(f"  Max members per group: {args.max_members if args.max_members else 'Unlimited'}")
    print(f"  Max TOC groups per sample: {args.max_toc}")
    # Collect low confidence strains (per-sample threshold)
    sample_min_conf = args._sample_min_conf
    low_confidence_strains = []
    for sample_name, species_groups in samples_dict.items():
        _mc = sample_min_conf.get(sample_name, _DEFAULT_CONF)
        for species_group in species_groups:
            for strain in get_all_strains(species_group):
                if should_include_strain(strain, args) and has_min_reads(strain, args.min_reads):
                    if not strain_passes_with_subkey_promotion(strain, species_group, _mc):
                        if strain.get("high_cons"):
                            low_confidence_strains.append({
                                'strain': strain,
                                'sample_name': sample_name,
                                'species_group': species_group
                            })

    valid_bookmarks = collect_all_bookmarks(samples_dict, low_confidence_strains)
    valid_bookmarks.update(taxid_to_bookmark.values())

    # ── Title ─────────────────────────────────────────────────────────────────
    story.append(OutlineDest('outline_title', 'Organism Discovery Report', level=0, closed=False, collector=outline_collector))
    story.append(Paragraph("Organism Discovery Report", title_style))
    story.append(Spacer(1, 0.05*inch))

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    generation_text = (
        f"This report was generated using "
        f"<link href=\"https://github.com/jhuapl-bio/taxtriage\" color=\"blue\">TaxTriage</link> "
        f"at <b>{current_time}</b> and is derived from "
        f"an in <link href=\"https://github.com/jhuapl-bio/taxtriage/blob/main/assets/pathogen_sheet.csv\" "
        f"color=\"blue\">development spreadsheet of human-host pathogens</link>."
    )
    story.append(Paragraph(generation_text, metadata_style))
    workflow_rev = None
    for _meta in getattr(args, '_input_metadata', {}).values():
        _rev = _meta.get('workflow_revision')
        if _rev:
            workflow_rev = _rev
            break
    if workflow_rev and str(workflow_rev).upper() != "NA":
        story.append(Paragraph(f"TaxTriage version: <b>{workflow_rev}</b>", metadata_style))
    elif not workflow_rev:
        story.append(Paragraph(f"TaxTriage Version: <b>Not Specified or Local Build</b>", metadata_style))
    commit_id = None
    for _meta in getattr(args, '_input_metadata', {}).values():
        _cid = _meta.get('commit_id')
        if _cid:
            commit_id = _cid
            break
    if commit_id and str(commit_id).upper() != "NA":
        story.append(Paragraph(f"Commit ID: <link href=\"https://github.com/jhuapl-bio/taxtriage/commit/{commit_id}\" color=\"blue\">{commit_id}</link>", metadata_style))
    story.append(Spacer(1, 0.02*inch))
    story.append(Paragraph("<b>★</b> = High Consequence Pathogen", small_style))
    story.append(Spacer(1, 0.02*inch))

    # ── Table of Contents ─────────────────────────────────────────────────────
    story.append(OutlineDest('outline_toc', 'Table of Contents', level=0, closed=True, collector=outline_collector))
    story.append(Paragraph("Table of Contents", heading_style))
    story.append(Spacer(1, 0.03*inch))
    story.append(Paragraph(
        "<i>Format: Sample Name (Total Alignments # - # Primary Pathogens) → "
        "Category Label ■ (# Primary Strains, Max TASS)</i>", small_style))
    story.append(Spacer(1, 0.15*inch))
    story.append(Paragraph(
        "Click on sample names or species groups to jump to their sections. Samples are listed in order of appearance in the main table with their top groups listed below. They are sorted by TASS Score by default or alphabetically if selected."
        "Only samples/groups with visible strains are shown here", small_style))
    story.append(Paragraph(
        "The table is organized by samples first, then in order of TASS Score by default or alphabetical if selected. "
        "Each genus group expands into species/subkey summary rows and then any child strains that also clear the cutoff. "
        "The number of visible child strains can still be limited with the max-members setting.",
        small_style))
    # add text for "Sample"
    story.append(Spacer(1, 0.12*inch))
    story.append(Paragraph(f"<b>Samples in Run:</b>", small_style))
    story.append(Spacer(1, 0.07*inch))
    for sample_name in sorted(samples_dict.keys()):
        bookmark_name = f"sample_{sanitize_bookmark_name(sample_name)}"
        species_groups = samples_dict[sample_name]
        total_alignments, primary_count = get_sample_stats(species_groups)
        _mc = sample_min_conf.get(sample_name, _DEFAULT_CONF)

        # Filter then sort — same logic as the main table so TOC order matches
        visible_groups = []
        for sg in species_groups:
            _qualifying = get_qualifying_strains(sg, args, _mc, min_reads=1)
            if _qualifying:
                best_tass = max(s.get('tass_score', 0) for s in _qualifying)
                visible_groups.append((sg, best_tass))

        if not visible_groups:
            continue

        if args.sort_alphabetical:
            visible_groups.sort(key=lambda pair: pair[0].get('toplevelname', 'Unknown'))
        else:
            visible_groups.sort(key=lambda pair: pair[1], reverse=True)

        visible_groups = [sg for sg, _ in visible_groups]

        _toc_meta = getattr(args, '_input_metadata', {}).get(sample_name, {})
        _toc_plat = _toc_meta.get('platform', '')
        _toc_plat_str = f" — {_toc_plat}" if _toc_plat and _toc_plat != 'unknown' else ''
        link_text = create_safe_link(
            f'<b>{sample_name}</b>{_toc_plat_str} ({total_alignments:,} Alignments - {primary_count} Primary Pathogens)',
            bookmark_name, valid_bookmarks)
        story.append(Paragraph(link_text, small_style))
        story.append(Spacer(1, 0.05*inch))

        toc_groups = visible_groups[:args.max_toc]
        has_more = len(visible_groups) > args.max_toc

        species_links = []
        for sg in toc_groups:
            sp_name = sg.get('toplevelname', 'Unknown')
            sp_key = sg.get('toplevelkey', sg.get('key', 'unknown'))
            sp_bm = f"species_{sanitize_bookmark_name(sample_name)}_{sp_key}"
            g_tass, g_primary, _ = get_species_group_stats(sg)
            g_cat = sg.get('microbial_category', 'Unknown')
            g_ann = sg.get('annClass', '')
            cat_color = get_category_color(g_cat, g_ann, alpha=1.0)
            color_hex = '#{:02x}{:02x}{:02x}'.format(
                int(cat_color.red * 255), int(cat_color.green * 255), int(cat_color.blue * 255))
            species_links.append(
                f'{create_safe_link(sp_name, sp_bm, valid_bookmarks)} '
                f'<font color="{color_hex}">■</font> '
                f'({g_primary}, {g_tass*100:.1f})'
            )

        if has_more:
            first_hidden = visible_groups[args.max_toc]
            h_key = first_hidden.get('toplevelkey', first_hidden.get('key', 'unknown'))
            h_bm = f"species_{sanitize_bookmark_name(sample_name)}_{h_key}"
            species_links.append(create_safe_link(
                f"... ({len(visible_groups) - args.max_toc} more)", h_bm, valid_bookmarks))

        story.append(Paragraph(f"→ {', '.join(species_links)}", indent_style))
        story.append(Spacer(1, 0.02*inch))

    story.append(Spacer(1, 0.05*inch))
    story.append(Paragraph('• ' + create_safe_link('Color Key', 'color_key', valid_bookmarks), styles['Normal']))
    story.append(Paragraph('• ' + create_safe_link('Column Explanations', 'column_explanations', valid_bookmarks), styles['Normal']))
    if low_confidence_strains:
        story.append(Paragraph(
            '• ' + create_safe_link('Low Confidence, High Consequence Detections', 'low_confidence', valid_bookmarks),
            styles['Normal']))
    if show_strain_appendix:
        story.append(Paragraph(
            '• ' + create_safe_link('Additional Lower-Level Strain Detail', 'strain_detail_table', valid_bookmarks),
            styles['Normal']))
    story.append(Paragraph('• ' + create_safe_link('Additional Information', 'additional_info', valid_bookmarks), styles['Normal']))
    story.append(Spacer(1, 0.00*inch))

    # ── Per-sample content ────────────────────────────────────────────────────
    # Also build sorted_groups_by_sample for the strain appendix.
    sorted_groups_by_sample = []  # list of (sample_name, sorted_groups)

    for sample_name in sorted(samples_dict.keys()):
        bookmark_name = f"sample_{sanitize_bookmark_name(sample_name)}"
        story.append(AnchorFlowable(bookmark_name))

        sampletype = (samples_dict[sample_name][0].get('sampletype', 'Unspecified Type')
                      if samples_dict[sample_name] else 'Unspecified Type')
        _mc = sample_min_conf.get(sample_name, _DEFAULT_CONF)
        _smeta_hdr = getattr(args, '_input_metadata', {}).get(sample_name, {})
        _plat_hdr = _smeta_hdr.get('platform', '')
        _plat_str = f" — {_plat_hdr}" if _plat_hdr and _plat_hdr != 'unknown' else ''
        # PDF outline: sample as top-level entry (collapsed by default)
        outline_sample_key = f"outline_{bookmark_name}"
        story.append(OutlineDest(outline_sample_key,
                                 f'Sample: {sample_name}{_plat_str}', level=0, closed=True,
                                 collector=outline_collector))
        story.append(Paragraph(f"Sample: {sample_name}{_plat_str} ({sampletype})", heading_style))
        # Sample metadata line: TASS cutoff + source + platform + read stats from input metadata
        _smeta = getattr(args, '_input_metadata', {}).get(sample_name, {})
        _conf_src = getattr(args, '_sample_conf_source', {}).get(sample_name, '')
        _meta_parts = [f"TASS cutoff: <b>{_mc:.2f}</b>"]
        if _conf_src:
            _meta_parts.append(f"Source: {_conf_src}")
        _tr = _smeta.get('total_reads')
        _ar = _smeta.get('aligned_reads')
        if _tr is not None:
            _meta_parts.append(f"Total reads: {int(_tr):,}")
        if _ar is not None:
            _meta_parts.append(f"Aligned: {int(_ar):,}")
        _nsg = _smeta.get('num_species_groups')
        if _nsg is not None:
            _meta_parts.append(f"Species groups: {_nsg}")
        story.append(Paragraph(
            f"<i>{' &bull; '.join(_meta_parts)}</i>", small_style))
        if sampletype in ['blood', 'csf', 'sterile', 'serum']:
            story.append(Paragraph(f"<font color=\"#666666\">\t&#8594; {sampletype} sample likely leads to lower TASS scores due to relatively low read count or coverage of organisms. All pathogens are defaulted to primary pathogens.</font>", small_style))
            if sampletype != "blood":
                story.append(Paragraph(f"<font color=\"#666666\">\t&#8594; {sampletype} follows the same anticipated clinical distribution as blood samples.</font>", small_style))

        # ── Check if any qualifying strains have below-threshold zscore ───
        # If so, add a note explaining the diamond symbol and faded rows.
        _zt_note = args.zscore_threshold
        if _zt_note is not None:
            _has_low_z = False
            _max_site_count = 0
            for sg in samples_dict[sample_name]:
                for m in get_all_strains(sg):
                    mz = float(m.get('zscore', 0) or 0)
                    _sc = int(m.get('hmp_site_count', 0) or 0)
                    if _sc > _max_site_count:
                        _max_site_count = _sc
                    if mz < _zt_note:
                        _has_low_z = True
            if _has_low_z:
                _total_bold = f'<b>{_max_site_count:,}</b>' if _max_site_count > 0 else 'N/A'
                story.append(Paragraph(
                    f'<font color="#666666">&#9830; Rows marked with a diamond '
                    f'(&#9830;) and faded styling indicate organisms whose abundance '
                    f'z-score is below the threshold of {_zt_note}. These organisms '
                    f'are within expected abundance ranges for this sample type and '
                    f'are less likely to be clinically significant. The number next to '
                    f'the diamond shows how many reference samples contained that '
                    f'organism, with the percentage of total samples in parentheses. '
                    f'The reference dataset was compiled from {_total_bold} SRA/R samples '
                    f'for this body site.</font>',
                    small_style))

        story.append(Spacer(1, 0.08*inch))

        species_groups = samples_dict[sample_name]
        # Use total reads from metadata (includes unaligned) when available;
        # fall back to sum of aligned reads across species groups.
        _smeta_reads = getattr(args, '_input_metadata', {}).get(sample_name, {})
        _meta_total = _smeta_reads.get('total_reads')
        sample_total_reads = int(_meta_total) if _meta_total else sum(sg.get('numreads', 0) for sg in species_groups)

        # ── Filter members first, THEN sort groups by best visible TASS ─────
        all_sample_strains = []
        species_group_map = {}
        # Collect qualifying strains per group so we can sort by post-filter TASS
        _group_qualified = []  # list of (sg, [qualifying strains])
        _zt = args.zscore_threshold
        for sg in species_groups:
            group_strains = get_qualifying_strains(sg, args, _mc, min_reads=1)
            if group_strains:
                group_strains.sort(key=lambda s: s.get('tass_score', 0), reverse=True)
                if _zt is not None:
                    _z_vals = [float(s.get('zscore', 0) or 0) for s in group_strains]
                    _has_hi = any(z >= _zt for z in _z_vals)
                    _has_lo = any(z < _zt for z in _z_vals)
                    if _has_hi and _has_lo:
                        group_strains.sort(
                            key=lambda s: (
                                float(s.get('zscore', 0) or 0) < _zt,
                                -float(s.get('tass_score', 0) or 0)
                            )
                        )
                _group_qualified.append((sg, group_strains))

        # ── Sort: elevated z-score groups first, then by TASS / alpha ─────
        def _group_zscore(pair):
            """Max zscore across the group-level and all qualifying members."""
            sg, strains = pair
            vals = [float(sg.get('zscore', 0) or 0)]
            vals.extend(float(s.get('zscore', 0) or 0) for s in strains)
            return max(vals)

        if args.sort_alphabetical:
            _group_qualified.sort(key=lambda pair: pair[0].get('toplevelname', 'Unknown'))
        else:
            _group_qualified.sort(
                key=lambda pair: pair[1][0].get('tass_score', 0) if pair[1] else 0,
                reverse=True)

        # Partition: elevated zscore groups come first (preserving order within)
        if _zt is not None:
            _elevated = [p for p in _group_qualified if _group_zscore(p) >= _zt]
            _normal = [p for p in _group_qualified if _group_zscore(p) < _zt]
            _group_qualified = _elevated + _normal

        # ── Move flora-on-sterile groups to the very end ──────────────────────
        # In sterile sample types (blood, csf, serum, etc.) organisms with
        # commensal-flora annotations are rendered after a "Potential
        # Contaminants" separator, so keep them together at the tail.
        _sampletype_norm = sampletype.lower().strip() if sampletype else ''
        if _sampletype_norm in _STERILE_TYPES:
            def _pair_has_flora(pair):
                sg, strains = pair
                if _is_flora_on_sterile(sg, sampletype):
                    return True
                return any(_is_flora_on_sterile(s, sampletype) for s in strains)
            _non_flora_pairs = [p for p in _group_qualified if not _pair_has_flora(p)]
            _flora_pairs = [p for p in _group_qualified if _pair_has_flora(p)]
            if _flora_pairs and _non_flora_pairs:
                _group_qualified = _non_flora_pairs + _flora_pairs

        sorted_groups = [sg for sg, _ in _group_qualified]

        # Record for appendix
        sorted_groups_by_sample.append((sample_name, sorted_groups))

        # Build subkey group lookup: (toplevelkey, subkey) → subkey_group_data,
        # plus a display lookup for visible species/subkey rows and child strains.
        _subkey_group_lookup = {}
        _subkey_display_lookup = {}
        for sg, group_strains in _group_qualified:
            if args.max_members is not None and args.max_members > 0:
                group_strains = group_strains[:args.max_members]
            for strain in group_strains:
                species_group_map[id(strain)] = sg
            all_sample_strains.extend(group_strains)
            tlk = str(sg.get('toplevelkey', sg.get('key', '')))
            for sk_grp in get_subkey_groups(sg):
                sk = str(sk_grp.get('subkey', sk_grp.get('key', '')))
                _subkey_group_lookup[(tlk, sk)] = sk_grp
            _subkey_display_lookup[tlk] = build_subkey_display_lookup(
                sg,
                args,
                _mc,
                min_reads=1,
                max_members=args.max_members,
            )

        if all_sample_strains:
            _derived_alerts = []
            if use_subkey:
                for _species_rows in _subkey_display_lookup.values():
                    for _row in _species_rows.values():
                        if _row.get('harmful_followup'):
                            _derived_alerts.append(_row)
            if _derived_alerts:
                _alert_names = ', '.join(sorted({a['display_name'] for a in _derived_alerts})[:4])
                _alert_extra = ''
                if len({a['display_name'] for a in _derived_alerts}) > 4:
                    _alert_extra = f' and {len({a["display_name"] for a in _derived_alerts}) - 4} more'
                # story.append(Paragraph(
                #     f'{_WARN_SYMBOL_HTML} '
                #     f'Potential harmful pathogen signal detected in one or more species/subkey rows. '
                #     f'These rows have qualifying child strains that require follow-up. '
                #     f'Affected rows are marked with {_WARN_SYMBOL_HTML}: <b>{_alert_names}{_alert_extra}</b>.',
                #     small_style))
                story.append(Spacer(1, 0.05*inch))
            # Auto-detect control data presence for the control bar column
            # (now also includes insilico_comparison data)
            _show_ctrl = getattr(args, 'show_control_bar', False) or _has_control_data(species_groups)
            # Gather missing positive controls for this sample (if not hidden)
            _smeta = getattr(args, '_input_metadata', {}).get(sample_name, {})
            _miss_pos_list = []
            if not getattr(args, 'hide_missing_pos_controls', False):
                _miss_pos_list = _smeta.get('missing_positive_controls') or []

            combined_table = create_combined_sample_table(
                all_sample_strains, species_group_map, small_style,
                show_ani_column, show_k2_column,
                taxid_to_bookmark, valid_bookmarks,
                sample_total_reads=sample_total_reads,
                sample_name=sample_name,
                available_width=available_width,
                use_subkey=use_subkey,
                show_strains_table=show_strains_table,
                outline_collector=outline_collector,
                zscore_threshold=args.zscore_threshold,
                sampletype=sampletype,
                show_control_bar=_show_ctrl,
                missing_pos_controls=_miss_pos_list,
                subkey_group_lookup=_subkey_group_lookup,
                subkey_display_lookup=_subkey_display_lookup,
            )
            story.append(combined_table)

            # ── Control bar legend (shown only when control data is present) ──
            if _show_ctrl:
                _smeta_ctrl = getattr(args, '_input_metadata', {}).get(sample_name, {})
                _neg_used = _smeta_ctrl.get('negative_controls_used') or []
                _pos_used = _smeta_ctrl.get('positive_controls_used') or []
                _fold_thresh = _smeta_ctrl.get('control_fold_threshold')
                _ctrl_legend_style = ParagraphStyle(
                    'CtrlLegend', parent=small_style,
                    fontSize=6.5, leading=9, textColor=colors.Color(0.35, 0.35, 0.35))
                _ctrl_parts = [
                    '<b>Ctrl column:</b> '
                    'The bar shows each organism\u2019s TASS score position (▲) relative to '
                    'control samples. '
                ]
                if _neg_used:
                    _neg_names = ', '.join(_neg_used)
                    _ctrl_parts.append(
                        f'<font color="#cc2222">\u25a0 Red zone</font> = '
                        f'negative control range ({_neg_names}). '
                    )
                if _pos_used:
                    _pos_names = ', '.join(_pos_used)
                    _ctrl_parts.append(
                        f'<font color="#2a8a2a">\u25a0 Green zone</font> = '
                        f'positive control range ({_pos_names}). '
                    )
                _ctrl_parts.append(
                    'The label below the bar shows: '
                    '<b>#\u00d7</b> = fold-change of sample TASS; '
                    '<b>#\u00d7 rd</b> = same ratio for read counts. '
                    'Prefix symbols indicate control type: '
                )
                if _neg_used:
                    _ctrl_parts.append(
                        '<b>\u2212</b> (minus) = vs. max negative control '
                        '(dark gray / <font color="#cc2222">red</font> text). '
                    )
                if _pos_used:
                    _ctrl_parts.append(
                        '<b>+</b> (plus) = vs. min positive control '
                        '(<font color="#2a8a2a">green text</font>). '
                    )
                _isil_used = _smeta_ctrl.get('insilico_controls_used') or []
                if _isil_used:
                    _isil_names = ', '.join(_isil_used)
                    _ctrl_parts.append(
                        f'<font color="#008B8B">\u25a0 Teal zone</font> = '
                        f'in-silico simulation range ({_isil_names}). '
                    )
                    _ctrl_parts.append(
                        '<b>\u221e</b> (loop) = vs. in-silico control '
                        '(<font color="#007373">teal text</font>). '
                    )
                if _fold_thresh is not None:
                    _ctrl_parts.append(
                        f'Organisms with fold &lt; {_fold_thresh}\u00d7 are considered '
                        f'within negative-control bounds (triangle shown in red).'
                    )
                story.append(Spacer(1, 0.03 * inch))
                story.append(Paragraph(''.join(_ctrl_parts), _ctrl_legend_style))

            # ── In-silico metrics summary tables (one per simulator type) ─────
            _miss_isil = _smeta.get('missing_insilico_controls') or []
            _miss_by_type = _smeta.get('missing_insilico_by_type') or {}
            _sim_types = _smeta.get('insilico_simulator_types') or []

            _isil_hdr_style = ParagraphStyle(
                'IsilHeader', parent=small_style,
                fontSize=7.5, leading=10, fontName='Helvetica-Bold',
                textColor=colors.HexColor('#008B8B'))
            _miss_hdr_style = ParagraphStyle(
                'MissIsilHeader', parent=small_style,
                fontSize=7.5, leading=10, fontName='Helvetica-Bold',
                textColor=colors.HexColor('#B22222'))

            _SIM_TYPE_LABELS = {
                'iss': 'InSilicoSeq (Illumina)',
                'nanosim': 'NanoSim (ONT)',
                'unknown': 'In-Silico',
            }

            def _render_isil_section(sim_label, comp_key, miss_list, so_comp_key):
                """Render one metrics table + missing + sample-only section."""
                _tbl = build_insilico_metrics_table(
                    species_groups, _mc, available_width,
                    missing_insilico=miss_list,
                    comparison_key=comp_key,
                )
                if not _tbl:
                    return
                story.append(Spacer(1, 0.08 * inch))
                story.append(Paragraph(
                    f'{sim_label} Metrics (TASS threshold: {_mc*100:.1f})',
                    _isil_hdr_style))
                story.append(Spacer(1, 0.03 * inch))
                story.append(_tbl)

                # Missing organisms detail
                if miss_list:
                    _miss_tbl = build_missing_insilico_detail_table(
                        miss_list, available_width)
                    if _miss_tbl:
                        story.append(Spacer(1, 0.06 * inch))
                        story.append(Paragraph(
                            f'Missing {sim_label} Organisms (present in simulation, absent from sample)',
                            _miss_hdr_style))
                        story.append(Spacer(1, 0.03 * inch))
                        story.append(_miss_tbl)

                # Sample-only organisms for this type
                _so_orgs = []
                for sg in species_groups:
                    ic = sg.get(so_comp_key) or {}
                    if ic.get('missing_from_insilico'):
                        _so_name = sg.get('name') or sg.get('toplevelname') or str(sg.get('toplevelkey', ''))
                        _so_tass = float(sg.get('tass_score', 0) or 0)
                        _so_cat = sg.get('microbial_category') or sg.get('is_pathogen') or 'Unknown'
                        _so_orgs.append((_so_name, _so_tass, _so_cat))
                if _so_orgs:
                    _so_hdr_style = ParagraphStyle(
                        'SampleOnlyHeader', parent=small_style,
                        fontSize=7.5, leading=10, fontName='Helvetica-Bold',
                        textColor=colors.HexColor('#555555'))
                    story.append(Spacer(1, 0.06 * inch))
                    story.append(Paragraph(
                        f'Sample-Only Organisms \u2014 {sim_label} (detected in sample but absent from simulation)',
                        _so_hdr_style))
                    _so_cell_style = ParagraphStyle(
                        'SampleOnlyCell', parent=small_style,
                        fontSize=7, leading=9)
                    _so_parts = []
                    for _so_name, _so_tass, _so_cat in _so_orgs:
                        _so_parts.append(
                            f'<b>{_so_name}</b> (TASS: {_so_tass*100:.2f}, {_so_cat})')
                    story.append(Spacer(1, 0.02 * inch))
                    story.append(Paragraph(
                        'These organisms are counted as FP in the metrics above: ' +
                        '; '.join(_so_parts) + '.',
                        _so_cell_style))

            if _sim_types and len(_sim_types) >= 1:
                # Render one table per simulator type
                for _st in _sim_types:
                    _st_label = _SIM_TYPE_LABELS.get(_st, f'In-Silico ({_st})')
                    _st_comp_key = f'insilico_comparison_{_st}'
                    _st_miss = _miss_by_type.get(_st, [])
                    _render_isil_section(_st_label, _st_comp_key, _st_miss, _st_comp_key)
            else:
                # Fallback: single combined table (no per-type data available)
                _isil_tbl = build_insilico_metrics_table(
                    species_groups, _mc, available_width,
                    missing_insilico=_miss_isil,
                )
                if _isil_tbl:
                    story.append(Spacer(1, 0.08 * inch))
                    story.append(Paragraph(
                        f'In-Silico Simulation Metrics (TASS threshold: {_mc*100:.1f})',
                        _isil_hdr_style))
                    story.append(Spacer(1, 0.03 * inch))
                    story.append(_isil_tbl)

                    if _miss_isil:
                        _miss_detail_tbl = build_missing_insilico_detail_table(
                            _miss_isil, available_width)
                        if _miss_detail_tbl:
                            story.append(Spacer(1, 0.06 * inch))
                            story.append(Paragraph(
                                'Missing In-Silico Organisms (present in simulation, absent from sample)',
                                _miss_hdr_style))
                            story.append(Spacer(1, 0.03 * inch))
                            story.append(_miss_detail_tbl)

                    _sample_only_orgs = []
                    for sg in species_groups:
                        ic = sg.get('insilico_comparison') or {}
                        if ic.get('missing_from_insilico'):
                            _so_name = sg.get('name') or sg.get('toplevelname') or str(sg.get('toplevelkey', ''))
                            _so_tass = float(sg.get('tass_score', 0) or 0)
                            _so_cat = sg.get('microbial_category') or sg.get('is_pathogen') or 'Unknown'
                            _sample_only_orgs.append((_so_name, _so_tass, _so_cat))
                    if _sample_only_orgs:
                        _so_hdr_style = ParagraphStyle(
                            'SampleOnlyHeader', parent=small_style,
                            fontSize=7.5, leading=10, fontName='Helvetica-Bold',
                            textColor=colors.HexColor('#555555'))
                        story.append(Spacer(1, 0.06 * inch))
                        story.append(Paragraph(
                            'Sample-Only Organisms (detected in sample but absent from in-silico simulation)',
                            _so_hdr_style))
                        _so_cell_style = ParagraphStyle(
                            'SampleOnlyCell', parent=small_style,
                            fontSize=7, leading=9)
                        _so_parts = []
                        for _so_name, _so_tass, _so_cat in _sample_only_orgs:
                            _so_parts.append(
                                f'<b>{_so_name}</b> (TASS: {_so_tass*100:.2f}, {_so_cat})')
                        story.append(Spacer(1, 0.02 * inch))
                        story.append(Paragraph(
                            'These organisms are counted as FP in the metrics above: ' +
                            '; '.join(_so_parts) + '.',
                            _so_cell_style))

            # ── Protein Annotation Hits (below main tables) ───────────────────
            _annot_min_pident = getattr(args, 'annot_min_pident', 90.0)
            _annot_hdr_style = ParagraphStyle(
                'AnnotHdr', parent=small_style,
                fontSize=9, leading=11, fontName='Helvetica-Bold',
                textColor=colors.HexColor('#2E7D32'))
            _annot_note_style = ParagraphStyle(
                'AnnotNote', parent=small_style,
                fontSize=6.5, leading=9,
                textColor=colors.Color(0.4, 0.4, 0.4))

            # ── Bubble chart: genus × category overview ───────────────────
            _bubble_img = create_annotation_bubble_chart(
                species_groups, available_width, min_pident=_annot_min_pident)
            if _bubble_img:
                story.append(Spacer(1, 0.12 * inch))
                _annot_bm = f"annot_{sanitize_bookmark_name(sample_name)}"
                story.append(AnchorFlowable(_annot_bm))
                story.append(Paragraph(
                    'Specialty Gene Hits (DIAMOND BLASTx)',
                    _annot_hdr_style))
                story.append(Spacer(1, 0.03 * inch))
                story.append(Paragraph(
                    'Protein-level specialty gene matches from de novo assembly '
                    'DIAMOND BLASTx, grouped by detected genus and annotation '
                    'category.  Bubble size reflects the number of distinct gene '
                    'hits in each (genus, category) group; representative gene '
                    f'names are shown beneath each bubble '
                    f'(pident ≥ {_annot_min_pident:.0f}%).',
                    _annot_note_style))
                story.append(Spacer(1, 0.06 * inch))
                story.append(_bubble_img)

                # ── Genus-card detail table ───────────────────────────────
                _genus_cards = create_annotation_genus_cards(
                    species_groups, small_style, available_width,
                    min_pident=_annot_min_pident)
                if _genus_cards:
                    story.append(Spacer(1, 0.10 * inch))
                    story.append(Paragraph(
                        'Specialty Gene Detail by Genus',
                        _annot_hdr_style))
                    story.append(Spacer(1, 0.03 * inch))
                    story.append(Paragraph(
                        'Each genus block lists detected specialty genes grouped '
                        'by category (AMR, Virulence, Drug Target, Transporter). '
                        'Classification or mechanism is shown in parentheses where '
                        'available.  Genes are sorted alphabetically; entries beyond '
                        'the first 8 per category are summarised as "+N more".',
                        _annot_note_style))
                    story.append(Spacer(1, 0.06 * inch))
                    story.append(_genus_cards)

        else:
            story.append(Paragraph(
                "<i>No data available for this sample above confidence threshold</i>",
                styles['Italic']))
            if low_confidence_strains:
                story.append(Spacer(1, 0.02*inch))
                story.append(Paragraph(
                    f"<i>However, {len(low_confidence_strains)} high-consequence detections below confidence threshold were found. See section on Low Confidence, High Consequence Detections.</i>",
                    small_style))
        if args.max_members is not None and args.max_members > 0:
            story.append(Paragraph(
                f"<i>Showing top {args.max_members} strains per group by TASS score</i>",
                small_style))
        story.append(Spacer(1, 0.1*inch))

    # ── Footer: Color Key ─────────────────────────────────────────────────────
    story.append(Spacer(1, 0.15*inch))
    story.append(AnchorFlowable('color_key'))
    story.append(OutlineDest('outline_color_key', 'Color Key', level=0, closed=True, collector=outline_collector))
    story.append(Paragraph("<b>Color Key:</b>", heading_style))
    url = 'https://github.com/jhuapl-bio/taxtriage/blob/main/assets/pathogen_sheet.csv'
    legend_data = [
        ['', Paragraph('Category', legend_text_style), Paragraph('Description', legend_text_style)],
        ['', Paragraph('Primary Pathogen (Direct)', legend_text_style),
         Paragraph(f'The organism is directly detected according to <link href="{url}" color="blue">taxonomic id</link> or assembly name and is listed as being of importance in your sample type.', legend_text_style)],
        ['', Paragraph('Primary Pathogen', legend_text_style),
         Paragraph('It is directly detected according to taxonomic id or assembly name and is listed as being of importance in a different sample type.', legend_text_style)],
        ['', Paragraph('Species/Subkey Summary (inferred)', legend_text_style),
         Paragraph(f'! shown when the species-level aggregation passes the threshold but all individual strains are below it. Dark red is used when any qualifying strain also passes. Rows marked with {_WARN_SYMBOL_HTML} list the highest-scoring Primary strain that is below the defined TASS threshold.', legend_text_style)],
        ['', Paragraph('Commensal', legend_text_style), Paragraph('Normal flora / non-pathogenic', legend_text_style)],
        ['', Paragraph('Opportunistic', legend_text_style), Paragraph('May cause disease in certain conditions', legend_text_style)],
        ['', Paragraph('Potential', legend_text_style), Paragraph('Potential pathogen requiring further investigation', legend_text_style)],
        ['', Paragraph('Unknown', legend_text_style), Paragraph('Classified organism with unknown significance', legend_text_style)],
    ]
    legend_table = Table(legend_data, colWidths=[0.3*inch, 1.6*inch, 3.3*inch])
    legend_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#E85F50')),   # Primary Pathogen (Direct) – crimson red
        ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#E67E22')),   # Primary Pathogen – orange
        ('BACKGROUND', (0, 3), (0, 3), colors.HexColor('#E8A0A0')),   # Species/Subkey Summary – light red
        ('BACKGROUND', (0, 4), (0, 4), colors.lightgreen),
        ('BACKGROUND', (0, 5), (0, 5), colors.HexColor('#ffe6a8')),
        ('BACKGROUND', (0, 6), (0, 6), colors.lightblue),
        ('BACKGROUND', (0, 7), (0, 7), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (1, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (1, 1), (-1, -1), 6),
    ]))
    story.append(legend_table)
    story.append(Spacer(1, 0.04*inch))
    _flora_note_style = ParagraphStyle(
        'FloraNote', parent=legend_text_style,
        fontSize=7, leading=9, textColor=colors.Color(0.35, 0.35, 0.35))
    story.append(Paragraph(
        '<font color="#e67e22"><b>[site flora]</b></font> '
        'Appears next to organism names in sterile samples (blood, CSF, serum) '
        'when the organism is a known commensal at the indicated body site. '
        'Detection in a normally sterile site may suggest contamination.',
        _flora_note_style))
    story.append(Spacer(1, 0.02*inch))

    # ── Footer: Column Explanations ───────────────────────────────────────────
    story.append(AnchorFlowable('column_explanations'))
    story.append(OutlineDest('outline_column_explanations', 'Column Explanations', level=0, closed=True, collector=outline_collector))
    story.append(Paragraph("<b>Column Explanations:</b>", heading_style))
    story.append(Spacer(1, 0.03*inch))
    explanations = [
        "• <b>Specimen ID (Taxonomic ID #):</b> The unique identifier for the sample including its name and taxonomic ID. The taxonomic ID is a link to the NCBI Taxonomy Browser for that organism.",
        "• <b>Detected Organism:</b> The organism detected in the sample, which could be a bacterium, virus, fungus, or parasite.",
        "• <b>TASS Score:</b> A metric between 0 and 100 that reflects the confidence of the organism's detection, with 100 being the highest value.",
        "• <b>K2 Reads:</b> The number of reads classified by Kraken2, a tool for taxonomic classification of sequencing data.",
        "• <b># Reads Aligned:</b> The number of reads from the sequencing data that align to the organism's genome. (%) refers to the percentage of total reads in the sample aligned to that species.",
        "• <b>RPM:</b> Reads Per Million (RPM). This normalized metric allows for comparison of abundance across samples and organisms of different sizes.",
        "• <b>Cov.:</b> The coverage of the organism's genome by aligned reads, expressed as a percentage.",
        "• <b>Species/Subkey Rows:</b> When subkey grouping is enabled (default), each genus group is followed by a species/subkey summary row and then any individual child strains that also pass the TASS cutoff. The species/subkey row color is promoted to the most clinically significant qualifying annotation among that species and its visible child strains.",
        f"• <b>Species/Subkey Warning Symbol:</b> Rows marked with {_WARN_SYMBOL_HTML} have one or more qualifying harmful child strains beneath that species/subkey and should be followed up.",
        "• <b>High ANI Matches:</b> When the High ANI column is shown, this column lists any strains in the dataset that have an Average Nucleotide Identity (ANI) above the specified threshold with the given strain. Each match includes a link to its section in the report if it is present, or just the taxonomic ID and ANI percentage if not present.",
    ]
    for e in explanations:
        story.append(Paragraph(e, metadata_style))
        story.append(Spacer(1, 0.02*inch))
    story.append(Spacer(1, 0.1*inch))

    # ── Footer: Low Confidence ────────────────────────────────────────────────
    if low_confidence_strains:
        story.append(AnchorFlowable('low_confidence'))
        story.append(OutlineDest('outline_low_confidence',
                                 'Low Confidence Detections', level=0, closed=True,
                                 collector=outline_collector))
        story.append(Paragraph("<b>Low Confidence, High Consequence Detections:</b>", heading_style))
        # Build a summary of per-sample thresholds for the description
        _unique_confs = sorted(set(sample_min_conf.values()))
        if len(_unique_confs) == 1:
            _conf_desc = f"the confidence threshold of {_unique_confs[0]}"
        else:
            _conf_desc = "their sample-specific confidence thresholds"
        story.append(Paragraph(
            f"The following strains were detected but fell below {_conf_desc} "
            f"and are listed here for reference.", metadata_style))
        story.append(Spacer(1, 0.05*inch))
        story.append(create_low_confidence_table(
            low_confidence_strains, small_style, show_k2_column, available_width))
        story.append(Spacer(1, 0.1*inch))

    # ── Strain Detail Appendix (when inline table is off) ────────────────────
    if show_strain_appendix:
        appendix_items = create_strain_detail_tables(
            samples_dict, sorted_groups_by_sample,
            small_style, show_k2_column, valid_bookmarks,
            args, available_width, outline_collector=outline_collector,
        )
        story.extend(appendix_items)

    # ── Footer: Additional Information ───────────────────────────────────────
    story.append(AnchorFlowable('additional_info'))
    story.append(OutlineDest('outline_additional_info', 'Additional Information', level=0, closed=True, collector=outline_collector))
    story.append(Paragraph("<b>Additional Information:</b>", heading_style))
    story.append(Spacer(1, 0.01*inch))
    url2 = "https://github.com/jhuapl-bio/taxtriage/blob/main/docs/usage.md#confidence-scoring"
    story.append(Paragraph(
        f'Please visit our <a href="{url2}"><b><font color="blue">DOCUMENTATION PAGE</font></b></a> '
        f'for more information on how confidence is calculated.', metadata_style))
    story.append(Spacer(1, 0.01*inch))
    story.append(Paragraph(
        "The following information highlights the description for the color combinations "
        "for each organism class in the annotated table(s).", metadata_style))
    story.append(Spacer(1, 0.01*inch))
    story.append(Paragraph(
        "Please see the relevant Discovery Analysis txt file for low confidence, "
        "high consequence annotations that were not present in the pdf.", metadata_style))
    story.append(Spacer(1, 0.01*inch))
    story.append(Paragraph(
        "Read amounts are represented as the <b>total number of aligned reads</b> of sufficient "
        "mapping quality <b>(% relative to total reads in sample)</b>.", metadata_style))
    story.append(Spacer(1, 0.01*inch))
    story.append(Paragraph(
        "If there are questions or issues with your report, please open an issue on GitHub as a "
        "discussion <link href=\"https://github.com/jhuapl-bio/taxtriage/discussions\" color=\"blue\">"
        "here</link>. Issues should be tracked/submitted at "
        "<link href=\"https://github.com/jhuapl-bio/taxtriage/issues\" color=\"blue\">this link</link>.",
        metadata_style))

    # Flush all outline entries in alphabetical order as the very last flowable
    story.append(FinalizeOutlines(outline_collector))

    doc.build(story)
    print(f"\nPDF created successfully: {output_path}")


def create_tabular_output(output_path, samples_dict, args):
    """
    Create a tabular output file (CSV/TSV/TXT/XLSX) with strain-level data.
    Includes ALL strains (not filtered). A 'Subkey' column is included.
    """
    file_ext = os.path.splitext(output_path)[1].lower()

    headers = [
        'Index', 'index', 'Detected Organism', 'Specimen ID', 'Sample Type',
        '% Reads', '# Reads Aligned', '% Aligned Reads', 'Coverage',
        'HHS Percentile', 'IsAnnotated', 'AnnClass', 'Microbial Category',
        'High Consequence', 'Taxonomic ID #', 'Status', 'Gini Coefficient',
        'Mean BaseQ', 'Mean MapQ', 'Mean Depth', 'isSpecies',
        'Pathogenic Subsp/Strains', 'K2 Reads', 'RPKM', 'RPM',
        'Parent K2 Reads', 'MapQ Score', 'Disparity Score', 'Minhash Score',
        'Diamond Identity', 'K2 Disparity Score', 'Siblings score',
        'Breadth Weight Score', 'TASS Score', 'MicrobeRT Probability',
        'MicrobeRT Model', 'Reads Aligned', 'Group', 'Subkey',
        'Flora Sites', 'Passes Threshold'
    ]

    all_rows = []
    global_index = 0

    for sample_name in sorted(samples_dict.keys()):
        species_groups = samples_dict[sample_name]

        def _max_member_tass(sg):
            return max((m.get('tass_score', 0) for m in get_all_strains(sg)), default=0)

        if args.sort_alphabetical:
            sorted_groups = sorted(species_groups, key=lambda sg: sg.get('toplevelname', 'Unknown'))
        else:
            sorted_groups = sorted(species_groups, key=_max_member_tass, reverse=True)

        # Use total reads from metadata (includes unaligned) when available;
        # fall back to sum of aligned reads across species groups.
        _tab_smeta = getattr(args, '_input_metadata', {}).get(sample_name, {})
        _tab_meta_total = _tab_smeta.get('total_reads')
        sample_total_reads = max(1, int(_tab_meta_total) if _tab_meta_total else sum(sg.get('numreads', 0) for sg in species_groups))
        # Per-sample confidence threshold (used for Passes Threshold column)
        _mc = getattr(args, '_sample_min_conf', {}).get(sample_name, _DEFAULT_CONF)

        for sg in sorted_groups:
            group_key = sg.get('toplevelkey', sg.get('key', ''))
            sample_type = sg.get('sampletype', 'unknown')
            strains = sorted(get_all_strains(sg),
                             key=lambda s: s.get('tass_score', 0), reverse=True)

            for local_idx, strain in enumerate(strains):
                if not has_min_reads(strain, 1):
                    continue
                strain_reads = float(strain.get('numreads', 0) or 0)
                pct_reads = strain_reads / sample_total_reads * 100.0

                # ── Flora Sites: commensal site tags for sterile sample types ──
                _is_sterile_tab = sample_type.lower().strip() in _STERILE_TYPES
                if _is_sterile_tab:
                    _sites_raw = strain.get('commensal_sites', [])
                    _sites_flat = (
                        [s for s in _sites_raw if isinstance(s, str)]
                        + [item for s in _sites_raw if isinstance(s, list) for item in s]
                    )
                    _flora_sites_str = ', '.join(sorted(set(_sites_flat))) if _sites_flat else ''
                else:
                    _flora_sites_str = ''
                # ── Passes Threshold: does this strain meet the sample's TASS cutoff ──
                _passes_thresh = 'TRUE' if passes_confidence_threshold(strain, _mc) else 'FALSE'

                all_rows.append([
                    global_index, local_idx,
                    strain.get('name', 'Unknown'), sample_name, sample_type,
                    f"{pct_reads:.4f}", int(strain_reads), f"{pct_reads:.4f}",
                    f"{(min(1, strain.get('coverage', 0)) or 0)*100:.0f}%",
                    '100.0',
                    'Yes' if strain.get('isAnnotated', True) else 'No',
                    strain.get('annClass', ''),
                    strain.get('microbial_category', 'Unknown'),
                    'True' if strain.get('high_cons', False) else 'False',
                    strain.get('key', ''), strain.get('status', ''),
                    f"{(strain.get('gini_coefficient', 0) or 0):.2f}",
                    f"{(strain.get('meanbaseq', 0) or 0):.2f}",
                    f"{(strain.get('meanmapq', 0) or 0):.2f}",
                    f"{(strain.get('meandepth', 0) or 0):.1f}",
                    'True' if strain.get('isSpecies', False) else 'False',
                    '',
                    int(strain.get('k2_reads', 0) or 0),
                    strain.get("rpkm", 0) or 0,
                    strain.get("rpm", 0) or 0,
                    int(strain.get('parent_k2_reads', 0) or 0),
                    f"{(strain.get('mapq_score', 0) or 0):.2f}",
                    f"{(strain.get('disparity', 0) or 0):.2f}",
                    f"{(strain.get('minhash_reduction', 0) or 0):.2f}",
                    f"{(strain.get('diamond_identity', 0) or 0):.1f}",
                    f"{(strain.get('k2_disparity_score', 0) or 0):.1f}",
                    f"{(strain.get('siblings_score', 0) or 0):.1f}",
                    f"{(strain.get('breadth_log_score', 0) or 0):.2f}",
                    int((strain.get('tass_score', 0) or 0) * 100),
                    f"{(strain.get('mmbert', 0) or 0):.4f}",
                    strain.get('mmbert_model', '') or '',
                    int(strain_reads),
                    group_key,
                    strain.get('subkey', strain.get('key', '')),
                    _flora_sites_str,
                    _passes_thresh,
                ])
                global_index += 1

    df = pd.DataFrame(all_rows, columns=headers)

    if file_ext == '.csv':
        df.to_csv(output_path, index=False)
        print(f"CSV output created: {output_path}")
    elif file_ext in ['.tsv', '.txt']:
        df.to_csv(output_path, sep='\t', index=False)
        print(f"TSV output created: {output_path}")
    elif file_ext in ['.xlsx', '.xls']:
        df.to_excel(output_path, index=False, engine='openpyxl')
        print(f"Excel output created: {output_path}")
    else:
        df.to_csv(output_path, sep='\t', index=False)
        print(f"TSV output created (default): {output_path}")

    print(f"  Total strains in output: {len(all_rows)}")
    print(f"  Note: Output includes ALL strains (not filtered by category or confidence)")


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Generate PDF report from pathogen discovery data",
        epilog="Example: python create_report.py -i confidences.json -o report.pdf",
    )
    parser.add_argument(
        "-i", "--input", metavar="INPUT", required=True, nargs="+", default=[],
        help="Base pathogen discovery table file(s), JSON format only. Can specify more than one",
    )
    parser.add_argument(
        "-d", "--distributions", metavar="DISTRIBUTIONS", required=False,
        help="TSV file that contains all the distribution information for body sites and organisms",
    )
    parser.add_argument(
        "-a", "--abundance_col", metavar="ABU", required=False, default='% Aligned Reads',
        help="Name of abundance column, default is abundance",
    )
    parser.add_argument(
        "-r", "--min_reads", metavar="READS", required=False, default=1, type=int,
        help="Minimum number of reads required to consider an organism for reporting. Default is 1.",
    )
    parser.add_argument(
        "-c", "--min_conf", metavar="MINCONF", required=False, default=None, type=float,
        help="TASS confidence threshold. If not set, auto-selects based on sample type: "
             "sterile/blood/csf=0.3, stool=0.65, nasal=0.60, skin/wound=0.5, default=0.5.",
    )
    parser.add_argument(
        "-x", "--id_col", metavar="IDCOL", required=False, default="Detected Organism",
        help="Name of id column, default is id",
    )
    parser.add_argument(
        "-p", "--percentile", metavar="PERCENTILE", required=False, type=float, default=0.75,
        help="Only show organisms that are in the top percentile of healthy subjects expected abu",
    )
    parser.add_argument('--sort_alphabetical', action="store_true", required=False,
                        help="Sort groups alphabetically instead of by TASS score")
    parser.add_argument('--enable_matrix', action="store_true", required=False,
                        help="Enable matrix view if available")
    parser.add_argument(
        "--max_members", metavar="MAX_MEMBERS", required=False, type=int, default=None,
        help="Maximum number of top strains (by TASS) to show per group. Default: show all",
    )
    parser.add_argument(
        "--max_toc", metavar="MAX_TOC", required=False, type=int, default=4,
        help="Maximum number of species groups to show in TOC per sample. Default: 4",
    )
    parser.add_argument(
        "--zscore_threshold", metavar="ZSCORE", required=False, type=float, default=2.0,
        help="HMP abundance z-score threshold. Organisms with zscore >= this value "
             "are shown above a visual separator in the primary table as 'elevated abundance'. "
             "Those below are listed underneath. Default: 2.0.",
    )
    parser.add_argument("--show_commensals", action="store_true", required=False,
                        help="Show the commensals table")
    parser.add_argument("--show_unidentified", action="store_true", required=False,
                        help="Show the all organisms now listed as commensal or pathogen")
    parser.add_argument("--show_potentials", action="store_true", required=False,
                        help="Show the potentials table")
    parser.add_argument("--show_opportunistic", action="store_true", required=False,
                        help="Show the opportunistic table")
    parser.add_argument(
        "-m", "--missing_samples", metavar="MISSING", required=False, default=None, nargs="+",
        help="Missing samples if any",
    )
    parser.add_argument(
        "-s", "--sitecol", metavar="SCOL", required=False, default='Sample Type',
        help="Name of site column, default is body_site",
    )
    parser.add_argument(
        "-t", "--type", metavar="TYPE", required=False, default='Detected Organism',
        help="What type of data is being processed. Options: 'Taxonomic ID #' or 'Detected Organism'.",
        choices=['Taxonomic ID #', 'Detected Organism'],
    )
    parser.add_argument(
        "--taxdump", metavar="TAXDUMP", required=False, default=None,
        help="Merge the entries on a specific rank args.rank, importing files from nodes.dmp, names.dmp and potentially merged.dmp",
    )
    parser.add_argument(
        "--rank", metavar="RANK", required=False, default="genus",
        help='IF merging with taxdump, what rank to merge on',
    )
    parser.add_argument(
        "-o", "--output", metavar="OUTPUT", required=True, type=str,
        help="Path of output file (pdf)",
    )
    parser.add_argument(
        "-u", "--output_txt", metavar="OUTPUT_TXT", required=False, type=str,
        help="Path of tabular output file. Format determined by extension: .csv, .tsv, .txt (TSV), or .xlsx",
    )
    parser.add_argument(
        "--output_annot_xlsx", metavar="OUTPUT_ANNOT_XLSX", required=False, type=str,
        help="Path for the protein annotation Excel workbook (.xlsx). "
             "Writes four sheets: Genus Summary, Per-Gene Hits, Sample Overview, AMR Genes.",
    )
    # ── subkey grouping control ──────────────────────────────────────────
    parser.add_argument(
        "--no_subkey",
        action="store_true",
        default=False,
        help=(
            "Disable subkey grouping (default: enabled). "
            "By default, members sharing the same 'subkey' value are collapsed into a "
            "single row. The row's left columns show the best-TASS member's metrics "
            "(TASS, K2 Reads, Reads, RPM, Coverage); the right 3 columns contain a "
            "nested mini-table listing every individual strain with its name, TASS score, "
            "and read count. Pass --no_subkey to revert to the original flat "
            "one-row-per-strain layout."
        ),
    )
    # ── Strain detail table toggle ────────────────────────────────────────────
    parser.add_argument(
        "--show_strains_table",
        action="store_true",
        default=False,
        help=(
            "Show the full strain detail mini-table inline in the right column of the "
            "main report table (default: off). When off (default), strain details are "
            "moved to a separate appendix table placed after the Low Confidence section, "
            "and the main table right column shows a compact strain-count link instead."
        ),
    )

    # ── Control comparison visualisation ──────────────────────────────────────
    parser.add_argument(
        "--show_control_bar",
        action="store_true",
        default=False,
        help="Force display of the control comparison spark-bar column in the "
             "primary table.  When not set, the column is auto-enabled whenever "
             "control_comparison data is present in the input JSON.",
    )
    parser.add_argument(
        "--hide_missing_pos_controls",
        action="store_true",
        default=False,
        help="Suppress the 'missing positive controls' rows at the bottom of "
             "the primary table.  By default, organisms present in the positive "
             "control but absent from the sample are shown as faded rows.",
    )

    parser.add_argument(
        "--annotate_report",
        default=None,
        type=str,
        help="Annotation report TSV (from annotate_report.py) for rendering a "
             "protein annotation summary table below the main organism tables. "
             "If not provided, annotation data embedded in the JSON "
             "(protein_annotations fields from match_paths.py) is used instead.",
    )
    parser.add_argument(
        "--annot_min_pident",
        default=50.0,
        type=float,
        metavar="PIDENT",
        help="Minimum %% identity threshold for the protein annotation plots. "
             "Annotations with pident below this value are excluded. Default: 96.0.",
    )

    return parser.parse_args()


def main():
    args = parse_args()

    taxdump_dict = {}
    names_map = {}
    merged_tax_data = {}

    if args.taxdump:
        if os.path.exists(f"{args.taxdump}/nodes.dmp"):
            taxdump_dict = load_taxdump(f"{args.taxdump}/nodes.dmp")
            print(f"Loaded nodes.dmp: {len(taxdump_dict)} entries")
        if os.path.exists(f"{args.taxdump}/names.dmp"):
            names_map = load_names(f"{args.taxdump}/names.dmp")
            print(f"Loaded names.dmp: {len(names_map)} entries")
        if os.path.exists(f"{args.taxdump}/merged.dmp"):
            merged_tax_data = load_merged(f"{args.taxdump}/merged.dmp")
            print(f"Loaded merged.dmp: {len(merged_tax_data)} entries")

    sample_data, input_metadata = load_json_samples(args.input)
    print(f"Loaded {len(sample_data)} species groups from JSON file(s)")

    # ── Inject protein annotations from standalone TSV if provided ────────
    # This is a fallback for when match_paths.py didn't embed annotations.
    if args.annotate_report and os.path.exists(args.annotate_report):
        print(f"Loading standalone annotation report: {args.annotate_report}")
        if args.annotate_report.endswith('.xlsx') or args.annotate_report.endswith('.xls'):
            _annot_df = pd.read_excel(args.annotate_report, dtype=str)
        else:
            _annot_df = pd.read_csv(args.annotate_report, sep='\t', dtype=str)
        _annot_df.columns = _annot_df.columns.str.strip()
        _annot_by_species = {}
        _annot_cols = [
            'sseqid', 'pident', 'evalue', 'bitscore', 'source_id',
            'gene_name', 'product', 'classification', 'antibiotics_class',
            'antibiotics', 'organism', 'genus', 'species', 'property',
            'source', 'level', 'host_name',
        ]
        _available_cols = [c for c in _annot_cols if c in _annot_df.columns]
        for _, row in _annot_df.iterrows():
            sp_taxid = str(row.get('species_taxid', '')).strip()
            if sp_taxid and sp_taxid != 'nan':
                entry = {c: row.get(c, '') for c in _available_cols}
                _annot_by_species.setdefault(sp_taxid, []).append(entry)

        _injected = 0
        for sg in sample_data:
            for sk_m in sg.get('members', []):
                if sk_m.get('protein_annotations'):
                    continue  # already has data from match_paths
                sk = str(sk_m.get('subkey', sk_m.get('key', '')))
                if sk in _annot_by_species:
                    _seen = set()
                    _unique = []
                    for a in _annot_by_species[sk]:
                        gn = a.get('gene_name', '')
                        if gn not in _seen:
                            _seen.add(gn)
                            _unique.append(a)
                    sk_m['protein_annotations'] = _unique
                    _injected += 1
        print(f"Injected protein annotations into {_injected} species groups from TSV")

    # Build per-sample metadata lookup from all input files
    # Key by sample_name from metadata; later accessible via args._input_metadata
    _per_sample_meta = {}
    for meta in input_metadata:
        sn = meta.get('sample_name')
        if sn:
            _per_sample_meta[sn] = meta
    args._input_metadata = _per_sample_meta
    if _per_sample_meta:
        for sn, m in _per_sample_meta.items():
            print(f"  Metadata for '{sn}': platform={m.get('platform', 'unknown')}, "
                  f"total_reads={m.get('total_reads', '?')}, "
                  f"aligned_reads={m.get('aligned_reads', '?')}, "
                  f"species_groups={m.get('num_species_groups', '?')}")

    samples_dict = organize_data_by_sample(sample_data)
    print(f"Found {len(samples_dict)} unique sample(s)")

    # ── Per-sample min_conf ──────────────────────────────────────────────────
    # Build a dict of per-sample thresholds so each sample can have its own
    # cutoff based on body-site type. If the user explicitly set --min_conf,
    # that value is applied uniformly to all samples.
    sample_min_conf = {}
    sample_conf_source = {}
    for sname, sgroups in samples_dict.items():
        _conf, _src = get_sample_min_conf(
            sname, sgroups, args.min_conf,
            input_metadata=_per_sample_meta,
        )
        sample_min_conf[sname] = _conf
        sample_conf_source[sname] = _src
        print(f"  Sample '{sname}' min_conf = {_conf}  ({_src})")
    # Store on args so downstream functions can access it
    args._sample_min_conf = sample_min_conf
    args._sample_conf_source = sample_conf_source

    print(f"\nConfiguration:")
    print(f"  Output PDF: {args.output}")
    if args.output_txt:
        print(f"  Output TXT: {args.output_txt}")
    if getattr(args, 'output_annot_xlsx', None):
        print(f"  Output Annotation XLSX: {args.output_annot_xlsx}")
    print(f"  Min Confidence: {args.min_conf if args.min_conf is not None else 'auto (per-sample)'}")
    print(f"  Show Potentials: {args.show_potentials}")
    print(f"  Show Unidentified: {args.show_unidentified}")
    print(f"  Show Commensals: {args.show_commensals}")
    print(f"  Subkey grouping: {'DISABLED' if args.no_subkey else 'ENABLED'}")
    print(f"  High ANI matches: read from 'high_ani_matches' field in JSON (set by match_paths.py)")

    create_pdf_template(args.output, samples_dict, args)

    if args.output_txt:
        create_tabular_output(args.output_txt, samples_dict, args)

    if getattr(args, 'output_annot_xlsx', None):
        create_protein_annotation_xlsx(args.output_annot_xlsx, samples_dict, args)

    return taxdump_dict, names_map, merged_tax_data, sample_data


if __name__ == "__main__":
    taxdump_dict, names_map, merged_tax_data, sample_data = main()
